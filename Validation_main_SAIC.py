# -*-coding:utf-8 -*-
import socket
import random
from binascii import a2b_hex
import openpyxl
import os
import time
import select
from serial import *
import subprocess
import binascii
import threading
import ctypes
from ctypes import *
import sys
from functools import reduce
from scapy.all import *
import hashlib
from tkinter import *
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
import tkinter
import numpy as np
import openpyxl.styles
from openpyxl.styles import PatternFill
import wmi
import pywintypes
import win32api
import pythoncom


'''Basic parameters'''
ip_incubator = '172.16.200.107'
ip_tester = '172.16.200.23'
port_incubator = 502
port_tester = random.randint(60000, 65000)
# port_tester = 503
# ser = Serial('COM11', 115200, timeout=0.5)
# channel_Open = [0x00, 0xf9, 0xff]
# channel_Close = [0x00, 0x09, 0xff]

ip_interface_1 = '192.168.2.201'
ip_interface_2 = '192.168.2.202'
ip_interface_3 = '192.168.2.203'
ip_interface_4 = '192.168.2.204'
ip_interface_5 = '192.168.2.205'
ip_interface_6 = '192.168.2.206'

#print(type(ip_interface_6))

def get_network_info(ipaddress):
  c = wmi.WMI()
  for interface in c.Win32_NetworkAdapterConfiguration(IPEnabled=1):
   if interface.IPAddress[0] == ipaddress:
    return interface.Description, interface.MACAddress

(net_if_1, mac_ecu_1) = get_network_info(ip_interface_1)

# (net_if_2, mac_ecu_2) = get_network_info(ip_interface_2)
# (net_if_3, mac_ecu_3) = get_network_info(ip_interface_3)
# (net_if_4, mac_ecu_4) = get_network_info(ip_interface_4)
# (net_if_5, mac_ecu_5) = get_network_info(ip_interface_5)
# (net_if_6, mac_ecu_6) = get_network_info(ip_interface_6)


port_1 = random.randint(60000, 65000)
port_2 = random.randint(60000, 65000)
port_3 = random.randint(60000, 65000)
port_4 = random.randint(60000, 65000)
port_5 = random.randint(60000, 65000)
port_6 = random.randint(60000, 65000)
ip_ecu = '192.168.2.50'
port_ecu = 9000

data_analysed_1 = []
data_analysed_2 = []
data_analysed_3 = []
data_analysed_4 = []
data_analysed_5 = []
data_analysed_6 = []
list_extend = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

'''CAN'''
err_CAN0_Status_1 = 0
err_CAN1_Status_1 = 0
err_CAN2_Status_1 = 0
err_CAN3_Status_1 = 0
err_CAN4_Status_1 = 0
err_CAN5_Status_1 = 0
#
# '''LSD'''
err_LSD1_AD_Status_1 = 0
err_LSD2_AD_Status_1 = 0
err_LSD3_AD_Status_1 = 0
err_LSD4_AD_Status_1 = 0

'''USS_PWR'''
err_USS1_Vout_1 = 0
err_USS2_Vout_1 = 0
err_USS3_Vout_1 = 0

'''ADC'''
err_ADC_VBAT_12V_1 = 0
err_ADC_ADC_VBoost_12V_1 = 0
err_DC_TJA1043_INH_1 = 0
err_ADC_IG_12V_1 = 0
err_ADC_SWITCH_INH_1 = 0
err_ADC_SYS_1V8A_1 = 0
err_ADC_SYS_1V8B_1 = 0
err_ADC_SYS_1V2A_1 = 0
err_ADC_SYS_1V2B_1 = 0
err_ADC_SYS_3V3_1 = 0
err_ADC_J3A_COREPD_0V8_G3_1 = 0
err_ADC_J3A_VDDQDDR_1V1_G1_1 = 0
err_ADC_J3B_COREPD_0V8_G3_1 = 0
err_ADC_J3B_VDDQDDR_1V1_G1_1 = 0
err_ADC_J3C_COREPD_0V8_G3_1 = 0
err_ADC_J3C_VDDQDDR_1V1_G1_1 = 0
err_ADC_GPSANT_PWR_AMUX_1 = 0
err_ADC_GPS_3V3_1 = 0
err_ADC_SYS_5V_1 = 0
err_ADC_ETH_0V9_1 = 0
err_ADC_HWID_1 = 0
err_ADC_Proj_ID_1 = 0

'''J3_Status'''
err_J3A_Alive_Status_1 = 0
err_J3B_Alive_Status_1 = 0
err_J3C_Alive_Status_1 = 0
err_MCU_J3A_SPI_Status_1 = 0
err_MCU_J3B_SPI_Status_1 = 0
err_MCU_J3C_SPI_Status_1 = 0
err_MCU_J3A_Ethernet_TimeOut_Staus_1 = 0
err_MCU_J3B_Ethernet_TimeOut_Staus_1 = 0
err_MCU_J3C_Ethernet_TimeOut_Staus_1 = 0
err_MCU_J3A_Ethernet_RollingCounter_Staus_1 = 0
err_MCU_J3B_Ethernet_RollingCounter_Staus_1 = 0
err_MCU_J3C_Ethernet_RollingCounter_Staus_1 = 0
err_J3A_CPU_Load_1 = 0
err_J3B_CPU_Load_1 = 0
err_J3C_CPU_Load_1 = 0
err_J3A_eMMC_Status_1 = 0
err_J3B_eMMC_Status_1 = 0
err_J3C_eMMC_Status_1 = 0

'''Surround_Camera'''
err_J3C_Surround_Camera_Input0_Link_Lock_1 = 0
err_J3C_Surround_Camera_Input0_Lock_1 = 0
err_J3C_Surround_Camera_Input0_Fps_1 = 0
err_J3C_Surround_Camera_Input0_Crc_1 = 0
err_J3C_Surround_Camera_Input0_OpenLoad_1 = 0
err_J3C_Surround_Camera_Input0_Short_1 = 0
err_J3C_Surround_Camera_Input1_Link_Lock_1 = 0
err_J3C_Surround_Camera_Input1_Lock_1 = 0
err_J3C_Surround_Camera_Input1_Fps_1 = 0
err_J3C_Surround_Camera_Input1_Crc_1 = 0
err_J3C_Surround_Camera_Input1_OpenLoad_1 = 0
err_J3C_Surround_Camera_Input1_Short_1 = 0
err_J3C_Surround_Camera_Input2_Link_Lock_1 = 0
err_J3C_Surround_Camera_Input2_Lock_1 = 0
err_J3C_Surround_Camera_Input2_Fps_1 = 0
err_J3C_Surround_Camera_Input2_Crc_1 = 0
err_J3C_Surround_Camera_Input2_OpenLoad_1 = 0
err_J3C_Surround_Camera_Input2_Short_1 = 0
err_J3C_Surround_Camera_Input3_Link_Lock_1 = 0
err_J3C_Surround_Camera_Input3_Lock_1 = 0
err_J3C_Surround_Camera_Input3_Fps_1 = 0
err_J3C_Surround_Camera_Input3_Crc_1 = 0
err_J3C_Surround_Camera_Input3_OpenLoad_1 = 0
err_J3C_Surround_Camera_Input3_Short_1 = 0
err_J3C_Surround_Camera_Output_Link_Lock_1 = 0
err_J3C_Surround_Camera_Output_Lock_1 = 0
err_J3C_Surround_Camera_Output_Fps_1 = 0
err_J3C_Surround_Camera_Output_Crc_1 = 0

'''FC'''
err_J3A_Front_Camera_Input_Link_Lock_1 = 0
err_J3A_Front_Camera_Input_Lock_1 = 0
err_J3A_Front_Camera_Input_Fps_1 = 0
err_J3A_Front_Camera_Input_Crc_1 = 0
err_J3A_Front_Camera_Input_OpenLoad_1 = 0
err_J3A_Front_Camera_Input_Short_1 = 0
err_J3A_Front_Camera_Output_Link_Lock_1 = 0
err_J3A_Front_Camera_Output_Lock_1 = 0
err_J3A_Front_Camera_Output_Fps_1 = 0
err_J3A_Front_Camera_Output_Crc_1 = 0

'''Rear_Camera'''
err_J3B_Rear_Camera_Input_Link_Lock_1 = 0
err_J3B_Rear_Camera_Input_Lock_1 = 0
err_J3B_Rear_Camera_Input_Fps_1 = 0
err_J3B_Rear_Camera_Input_Crc_1 = 0
err_J3B_Rear_Camera_Input_OpenLoad_1 = 0
err_J3B_Rear_Camera_Input_Short_1 = 0

'''Side_Camera'''
err_J3B_Side_Camera_Input0_Link_Lock_1 = 0
err_J3B_Side_Camera_Input0_Lock_1 = 0
err_J3B_Side_Camera_Input0_Fps_1 = 0
err_J3B_Side_Camera_Input0_Crc_1 = 0
err_J3B_Side_Camera_Input0_OpenLoad_1 = 0
err_J3B_Side_Camera_Input0_Short_1 = 0
err_J3B_Side_Camera_Input1_Link_Lock_1 = 0
err_J3B_Side_Camera_Input1_Lock_1 = 0
err_J3B_Side_Camera_Input1_Fps_1 = 0
err_J3B_Side_Camera_Input1_Crc_1 = 0
err_J3B_Side_Camera_Input1_OpenLoad_1 = 0
err_J3B_Side_Camera_Input1_Short_1 = 0
err_J3B_Side_Camera_Input2_Link_Lock_1 = 0
err_J3B_Side_Camera_Input2_Lock_1 = 0
err_J3B_Side_Camera_Input2_Fps_1 = 0
err_J3B_Side_Camera_Input2_Crc_1 = 0
err_J3B_Side_Camera_Input2_OpenLoad_1 = 0
err_J3B_Side_Camera_Input2_Short_1 = 0
err_J3B_Side_Camera_Input3_Link_Lock_1 = 0
err_J3B_Side_Camera_Input3_Lock_1 = 0
err_J3B_Side_Camera_Input3_Fps_1 = 0
err_J3B_Side_Camera_Input3_Crc_1 = 0
err_J3B_Side_Camera_Input3_OpenLoad_1 = 0
err_J3B_Side_Camera_Input3_Short_1 = 0

'''GNSS'''
err_J3A_F9K_GNSS_Status_1 = 0
err_MCU_F9K_GNSS_OpenLoad_1 = 0
err_MCU_F9K_GNSS_Short_1 = 0
err_MCU_F9K_GNSS_Supply_1 = 0
err_J3A_F9K_IMU_Z_1 = 0
# err_J3A_IAM_20685_IMU_Z_1 = 0
err_J3A_F9K_CN_Value_1 = 0

'''TEMP'''
err_MCU_TEMP_1 = 0
err_J3A_5024_TEMP_1 = 0
err_J3B_5024_TEMP_1 = 0
err_J3C_5024_TEMP_1 = 0
err_PCB_TEMP_1 = 0

'''J3A'''
err_J3A_CPU_0V8_G3_1 = 0
err_J3A_DDR_0V8_G3_1 = 0
err_J3A_VDD_0V8_G4_1 = 0
err_J3A_COREAO_0V8_G2_1 = 0
err_J3A_CNN0_0V8_G3_1 = 0
err_J3A_CNN1_0V8_G3_1 = 0
err_J3A_VDD_1V8_G1_1 = 0
err_J3A_EMMC_3V3_G1_1 = 0
err_J3A_VDD_3V3_G1_1 = 0
err_J3A_VCC_3V3_G4_1 = 0
err_J3A_VCC_1V8_G4_1 = 0
err_J3A_TEMP_IC_1 = 0
err_J3A_TEMP_SW1_1 = 0
err_J3A_TEMP_SW2_1 = 0
err_J3A_TEMP_SW3_1 = 0
err_J3A_TEMP_SW4_1 = 0
err_J3A_TEMP_SW5_1 = 0
err_J3A_TEMP_SW6_1 = 0
err_J3A_TEMP_SW7_1 = 0
err_J3A_TEMP_TEMP_LDO1_2_1 = 0
err_J3A_TEMP_TEMP_LDO3_4_1 = 0

'''J3B'''
err_J3B_CPU_0V8_G3_1 = 0
err_J3B_DDR_0V8_G3_1 = 0
err_J3B_VDD_0V8_G4_1 = 0
err_J3B_COREAO_0V8_G2_1 = 0
err_J3B_CNN0_0V8_G3_1 = 0
err_J3B_CNN1_0V8_G3_1 = 0
err_J3B_VDD_1V8_G1_1 = 0
err_J3B_EMMC_3V3_G1_1 = 0
err_J3B_VDD_3V3_G1_1 = 0
err_J3B_VCC_3V3_G4_1 = 0
err_J3B_VCC_1V8_G4_1 = 0
err_J3B_TEMP_IC_1 = 0
err_J3B_TEMP_SW1_1 = 0
err_J3B_TEMP_SW2_1 = 0
err_J3B_TEMP_SW3_1 = 0
err_J3B_TEMP_SW4_1 = 0
err_J3B_TEMP_SW5_1 = 0
err_J3B_TEMP_SW6_1 = 0
err_J3B_TEMP_SW7_1 = 0
err_J3B_TEMP_TEMP_LDO1_2_1 = 0
err_J3B_TEMP_TEMP_LDO3_4_1 = 0

'''J3C'''
err_J3C_CPU_0V8_G3_1 = 0
err_J3C_DDR_0V8_G3_1 = 0
err_J3C_VDD_0V8_G4_1 = 0
err_J3C_COREAO_0V8_G2_1 = 0
err_J3C_CNN0_0V8_G3_1 = 0
err_J3C_CNN1_0V8_G3_1 = 0
err_J3C_VDD_1V8_G1_1 = 0
err_J3C_EMMC_3V3_G1_1 = 0
err_J3C_VDD_3V3_G1_1 = 0
err_J3C_VCC_3V3_G4_1 = 0
err_J3C_VCC_1V8_G4_1 = 0
err_J3C_TEMP_IC_1 = 0
err_J3C_TEMP_SW1_1 = 0
err_J3C_TEMP_SW2_1 = 0
err_J3C_TEMP_SW3_1 = 0
err_J3C_TEMP_SW4_1 = 0
err_J3C_TEMP_SW5_1 = 0
err_J3C_TEMP_SW6_1 = 0
err_J3C_TEMP_SW7_1 = 0
err_J3C_TEMP_TEMP_LDO1_2_1 = 0
err_J3C_TEMP_TEMP_LDO3_4_1 = 0

'''I2C'''
err_MCU_MAX20084_I2C_Status_1 = 0
err_MCU_PMIC8100_Index0_I2C_Status_1 = 0
err_MCU_PMIC8100_Index1_I2C_Status_1 = 0
err_MCU_PMIC8100_Index2_I2C_Status_1 = 0
err_MCU_PMIC5024_Index0_I2C_Status_1 = 0
err_MCU_PMIC5024_Index1_I2C_Status_1 = 0
err_MCU_PMIC5024_Index2_I2C_Status_1 = 0
err_J3A_MAX9296_I2C_Status_1 = 0
err_J3A_MAX96717_I2C_Status_1 = 0
err_J3A_MAX20089_I2C_Status_1 = 0
err_J3B_MAX9296_I2C_Status_1 = 0
err_J3B_MAX96712_I2C_Status_1 = 0
err_J3B_MAX20089_I2C_Status_1 = 0
err_J3B_MAX20087_I2C_Status_1 = 0
err_J3C_MAX96712_I2C_Status_1 = 0
err_J3C_MAX96717_I2C_Status_1 = 0
err_J3C_MAX20087_I2C_Status_1 = 0
'''I2C add'''
err_J3A_Front_Camera_max9295_ID_1 = 0
err_J3B_Rear_Camera_max9295_ID_1 = 0
err_J3B_Side_Camera_0_max9295_ID_1 = 0
err_J3B_Side_Camera_1_max9295_ID_1 = 0
err_J3B_Side_Camera_2_max9295_ID_1 = 0
err_J3B_Side_Camera_3_max9295_ID_1 = 0
err_J3C_Surround_Camera_0_max9295_ID_1 = 0
err_J3C_Surround_Camera_1_max9295_ID_1 = 0
err_J3C_Surround_Camera_2_max9295_ID_1 = 0
err_J3C_Surround_Camera_3_max9295_ID_1 = 0


'''TEMP'''
err_J3A_TEMP_1 = 0
err_J3B_TEMP_1 = 0
err_J3C_TEMP_1 = 0
err_J3A_F9K_TEMP_1 = 0

'''ADC add'''
err_ADC_SW1_3V3_1 = 0
err_ADC_SW1_1V8_1 = 0
err_ADC_SW1_1V2_1 = 0
err_ADC_J3A_VDD_0V8A_G4_1 = 0
err_ADC_J3B_VDD_0V8A_G4_1 = 0
err_ADC_J3C_VDD_0V8A_G4_1 = 0


'''USS add'''
err_USS1_PWM_PERIOD_1 = 0
err_USS1_PWM_DUTY_1 = 0
err_USS2_PWM_PERIOD_1 = 0
err_USS2_PWM_DUTY_1 = 0
err_USS3_PWM_PERIOD_1 = 0
err_USS3_PWM_DUTY_1 = 0
err_USS4_PWM_PERIOD_1 = 0
err_USS4_PWM_DUTY_1 = 0
err_USS5_PWM_PERIOD_1 = 0
err_USS5_PWM_DUTY_1 = 0
err_USS6_PWM_PERIOD_1 = 0
err_USS6_PWM_DUTY_1 = 0
err_USS7_PWM_PERIOD_1 = 0
err_USS7_PWM_DUTY_1 = 0
err_USS8_PWM_PERIOD_1 = 0
err_USS8_PWM_DUTY_1 = 0
err_USS9_PWM_PERIOD_1 = 0
err_USS9_PWM_DUTY_1 = 0
err_USS10_PWM_PERIOD_1 = 0
err_USS10_PWM_DUTY_1 = 0
err_USS11_PWM_PERIOD_1 = 0
err_USS11_PWM_DUTY_1 = 0
err_USS12_PWM_PERIOD_1 = 0
err_USS12_PWM_DUTY_1 = 0


'''Camera_Supply'''
err_J3A_Front_Camera_Supply_AD_1 = 0
err_J3B_Rear_Camera_Supply_AD_1 = 0
err_J3B_Side_Camera_0_Supply_AD_1 = 0
err_J3B_Side_Camera_1_Supply_AD_1 = 0
err_J3B_Side_Camera_2_Supply_AD_1 = 0
err_J3B_Side_Camera_3_Supply_AD_1 = 0
err_J3C_Surround_Camera_0_Supply_AD_1 = 0
err_J3C_Surround_Camera_1_Supply_AD_1 = 0
err_J3C_Surround_Camera_2_Supply_AD_1 = 0
err_J3C_Surround_Camera_3_Supply_AD_1 = 0


'''Basic functions'''
def rev_msg(condition, buff=1024):
    ready = select.select([tcp_sk], [], [], 1)
    i = condition
    if ready[0]:
        msg = list(tcp_sk.recv(buff))
        if msg[0] == 0x0 and msg[1] == 0x0 and msg[2] == 0x0 and msg[3] == 0x0 and msg[7] == i:
            return msg[9:]
        else:
            return rev_msg(i)
    else:
        return rev_msg(i)

def ping(source_ip, destination_ip):
    command = ['ping', '-n', '2', '-S', source_ip, destination_ip]
    return subprocess.call(command) == 0

def str2int(s):
    return reduce(lambda x, y: x * 16 + y, map(lambda s: {'0': 0, '1': 1, '2': 2,\
                                                          '3': 3, '4': 4, '5': 5, \
                                                          '6': 6, '7': 7,  '8': 8,\
                                                          '9': 9, 'a': 10, 'b': 11,\
                                                          'c': 12, 'd': 13, 'e': 14,\
                                                          'f': 15}[s], s))
def float_cvt(float_list):
    data_bytes = np.array(float_list, dtype=np.uint8)
    data_as_float = data_bytes.view(dtype=np.float32)
    return data_as_float

def packet_rev_udp_1(response):
    global data_analysed_1, list_extend, err_CAN0_Status_1, err_CAN1_Status_1, err_CAN2_Status_1, err_CAN3_Status_1, \
        err_CAN5_Status_1,err_LSD1_AD_Status_1,err_LSD2_AD_Status_1,err_LSD3_AD_Status_1,err_LSD4_AD_Status_1,\
        err_CAN4_Status_1, err_CAN5_Status_1, err_LSD1_AD_Status_1, err_LSD2_AD_Status_1, err_LSD3_AD_Status_1, err_LSD4_AD_Status_1,\
        err_USS1_Vout_1, err_USS2_Vout_1, err_USS3_Vout_1,err_ADC_VBAT_12V_1, err_ADC_ADC_VBoost_12V_1, err_DC_TJA1043_INH_1,\
        err_ADC_IG_12V_1,err_ADC_SWITCH_INH_1,err_ADC_SYS_1V8A_1,err_ADC_SYS_1V8B_1,err_ADC_SYS_1V2A_1,err_ADC_SYS_1V2B_1,\
        err_ADC_SYS_3V3_1,err_ADC_J3A_COREPD_0V8_G3_1,err_ADC_J3A_VDDQDDR_1V1_G1_1,err_ADC_J3B_COREPD_0V8_G3_1,err_ADC_J3B_VDDQDDR_1V1_G1_1,\
        err_ADC_J3C_COREPD_0V8_G3_1,err_ADC_J3C_VDDQDDR_1V1_G1_1,err_ADC_GPSANT_PWR_AMUX_1,err_ADC_GPS_3V3_1,err_ADC_SYS_5V_1,\
        err_ADC_ETH_0V9_1,err_ADC_HWID_1,err_ADC_Proj_ID_1,err_J3A_Alive_Status_1,err_J3B_Alive_Status_1,err_J3C_Alive_Status_1,\
        err_MCU_J3A_SPI_Status_1,err_MCU_J3B_SPI_Status_1,err_MCU_J3C_SPI_Status_1,err_MCU_J3A_Ethernet_TimeOut_Staus_1,\
        err_MCU_J3B_Ethernet_TimeOut_Staus_1,err_MCU_J3C_Ethernet_TimeOut_Staus_1,err_MCU_J3A_Ethernet_RollingCounter_Staus_1,\
        err_MCU_J3B_Ethernet_RollingCounter_Staus_1,err_MCU_J3C_Ethernet_RollingCounter_Staus_1,err_J3A_CPU_Load_1,err_J3B_CPU_Load_1,\
        err_J3C_CPU_Load_1,err_J3A_eMMC_Status_1,err_J3B_eMMC_Status_1,err_J3C_eMMC_Status_1,err_J3C_Surround_Camera_Input0_Link_Lock_1,\
        err_J3C_Surround_Camera_Input0_Lock_1,err_J3C_Surround_Camera_Input0_Fps_1,err_J3C_Surround_Camera_Input0_Crc_1,\
        err_J3C_Surround_Camera_Input0_OpenLoad_1,err_J3C_Surround_Camera_Input0_Short_1,err_J3C_Surround_Camera_Input1_Link_Lock_1,\
        err_J3C_Surround_Camera_Input1_Lock_1,err_J3C_Surround_Camera_Input1_Fps_1,err_J3C_Surround_Camera_Input1_Crc_1,\
        err_J3C_Surround_Camera_Input1_OpenLoad_1,err_J3C_Surround_Camera_Input1_Short_1,err_J3C_Surround_Camera_Input2_Link_Lock_1,\
        err_J3C_Surround_Camera_Input2_Lock_1,err_J3C_Surround_Camera_Input2_Fps_1,err_J3C_Surround_Camera_Input2_Crc_1,\
        err_J3C_Surround_Camera_Input2_OpenLoad_1,err_J3C_Surround_Camera_Input2_Short_1,err_J3C_Surround_Camera_Input3_Link_Lock_1,\
        err_J3C_Surround_Camera_Input3_Lock_1,err_J3C_Surround_Camera_Input3_Fps_1,err_J3C_Surround_Camera_Input3_Crc_1,\
        err_J3C_Surround_Camera_Input3_OpenLoad_1,err_J3C_Surround_Camera_Input3_Short_1,err_J3C_Surround_Camera_Output_Link_Lock_1,\
        err_J3C_Surround_Camera_Output_Lock_1,err_J3C_Surround_Camera_Output_Fps_1,err_J3C_Surround_Camera_Output_Crc_1,\
        err_J3A_Front_Camera_Input_Link_Lock_1,err_J3A_Front_Camera_Input_Lock_1,err_J3A_Front_Camera_Input_Fps_1,err_J3A_Front_Camera_Input_Crc_1,\
        err_J3A_Front_Camera_Input_OpenLoad_1,err_J3A_Front_Camera_Input_Short_1,err_J3A_Front_Camera_Output_Link_Lock_1,err_J3A_Front_Camera_Output_Lock_1,\
        err_J3A_Front_Camera_Output_Fps_1,err_J3A_Front_Camera_Output_Crc_1,err_J3B_Rear_Camera_Input_Link_Lock_1,err_J3B_Rear_Camera_Input_Lock_1,\
        err_J3B_Rear_Camera_Input_Fps_1,err_J3B_Rear_Camera_Input_Crc_1,err_J3B_Rear_Camera_Input_OpenLoad_1,err_J3B_Rear_Camera_Input_Short_1,\
        err_J3B_Side_Camera_Input0_Link_Lock_1,err_J3B_Side_Camera_Input0_Lock_1,err_J3B_Side_Camera_Input0_Fps_1,err_J3B_Side_Camera_Input0_Crc_1,\
        err_J3B_Side_Camera_Input0_OpenLoad_1,err_J3B_Side_Camera_Input0_Short_1,err_J3B_Side_Camera_Input1_Link_Lock_1,err_J3B_Side_Camera_Input1_Lock_1,\
        err_J3B_Side_Camera_Input1_Fps_1,err_J3B_Side_Camera_Input1_Crc_1,err_J3B_Side_Camera_Input1_OpenLoad_1,err_J3B_Side_Camera_Input1_Short_1,\
        err_J3B_Side_Camera_Input2_Link_Lock_1,err_J3B_Side_Camera_Input2_Lock_1,err_J3B_Side_Camera_Input2_Fps_1,err_J3B_Side_Camera_Input2_Crc_1,\
        err_J3B_Side_Camera_Input2_OpenLoad_1,err_J3B_Side_Camera_Input2_Short_1,err_J3B_Side_Camera_Input3_Link_Lock_1,err_J3B_Side_Camera_Input3_Lock_1,\
        err_J3B_Side_Camera_Input3_Fps_1,err_J3B_Side_Camera_Input3_Crc_1,err_J3B_Side_Camera_Input3_OpenLoad_1,err_J3B_Side_Camera_Input3_Short_1,\
        err_J3A_F9K_GNSS_Status_1,err_MCU_F9K_GNSS_OpenLoad_1,err_MCU_F9K_GNSS_Short_1,err_MCU_F9K_GNSS_Supply_1,err_J3A_F9K_IMU_Z_1,err_J3A_IAM_20685_IMU_Z_1,\
        err_J3A_F9K_CN_Value_1,err_MCU_TEMP_1,err_J3A_5024_TEMP_1,err_J3B_5024_TEMP_1,err_J3C_5024_TEMP_1,err_PCB_TEMP_1, err_J3A_CPU_0V8_G3_1,err_J3A_DDR_0V8_G3_1,\
        err_J3A_VDD_0V8_G4_1,err_J3A_COREAO_0V8_G2_1,err_J3A_CNN0_0V8_G3_1,err_J3A_CNN1_0V8_G3_1,err_J3A_VDD_1V8_G1_1,err_J3A_EMMC_3V3_G1_1,err_J3A_VDD_3V3_G1_1,\
        err_J3A_VCC_3V3_G4_1,err_J3A_VCC_1V8_G4_1,err_J3A_TEMP_IC_1,err_J3A_TEMP_SW1_1,err_J3A_TEMP_SW2_1,err_J3A_TEMP_SW3_1,err_J3A_TEMP_SW4_1,err_J3A_TEMP_SW5_1,\
        err_J3A_TEMP_SW6_1,err_J3A_TEMP_SW7_1,err_J3A_TEMP_TEMP_LDO1_2_1,err_J3A_TEMP_TEMP_LDO3_4_1,err_J3B_CPU_0V8_G3_1,err_J3B_DDR_0V8_G3_1,err_J3B_VDD_0V8_G4_1,\
        err_J3B_COREAO_0V8_G2_1,err_J3B_CNN0_0V8_G3_1,err_J3B_CNN1_0V8_G3_1,err_J3B_VDD_1V8_G1_1,err_J3B_EMMC_3V3_G1_1,err_J3B_VDD_3V3_G1_1,err_J3B_VCC_3V3_G4_1,\
        err_J3B_VCC_1V8_G4_1,err_J3B_TEMP_IC_1,err_J3B_TEMP_SW1_1,err_J3B_TEMP_SW2_1,err_J3B_TEMP_SW3_1,err_J3B_TEMP_SW4_1,err_J3B_TEMP_SW5_1,err_J3B_TEMP_SW6_1,\
        err_J3B_TEMP_SW7_1,err_J3B_TEMP_TEMP_LDO1_2_1,err_J3B_TEMP_TEMP_LDO3_4_1,err_J3C_CPU_0V8_G3_1,err_J3C_DDR_0V8_G3_1,err_J3C_VDD_0V8_G4_1,err_J3C_COREAO_0V8_G2_1,\
        err_J3C_CNN0_0V8_G3_1,err_J3C_CNN1_0V8_G3_1,err_J3C_VDD_1V8_G1_1,err_J3C_EMMC_3V3_G1_1,err_J3C_VDD_3V3_G1_1,err_J3C_VCC_3V3_G4_1,err_J3C_VCC_1V8_G4_1,err_J3C_TEMP_IC_1,\
        err_J3C_TEMP_SW1_1,err_J3C_TEMP_SW2_1,err_J3C_TEMP_SW3_1,err_J3C_TEMP_SW4_1,err_J3C_TEMP_SW5_1,err_J3C_TEMP_SW6_1,err_J3C_TEMP_SW7_1,err_J3C_TEMP_TEMP_LDO1_2_1,\
        err_J3C_TEMP_TEMP_LDO3_4_1,err_MCU_MAX20084_I2C_Status_1,err_MCU_PMIC8100_Index0_I2C_Status_1,err_MCU_PMIC8100_Index1_I2C_Status_1,err_MCU_PMIC8100_Index2_I2C_Status_1,\
        err_MCU_PMIC5024_Index0_I2C_Status_1,err_MCU_PMIC5024_Index1_I2C_Status_1,err_MCU_PMIC5024_Index2_I2C_Status_1,err_J3A_MAX9296_I2C_Status_1,err_J3A_MAX96717_I2C_Status_1,\
        err_J3A_MAX20089_I2C_Status_1,err_J3B_MAX9296_I2C_Status_1,err_J3B_MAX96712_I2C_Status_1,err_J3B_MAX20089_I2C_Status_1,err_J3B_MAX20087_I2C_Status_1,\
        err_J3C_MAX96712_I2C_Status_1,err_J3C_MAX96717_I2C_Status_1,err_J3C_MAX20087_I2C_Status_1, err_J3A_TEMP_1, err_J3B_TEMP_1, err_J3C_TEMP_1, err_J3A_F9K_TEMP_1,\
        err_J3A_Front_Camera_Supply_AD_1,err_J3B_Rear_Camera_Supply_AD_1, J3B_Side_Camera_0_Supply_AD_1, err_J3B_Side_Camera_1_Supply_AD_1, err_J3B_Side_Camera_2_Supply_AD_1, \
        err_J3B_Side_Camera_3_Supply_AD_1, err_J3C_Surround_Camera_0_Supply_AD_1, err_J3C_Surround_Camera_1_Supply_AD_1, err_J3C_Surround_Camera_2_Supply_AD_1, err_J3C_Surround_Camera_3_Supply_AD_1
    global err_J3A_Front_Camera_max9295_ID_1, err_J3B_Rear_Camera_max9295_ID_1, err_J3B_Side_Camera_0_max9295_ID_1,err_J3B_Side_Camera_1_max9295_ID_1,\
        err_J3B_Side_Camera_2_max9295_ID_1,err_J3B_Side_Camera_3_max9295_ID_1,err_J3C_Surround_Camera_0_max9295_ID_1,err_J3C_Surround_Camera_1_max9295_ID_1,\
        err_J3C_Surround_Camera_2_max9295_ID_1,err_J3C_Surround_Camera_3_max9295_ID_1, err_ADC_SW1_3V3_1, err_ADC_SW1_1V8_1, err_ADC_SW1_1V2_1, \
        err_ADC_J3A_VDD_0V8A_G4_1, err_ADC_J3B_VDD_0V8A_G4_1, err_ADC_J3C_VDD_0V8A_G4_1, \
        err_USS1_PWM_PERIOD_1, err_USS1_PWM_DUTY_1, err_USS2_PWM_PERIOD_1, err_USS2_PWM_DUTY_1, err_USS3_PWM_PERIOD_1, err_USS3_PWM_DUTY_1,\
        err_USS4_PWM_PERIOD_1, err_USS4_PWM_DUTY_1, err_USS5_PWM_PERIOD_1, err_USS5_PWM_DUTY_1, err_USS6_PWM_PERIOD_1, err_USS6_PWM_DUTY_1,\
        err_USS7_PWM_PERIOD_1, err_USS7_PWM_DUTY_1, err_USS8_PWM_PERIOD_1, err_USS8_PWM_DUTY_1, err_USS9_PWM_PERIOD_1, err_USS9_PWM_DUTY_1,\
        err_USS10_PWM_PERIOD_1, err_USS10_PWM_DUTY_1, err_USS11_PWM_PERIOD_1, err_USS11_PWM_DUTY_1, err_USS12_PWM_PERIOD_1, err_USS12_PWM_DUTY_1
    reply_payload = response[UDP].payload
    reply_bytes = binascii.hexlify(bytes(reply_payload))
    reply_str = reply_bytes.decode()
    num = 0
    reply_list = []
    while num < len(reply_str):
        ab = reply_str[num:num + 2]
        reply_list.append(str2int(ab))
        num += 2
    reply_extend = reply_list + list_extend

    # if reply_extend[0] == 0x55 and reply_extend[1] == 0xAA and  reply_extend[2] == 0x55:
    if reply_extend:
        '''Rolling Counter'''
        exchange_list = [0, 0, 0]
        exchange_list[0] = reply_extend[9]
        data_analysed_1.append(exchange_list)       # data_analysed_1[0]

        '''CAN0_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[10] == 0:
            exchange_list[0] = reply_extend[10]   # value
            exchange_list[1] = 0                  # correct:0    error:1
            exchange_list[2] = err_CAN0_Status_1       # error count
        else:
            err_CAN0_Status_1 += 1
            exchange_list[0] = reply_extend[10]
            exchange_list[1] = 1
            exchange_list[2] = err_CAN0_Status_1
        data_analysed_1.append(exchange_list)       # data_analysed_1[1]

        '''CAN1_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[11] == 0:
            exchange_list[0] = reply_extend[11]  # value
            exchange_list[1] = 0  # correct:0    error:1
            exchange_list[2] = err_CAN1_Status_1  # error count
        else:
            err_CAN1_Status_1 += 1
            exchange_list[0] = reply_extend[11]
            exchange_list[1] = 1
            exchange_list[2] = err_CAN1_Status_1
        data_analysed_1.append(exchange_list)       # data_analysed_1[2]

        '''CAN2_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[12] == 0:
            exchange_list[0] = reply_extend[12]
            exchange_list[1] = 0
            exchange_list[2] = err_CAN2_Status_1
        else:
            err_CAN2_Status_1 += 1
            exchange_list[0] = reply_extend[12]
            exchange_list[1] = 1
            exchange_list[2] = err_CAN2_Status_1
        data_analysed_1.append(exchange_list)       # data_analysed_1[3]

        '''CAN3_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[13] == 0:
            exchange_list[0] = reply_extend[13]
            exchange_list[1] = 0
            exchange_list[2] = err_CAN3_Status_1
        else:
            err_CAN3_Status_1 += 1
            exchange_list[0] = reply_extend[13]
            exchange_list[1] = 1
            exchange_list[2] = err_CAN3_Status_1
        data_analysed_1.append(exchange_list)       # data_analysed_1[4]

        '''CAN4_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[14] == 0:
            exchange_list[0] = reply_extend[14]
            exchange_list[1] = 0
            exchange_list[2] = err_CAN4_Status_1
        else:
            err_CAN4_Status_1 += 1
            exchange_list[0] = reply_extend[14]
            exchange_list[1] = 1
            exchange_list[2] = err_CAN4_Status_1
        data_analysed_1.append(exchange_list)       # data_analysed_1[5]

        '''CAN5_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[15] == 0:
            exchange_list[0] = reply_extend[15]
            exchange_list[1] = 0
            exchange_list[2] = err_CAN5_Status_1
        else:
            err_CAN5_Status_1 += 1
            exchange_list[0] = reply_extend[15]
            exchange_list[1] = 1
            exchange_list[2] = err_CAN5_Status_1
        data_analysed_1.append(exchange_list)       # data_analysed_1[6]

        '''LSD1_AD_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[16] == 0:
            exchange_list[0] = reply_extend[16]
            exchange_list[1] = 0
            exchange_list[2] = err_LSD1_AD_Status_1
        else:
            err_LSD1_AD_Status_1 += 1
            exchange_list[0] = reply_extend[16]
            exchange_list[1] = 1
            exchange_list[2] = err_LSD1_AD_Status_1
        data_analysed_1.append(exchange_list)       # data_analysed_1[7]

        '''LSD2_AD_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[17] == 0:
            exchange_list[0] = reply_extend[17]
            exchange_list[1] = 0
            exchange_list[2] = err_LSD2_AD_Status_1
        else:
            err_LSD2_AD_Status_1 += 1
            exchange_list[0] = reply_extend[17]
            exchange_list[1] = 1
            exchange_list[2] = err_LSD2_AD_Status_1
        data_analysed_1.append(exchange_list)       # data_analysed_1[8]

        '''LSD3_AD_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[18] == 0:
            exchange_list[0] = reply_extend[18]
            exchange_list[1] = 0
            exchange_list[2] = err_LSD3_AD_Status_1
        else:
            err_LSD3_AD_Status_1 += 1
            exchange_list[0] = reply_extend[18]
            exchange_list[1] = 1
            exchange_list[2] = err_LSD3_AD_Status_1
        data_analysed_1.append(exchange_list)       # data_analysed_1[9]

        '''LSD4_AD_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[19] == 0:
            exchange_list = [reply_extend[19], 0, err_LSD4_AD_Status_1]
        else:
            err_LSD4_AD_Status_1 += 1
            exchange_list = [reply_extend[19], 1, err_LSD4_AD_Status_1]
        data_analysed_1.append(exchange_list)       # data_analysed_1[10]

        '''USS1_Vout'''
        float_list = [reply_extend[20],reply_extend[21],reply_extend[22],reply_extend[23]]
        float_USS1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 11.64 <= float_USS1 <= 12.36:
            exchange_list = [float_USS1, 0, err_USS1_Vout_1]
        else:
            err_USS1_Vout_1 += 1
            exchange_list = [float_USS1, 1, err_USS1_Vout_1]
        data_analysed_1.append(exchange_list)       # data_analysed_1[11]

        '''USS2_Vout'''
        float_list = [reply_extend[24], reply_extend[25], reply_extend[26], reply_extend[27]]
        float_USS2 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 11.64 <= float_USS2 <= 12.36:
            exchange_list = [float_USS2, 0, err_USS2_Vout_1]
        else:
            err_USS2_Vout_1 += 1
            exchange_list = [float_USS2, 1, err_USS2_Vout_1]
        data_analysed_1.append(exchange_list)     # data_analysed_1[12]

        '''USS3_Vout'''
        float_list = [reply_extend[28], reply_extend[29], reply_extend[30], reply_extend[31]]
        float_USS3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 11.64 <= float_USS3 <= 12.36:
            exchange_list = [float_USS3, 0, err_USS3_Vout_1]
        else:
            err_USS3_Vout_1 += 1
            exchange_list = [float_USS3, 1, err_USS3_Vout_1]
        data_analysed_1.append(exchange_list)    # data_analysed_1[13]

        '''ADC_VBAT_12V'''
        float_list = [reply_extend[56], reply_extend[57], reply_extend[58], reply_extend[59]]
        ADC_VBAT_12V = float_cvt(float_list)
        exchange_list = [ADC_VBAT_12V, 0, 0]
        data_analysed_1.append(exchange_list)           # data_analysed_1[14]

        '''ADC_VBoost_12V'''
        float_list = [reply_extend[60], reply_extend[61], reply_extend[62], reply_extend[63]]
        ADC_VBoost_12V = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 11.075 <= ADC_VBoost_12V <= 12.980:
            exchange_list = [ADC_VBoost_12V, 0, err_ADC_ADC_VBoost_12V_1]
        else:
            err_ADC_ADC_VBoost_12V_1 += 1
            exchange_list = [ADC_VBoost_12V, 1, err_ADC_ADC_VBoost_12V_1]
        data_analysed_1.append(exchange_list)           # data_analysed_1[15]

        '''ADC_TJA1043_INH'''
        float_list = [reply_extend[64], reply_extend[65], reply_extend[66], reply_extend[67]]
        ADC_TJA1043_INH = float_cvt(float_list)
        exchange_list = [ADC_TJA1043_INH, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[16]

        '''ADC_IG_12V'''
        float_list = [reply_extend[68], reply_extend[69], reply_extend[70], reply_extend[71]]
        ADC_IG_12V = float_cvt(float_list)
        exchange_list = [ADC_IG_12V, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[17]

        '''ADC_SWITCH_INH'''
        float_list = [reply_extend[72], reply_extend[73], reply_extend[74], reply_extend[75]]
        ADC_SWITCH_INH = float_cvt(float_list)
        exchange_list = [ADC_SWITCH_INH, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[18]

        '''ADC_SYS_1V8A'''
        float_list = [reply_extend[76], reply_extend[77], reply_extend[78], reply_extend[79]]
        ADC_SYS_1V8A = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.697 <= ADC_SYS_1V8A <= 1.914:
            exchange_list = [ADC_SYS_1V8A, 0, err_ADC_SYS_1V8A_1]
        else:
            err_ADC_SYS_1V8A_1 += 1
            exchange_list = [ADC_SYS_1V8A, 1, err_ADC_SYS_1V8A_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[19]

        '''ADC_SYS_1V8B'''
        float_list = [reply_extend[80], reply_extend[81], reply_extend[82], reply_extend[83]]
        ADC_SYS_1V8B = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.697 <= ADC_SYS_1V8B <= 1.914:
            exchange_list = [ADC_SYS_1V8B, 0, err_ADC_SYS_1V8B_1]
        else:
            err_ADC_SYS_1V8B_1 += 1
            exchange_list = [ADC_SYS_1V8B, 1, err_ADC_SYS_1V8B_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[20]

        '''ADC_SYS_1V2A'''
        float_list = [reply_extend[84], reply_extend[85], reply_extend[86], reply_extend[87]]
        ADC_SYS_1V2A = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.141 <= ADC_SYS_1V2A <= 1.262:
            exchange_list = [ADC_SYS_1V2A, 0, err_ADC_SYS_1V2A_1]
        else:
            err_ADC_SYS_1V2A_1 += 1
            exchange_list = [ADC_SYS_1V2A, 1, err_ADC_SYS_1V2A_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[21]

        '''ADC_SYS_1V2B'''
        float_list = [reply_extend[88], reply_extend[89], reply_extend[90], reply_extend[91]]
        ADC_SYS_1V2B = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.141 <= ADC_SYS_1V2B <= 1.262:
            exchange_list = [ADC_SYS_1V2B, 0, err_ADC_SYS_1V2B_1]
        else:
            err_ADC_SYS_1V2B_1 += 1
            exchange_list = [ADC_SYS_1V2B, 1, err_ADC_SYS_1V2B_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[22]

        '''ADC_SYS_3V3'''
        float_list = [reply_extend[92], reply_extend[93], reply_extend[94], reply_extend[95]]
        ADC_SYS_3V3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.136 <= ADC_SYS_3V3 <= 3.53:
            exchange_list = [ADC_SYS_3V3, 0, err_ADC_SYS_3V3_1]
        else:
            err_ADC_SYS_3V3_1 += 1
            exchange_list = [ADC_SYS_3V3, 1, err_ADC_SYS_3V3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[23]

        '''ADC_J3A_COREPD_0V8_G3'''
        float_list = [reply_extend[96], reply_extend[97], reply_extend[98], reply_extend[99]]
        ADC_J3A_COREPD_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= ADC_J3A_COREPD_0V8_G3 <= 0.81:
            exchange_list = [ADC_J3A_COREPD_0V8_G3, 0, err_ADC_J3A_COREPD_0V8_G3_1]
        else:
            err_ADC_J3A_COREPD_0V8_G3_1 += 1
            exchange_list = [ADC_J3A_COREPD_0V8_G3, 1, err_ADC_J3A_COREPD_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[24]

        '''ADC_J3A_VDDQDDR_1V1_G1'''
        float_list = [reply_extend[100], reply_extend[101], reply_extend[102], reply_extend[103]]
        ADC_J3A_VDDQDDR_1V1_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.0835 <= ADC_J3A_VDDQDDR_1V1_G1 <= 1.1165:
            exchange_list = [ADC_J3A_VDDQDDR_1V1_G1, 0, err_ADC_J3A_VDDQDDR_1V1_G1_1]
        else:
            err_ADC_J3A_VDDQDDR_1V1_G1_1 += 1
            exchange_list = [ADC_J3A_VDDQDDR_1V1_G1, 1, err_ADC_J3A_VDDQDDR_1V1_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[25]

        '''ADC_J3B_COREPD_0V8_G3'''
        float_list = [reply_extend[104], reply_extend[105], reply_extend[106], reply_extend[107]]
        ADC_J3B_COREPD_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= ADC_J3B_COREPD_0V8_G3 <= 0.81:
            exchange_list = [ADC_J3B_COREPD_0V8_G3, 0, err_ADC_J3B_COREPD_0V8_G3_1]
        else:
            err_ADC_J3B_COREPD_0V8_G3_1 += 1
            exchange_list = [ADC_J3B_COREPD_0V8_G3, 1, err_ADC_J3B_COREPD_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[26]

        '''ADC_J3B_VDDQDDR_1V1_G1'''
        float_list = [reply_extend[108], reply_extend[109], reply_extend[110], reply_extend[111]]
        ADC_J3B_VDDQDDR_1V1_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.0835 <= ADC_J3B_VDDQDDR_1V1_G1 <= 1.1165:
            exchange_list = [ADC_J3B_VDDQDDR_1V1_G1, 0, err_ADC_J3B_VDDQDDR_1V1_G1_1]
        else:
            err_ADC_J3B_VDDQDDR_1V1_G1_1 += 1
            exchange_list = [ADC_J3B_VDDQDDR_1V1_G1, 1, err_ADC_J3B_VDDQDDR_1V1_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[27]

        '''ADC_J3C_COREPD_0V8_G3'''
        float_list = [reply_extend[112], reply_extend[113], reply_extend[114], reply_extend[115]]
        ADC_J3C_COREPD_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= ADC_J3C_COREPD_0V8_G3 <= 0.81:
            exchange_list = [ADC_J3C_COREPD_0V8_G3, 0, err_ADC_J3C_COREPD_0V8_G3_1]
        else:
            err_ADC_J3C_COREPD_0V8_G3_1 += 1
            exchange_list = [ADC_J3C_COREPD_0V8_G3, 1, err_ADC_J3C_COREPD_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[28]

        '''ADC_J3C_VDDQDDR_1V1_G1'''
        float_list = [reply_extend[116], reply_extend[117], reply_extend[118], reply_extend[119]]
        ADC_J3C_VDDQDDR_1V1_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.0835 <= ADC_J3C_VDDQDDR_1V1_G1 <= 1.1165:
            exchange_list = [ADC_J3C_VDDQDDR_1V1_G1, 0, err_ADC_J3C_VDDQDDR_1V1_G1_1]
        else:
            err_ADC_J3C_VDDQDDR_1V1_G1_1 += 1
            exchange_list = [ADC_J3C_VDDQDDR_1V1_G1, 1, err_ADC_J3C_VDDQDDR_1V1_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[29]

        '''ADC_GPSANT_PWR_AMUX'''
        float_list = [reply_extend[120], reply_extend[121], reply_extend[122], reply_extend[123]]
        ADC_GPSANT_PWR_AMUX = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.5 <= ADC_GPSANT_PWR_AMUX <= 5.5:
            exchange_list = [ADC_GPSANT_PWR_AMUX, 0, err_ADC_GPSANT_PWR_AMUX_1]
        else:
            # err_ADC_GPSANT_PWR_AMUX_1 += 1
            exchange_list = [ADC_GPSANT_PWR_AMUX, 0, err_ADC_GPSANT_PWR_AMUX_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[30]

        '''ADC_GPS_3V3'''
        float_list = [reply_extend[124], reply_extend[125], reply_extend[126], reply_extend[127]]
        ADC_GPS_3V3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.16 <= ADC_GPS_3V3 <= 3.461:
            exchange_list = [ADC_GPS_3V3, 0, err_ADC_GPS_3V3_1]
        else:
            err_ADC_GPS_3V3_1 += 1
            exchange_list = [ADC_GPS_3V3, 1, err_ADC_GPS_3V3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[31]

        '''ADC_SYS_5V'''
        float_list = [reply_extend[128], reply_extend[129], reply_extend[130], reply_extend[131]]
        ADC_SYS_5V = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.566 <= ADC_SYS_5V <= 5.528:
            exchange_list = [ADC_SYS_5V, 0, err_ADC_SYS_5V_1]
        else:
            err_ADC_SYS_5V_1 += 1
            exchange_list = [ADC_SYS_5V, 1, err_ADC_SYS_5V_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[32]

        '''ADC_ETH_0V9'''
        float_list = [reply_extend[132], reply_extend[133], reply_extend[134], reply_extend[135]]
        ADC_ETH_0V9 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.855 <= ADC_ETH_0V9 <= 0.945:
            exchange_list = [ADC_ETH_0V9, 0, err_ADC_ETH_0V9_1]
        else:
            err_ADC_ETH_0V9_1 += 1
            exchange_list = [ADC_ETH_0V9, 1, err_ADC_ETH_0V9_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[33]

        '''ADC_HWID'''
        float_list = [reply_extend[136], reply_extend[137], reply_extend[138], reply_extend[139]]
        ADC_HWID = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.95 <= ADC_HWID <= 5.05:
            exchange_list = [ADC_HWID, 0, err_ADC_HWID_1]
        else:
            err_ADC_HWID_1 += 1
            exchange_list = [ADC_HWID, 1, err_ADC_HWID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[34]

        '''ADC_Proj_ID'''
        float_list = [reply_extend[140], reply_extend[141], reply_extend[142], reply_extend[143]]
        ADC_Proj_ID = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.95 <= ADC_Proj_ID <= 5.05:
            exchange_list = [ADC_Proj_ID, 0, err_ADC_Proj_ID_1]
        else:
            err_ADC_Proj_ID_1 += 1
            exchange_list = [ADC_Proj_ID, 1, err_ADC_Proj_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[35]

        '''J3A_Alive_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[148] == 0:
            exchange_list = [reply_extend[148], 0, err_J3A_Alive_Status_1]
        else:
            err_J3A_Alive_Status_1 += 1
            exchange_list = [reply_extend[148], 1, err_J3A_Alive_Status_1]
        data_analysed_1.append(exchange_list)       # data_analysed_1[36]

        '''J3B_Alive_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[149] == 0:
            exchange_list = [reply_extend[149], 0, err_J3B_Alive_Status_1]
        else:
            err_J3B_Alive_Status_1 += 1
            exchange_list = [reply_extend[149], 1, err_J3B_Alive_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[37]

        '''J3C_Alive_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[150] == 0:
            exchange_list = [reply_extend[150], 0, err_J3C_Alive_Status_1]
        else:
            err_J3C_Alive_Status_1 += 1
            exchange_list = [reply_extend[150], 1, err_J3C_Alive_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[38]

        '''MCU_J3A_SPI_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[151] == 0:
            exchange_list = [reply_extend[151], 0, err_MCU_J3A_SPI_Status_1]
        else:
            err_MCU_J3A_SPI_Status_1 += 1
            exchange_list = [reply_extend[151], 1, err_MCU_J3A_SPI_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[39]

        '''MCU_J3B_SPI_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[152] == 0:
            exchange_list = [reply_extend[152], 0, err_MCU_J3B_SPI_Status_1]
        else:
            err_MCU_J3B_SPI_Status_1 += 1
            exchange_list = [reply_extend[152], 1, err_MCU_J3B_SPI_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[40]

        '''MCU_J3C_SPI_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[153] == 0:
            exchange_list = [reply_extend[153], 0, err_MCU_J3C_SPI_Status_1]
        else:
            err_MCU_J3C_SPI_Status_1 += 1
            exchange_list = [reply_extend[153], 1, err_MCU_J3C_SPI_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[41]

        '''MCU_J3A_Ethernet_TimeOut_Staus'''
        exchange_list = [0, 0, 0]
        if reply_extend[154] == 0:
            exchange_list = [reply_extend[154], 0, err_MCU_J3A_Ethernet_TimeOut_Staus_1]
        else:
            err_MCU_J3A_Ethernet_TimeOut_Staus_1 += 1
            exchange_list = [reply_extend[154], 1, err_MCU_J3A_Ethernet_TimeOut_Staus_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[42]

        '''MCU_J3B_Ethernet_TimeOut_Staus'''
        exchange_list = [0, 0, 0]
        if reply_extend[155] == 0:
            exchange_list = [reply_extend[155], 0, err_MCU_J3B_Ethernet_TimeOut_Staus_1]
        else:
            err_MCU_J3B_Ethernet_TimeOut_Staus_1 += 1
            exchange_list = [reply_extend[155], 1, err_MCU_J3B_Ethernet_TimeOut_Staus_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[43]

        '''MCU_J3C_Ethernet_TimeOut_Staus'''
        exchange_list = [0, 0, 0]
        if reply_extend[156] == 0:
            exchange_list = [reply_extend[156], 0, err_MCU_J3C_Ethernet_TimeOut_Staus_1]
        else:
            err_MCU_J3C_Ethernet_TimeOut_Staus_1 += 1
            exchange_list = [reply_extend[156], 1, err_MCU_J3C_Ethernet_TimeOut_Staus_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[44]

        '''MCU_J3A_Ethernet_RollingCounter_Staus'''
        exchange_list = [0, 0, 0]
        if reply_extend[157] == 0:
            exchange_list = [reply_extend[157], 0, err_MCU_J3A_Ethernet_RollingCounter_Staus_1]
        else:
            err_MCU_J3A_Ethernet_RollingCounter_Staus_1 += 1
            exchange_list = [reply_extend[157], 1, err_MCU_J3A_Ethernet_RollingCounter_Staus_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[45]

        '''MCU_J3B_Ethernet_RollingCounter_Staus'''
        exchange_list = [0, 0, 0]
        if reply_extend[158] == 0:
            exchange_list = [reply_extend[158], 0, err_MCU_J3B_Ethernet_RollingCounter_Staus_1]
        else:
            err_MCU_J3B_Ethernet_RollingCounter_Staus_1 += 1
            exchange_list = [reply_extend[158], 1, err_MCU_J3B_Ethernet_RollingCounter_Staus_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[46]

        '''MCU_J3C_Ethernet_RollingCounter_Staus'''
        exchange_list = [0, 0, 0]
        if reply_extend[159] == 0:
            exchange_list = [reply_extend[159], 0, err_MCU_J3C_Ethernet_RollingCounter_Staus_1]
        else:
            err_MCU_J3C_Ethernet_RollingCounter_Staus_1 += 1
            exchange_list = [reply_extend[159], 1, err_MCU_J3C_Ethernet_RollingCounter_Staus_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[47]


        '''J3A_CPU_Load'''
        float_list = [reply_extend[160], reply_extend[161], reply_extend[162], reply_extend[163]]
        J3A_CPU_Load = float_cvt(float_list)
        exchange_list = [J3A_CPU_Load, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[48]

        '''J3B_CPU_Load'''
        float_list = [reply_extend[164], reply_extend[165], reply_extend[166], reply_extend[167]]
        J3B_CPU_Load = float_cvt(float_list)
        exchange_list = [J3B_CPU_Load, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[49]

        '''J3C_CPU_Load'''
        float_list = [reply_extend[168], reply_extend[169], reply_extend[170], reply_extend[171]]
        J3C_CPU_Load = float_cvt(float_list)
        exchange_list = [J3C_CPU_Load, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[50]

        '''J3A_eMMC_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[172] == 0:
            exchange_list = [reply_extend[172], 0, err_J3A_eMMC_Status_1]
        else:
            err_J3A_eMMC_Status_1 += 1
            exchange_list = [reply_extend[172], 1, err_J3A_eMMC_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[51]

        '''J3B_eMMC_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[173] == 0:
            exchange_list = [reply_extend[173], 0, err_J3B_eMMC_Status_1]
        else:
            err_J3B_eMMC_Status_1 += 1
            exchange_list = [reply_extend[173], 1, err_J3B_eMMC_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[52]

        '''J3C_eMMC_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[174] == 0:
            exchange_list = [reply_extend[174], 0, err_J3C_eMMC_Status_1]
        else:
            err_J3C_eMMC_Status_1 += 1
            exchange_list = [reply_extend[174], 1, err_J3C_eMMC_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[53]

        '''J3C_Surround_Camera_Input0_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[175] == 0:
            exchange_list = [reply_extend[175], 0, err_J3C_Surround_Camera_Input0_Link_Lock_1]
        else:
            err_J3C_Surround_Camera_Input0_Link_Lock_1 += 1
            exchange_list = [reply_extend[175], 1, err_J3C_Surround_Camera_Input0_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[54]

        '''J3C_Surround_Camera_Input0_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[176] == 0:
            exchange_list = [reply_extend[176], 0, err_J3C_Surround_Camera_Input0_Lock_1]
        else:
            err_J3C_Surround_Camera_Input0_Lock_1 += 1
            exchange_list = [reply_extend[176], 1, err_J3C_Surround_Camera_Input0_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[55]

        '''J3C_Surround_Camera_Input0_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[177] == 0:
            exchange_list = [reply_extend[177], 0, err_J3C_Surround_Camera_Input0_Fps_1]
        else:
            err_J3C_Surround_Camera_Input0_Fps_1 += 1
            exchange_list = [reply_extend[177], 1, err_J3C_Surround_Camera_Input0_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[56]

        '''J3C_Surround_Camera_Input0_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[178] == 0:
            exchange_list = [reply_extend[178], 0, err_J3C_Surround_Camera_Input0_Crc_1]
        else:
            err_J3C_Surround_Camera_Input0_Crc_1 += 1
            exchange_list = [reply_extend[178], 1, err_J3C_Surround_Camera_Input0_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[57]

        '''J3C_Surround_Camera_Input0_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[179] == 0:
            exchange_list = [reply_extend[179], 0, err_J3C_Surround_Camera_Input0_OpenLoad_1]
        else:
            err_J3C_Surround_Camera_Input0_OpenLoad_1 += 1
            exchange_list = [reply_extend[179], 1, err_J3C_Surround_Camera_Input0_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[58]

        '''J3C_Surround_Camera_Input0_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[180] == 0:
            exchange_list = [reply_extend[180], 0, err_J3C_Surround_Camera_Input0_Short_1]
        else:
            err_J3C_Surround_Camera_Input0_Short_1 += 1
            exchange_list = [reply_extend[180], 1, err_J3C_Surround_Camera_Input0_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[59]

        '''J3C_Surround_Camera_Input1_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[181] == 0:
            exchange_list = [reply_extend[181], 0, err_J3C_Surround_Camera_Input1_Link_Lock_1]
        else:
            err_J3C_Surround_Camera_Input1_Link_Lock_1 += 1
            exchange_list = [reply_extend[181], 1, err_J3C_Surround_Camera_Input1_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[60]

        '''J3C_Surround_Camera_Input1_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[182] == 0:
            exchange_list = [reply_extend[182], 0, err_J3C_Surround_Camera_Input1_Lock_1]
        else:
            err_J3C_Surround_Camera_Input1_Lock_1 += 1
            exchange_list = [reply_extend[182], 1, err_J3C_Surround_Camera_Input1_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[61]

        '''J3C_Surround_Camera_Input1_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[183] == 0:
            exchange_list = [reply_extend[183], 0, err_J3C_Surround_Camera_Input1_Fps_1]
        else:
            err_J3C_Surround_Camera_Input1_Fps_1 += 1
            exchange_list = [reply_extend[183], 1, err_J3C_Surround_Camera_Input1_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[62]

        '''J3C_Surround_Camera_Input1_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[184] == 0:
            exchange_list = [reply_extend[184], 0, err_J3C_Surround_Camera_Input1_Crc_1]
        else:
            err_J3C_Surround_Camera_Input1_Crc_1 += 1
            exchange_list = [reply_extend[184], 1, err_J3C_Surround_Camera_Input1_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[63]

        '''J3C_Surround_Camera_Input1_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[185] == 0:
            exchange_list = [reply_extend[185], 0, err_J3C_Surround_Camera_Input1_OpenLoad_1]
        else:
            err_J3C_Surround_Camera_Input1_OpenLoad_1 += 1
            exchange_list = [reply_extend[185], 1, err_J3C_Surround_Camera_Input1_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[64]

        '''J3C_Surround_Camera_Input1_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[186] == 0:
            exchange_list = [reply_extend[186], 0, err_J3C_Surround_Camera_Input1_Short_1]
        else:
            err_J3C_Surround_Camera_Input1_Short_1 += 1
            exchange_list = [reply_extend[186], 1, err_J3C_Surround_Camera_Input1_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[65]

        '''J3C_Surround_Camera_Input2_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[187] == 0:
            exchange_list = [reply_extend[187], 0, err_J3C_Surround_Camera_Input2_Link_Lock_1]
        else:
            err_J3C_Surround_Camera_Input2_Link_Lock_1 += 1
            exchange_list = [reply_extend[187], 1, err_J3C_Surround_Camera_Input2_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[66]

        '''J3C_Surround_Camera_Input2_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[188] == 0:
            exchange_list = [reply_extend[188], 0, err_J3C_Surround_Camera_Input2_Lock_1]
        else:
            err_J3C_Surround_Camera_Input2_Lock_1 += 1
            exchange_list = [reply_extend[188], 1, err_J3C_Surround_Camera_Input2_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[67]

        '''J3C_Surround_Camera_Input2_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[189] == 0:
            exchange_list = [reply_extend[189], 0, err_J3C_Surround_Camera_Input2_Fps_1]
        else:
            err_J3C_Surround_Camera_Input2_Fps_1 += 1
            exchange_list = [reply_extend[189], 1, err_J3C_Surround_Camera_Input2_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[68]

        '''J3C_Surround_Camera_Input2_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[190] == 0:
            exchange_list = [reply_extend[190], 0, err_J3C_Surround_Camera_Input2_Crc_1]
        else:
            err_J3C_Surround_Camera_Input2_Crc_1 += 1
            exchange_list = [reply_extend[190], 1, err_J3C_Surround_Camera_Input2_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[69]

        '''J3C_Surround_Camera_Input2_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[191] == 0:
            exchange_list = [reply_extend[191], 0, err_J3C_Surround_Camera_Input2_OpenLoad_1]
        else:
            err_J3C_Surround_Camera_Input2_OpenLoad_1 += 1
            exchange_list = [reply_extend[191], 1, err_J3C_Surround_Camera_Input2_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[70]

        '''J3C_Surround_Camera_Input2_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[192] == 0:
            exchange_list = [reply_extend[192], 0, err_J3C_Surround_Camera_Input2_Short_1]
        else:
            err_J3C_Surround_Camera_Input2_Short_1 += 1
            exchange_list = [reply_extend[192], 1, err_J3C_Surround_Camera_Input2_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[71]

        '''J3C_Surround_Camera_Input3_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[193] == 0:
            exchange_list = [reply_extend[193], 0, err_J3C_Surround_Camera_Input3_Link_Lock_1]
        else:
            err_J3C_Surround_Camera_Input3_Link_Lock_1 += 1
            exchange_list = [reply_extend[193], 1, err_J3C_Surround_Camera_Input3_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[72]

        '''J3C_Surround_Camera_Input3_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[194] == 0:
            exchange_list = [reply_extend[194], 0, err_J3C_Surround_Camera_Input3_Lock_1]
        else:
            err_J3C_Surround_Camera_Input3_Lock_1 += 1
            exchange_list = [reply_extend[194], 1, err_J3C_Surround_Camera_Input3_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[73]

        '''J3C_Surround_Camera_Input3_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[195] == 0:
            exchange_list = [reply_extend[195], 0, err_J3C_Surround_Camera_Input3_Fps_1]
        else:
            err_J3C_Surround_Camera_Input3_Fps_1 += 1
            exchange_list = [reply_extend[195], 1, err_J3C_Surround_Camera_Input3_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[74]

        '''J3C_Surround_Camera_Input3_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[196] == 0:
            exchange_list = [reply_extend[196], 0, err_J3C_Surround_Camera_Input3_Crc_1]
        else:
            err_J3C_Surround_Camera_Input3_Crc_1 += 1
            exchange_list = [reply_extend[196], 1, err_J3C_Surround_Camera_Input3_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[75]

        '''J3C_Surround_Camera_Input3_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[197] == 0:
            exchange_list = [reply_extend[197], 0, err_J3C_Surround_Camera_Input3_OpenLoad_1]
        else:
            err_J3C_Surround_Camera_Input3_OpenLoad_1 += 1
            exchange_list = [reply_extend[197], 1, err_J3C_Surround_Camera_Input3_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[76]

        '''J3C_Surround_Camera_Input3_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[198] == 0:
            exchange_list = [reply_extend[198], 0, err_J3C_Surround_Camera_Input3_Short_1]
        else:
            err_J3C_Surround_Camera_Input3_Short_1 += 1
            exchange_list = [reply_extend[198], 1, err_J3C_Surround_Camera_Input3_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[77]

        '''J3C_Surround_Camera_Output_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[199] == 0:
            exchange_list = [reply_extend[199], 0, err_J3C_Surround_Camera_Output_Link_Lock_1]
        else:
            err_J3C_Surround_Camera_Output_Link_Lock_1 += 1
            exchange_list = [reply_extend[199], 1, err_J3C_Surround_Camera_Output_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[78]

        '''J3C_Surround_Camera_Output_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[200] == 0:
            exchange_list = [reply_extend[200], 0, err_J3C_Surround_Camera_Output_Lock_1]
        else:
            err_J3C_Surround_Camera_Output_Lock_1 += 1
            exchange_list = [reply_extend[200], 1, err_J3C_Surround_Camera_Output_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[79]

        '''J3C_Surround_Camera_Output_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[201] == 0:
            exchange_list = [reply_extend[201], 0, err_J3C_Surround_Camera_Output_Fps_1]
        else:
            err_J3C_Surround_Camera_Output_Fps_1 += 1
            exchange_list = [reply_extend[201], 1, err_J3C_Surround_Camera_Output_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[80]

        '''J3C_Surround_Camera_Output_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[202] == 0:
            exchange_list = [reply_extend[202], 0, err_J3C_Surround_Camera_Output_Crc_1]
        else:
            err_J3C_Surround_Camera_Output_Crc_1 += 1
            exchange_list = [reply_extend[202], 1, err_J3C_Surround_Camera_Output_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[81]

        '''J3A_Front_Camera_Input_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[203] == 0:
            exchange_list = [reply_extend[203], 0, err_J3A_Front_Camera_Input_Link_Lock_1]
        else:
            err_J3A_Front_Camera_Input_Link_Lock_1 += 1
            exchange_list = [reply_extend[203], 1, err_J3A_Front_Camera_Input_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[82]

        '''J3A_Front_Camera_Input_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[204] == 0:
            exchange_list = [reply_extend[204], 0, err_J3A_Front_Camera_Input_Lock_1]
        else:
            err_J3A_Front_Camera_Input_Lock_1 += 1
            exchange_list = [reply_extend[204], 1, err_J3A_Front_Camera_Input_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[83]

        '''J3A_Front_Camera_Input_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[205] == 0:
            exchange_list = [reply_extend[205], 0, err_J3A_Front_Camera_Input_Fps_1]
        else:
            err_J3A_Front_Camera_Input_Fps_1 += 1
            exchange_list = [reply_extend[205], 1, err_J3A_Front_Camera_Input_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[84]

        '''J3A_Front_Camera_Input_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[206] == 0:
            exchange_list = [reply_extend[206], 0, err_J3A_Front_Camera_Input_Crc_1]
        else:
            err_J3A_Front_Camera_Input_Crc_1 += 1
            exchange_list = [reply_extend[206], 1, err_J3A_Front_Camera_Input_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[85]

        '''J3A_Front_Camera_Input_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[207] == 0:
            exchange_list = [reply_extend[207], 0, err_J3A_Front_Camera_Input_OpenLoad_1]
        else:
            err_J3A_Front_Camera_Input_OpenLoad_1 += 1
            exchange_list = [reply_extend[207], 1, err_J3A_Front_Camera_Input_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[86]

        '''J3A_Front_Camera_Input_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[208] == 0:
            exchange_list = [reply_extend[208], 0, err_J3A_Front_Camera_Input_Short_1]
        else:
            err_J3A_Front_Camera_Input_Short_1 += 1
            exchange_list = [reply_extend[208], 1, err_J3A_Front_Camera_Input_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[87]

        '''J3A_Front_Camera_Output_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[209] == 0:
            exchange_list = [reply_extend[209], 0, err_J3A_Front_Camera_Output_Link_Lock_1]
        else:
            err_J3A_Front_Camera_Output_Link_Lock_1 += 1
            exchange_list = [reply_extend[209], 1, err_J3A_Front_Camera_Output_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[88]

        '''J3A_Front_Camera_Output_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[210] == 0:
            exchange_list = [reply_extend[210], 0, err_J3A_Front_Camera_Output_Lock_1]
        else:
            err_J3A_Front_Camera_Output_Lock_1 += 1
            exchange_list = [reply_extend[210], 1, err_J3A_Front_Camera_Output_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[89]

        '''J3A_Front_Camera_Output_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[211] == 0:
            exchange_list = [reply_extend[211], 0, err_J3A_Front_Camera_Output_Fps_1]
        else:
            err_J3A_Front_Camera_Output_Fps_1 += 1
            exchange_list = [reply_extend[211], 1, err_J3A_Front_Camera_Output_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[90]

        '''J3A_Front_Camera_Output_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[212] == 0:
            exchange_list = [reply_extend[212], 0, err_J3A_Front_Camera_Output_Crc_1]
        else:
            err_J3A_Front_Camera_Output_Crc_1 += 1
            exchange_list = [reply_extend[212], 1, err_J3A_Front_Camera_Output_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[91]

        '''Rear_Camera'''
        '''J3B_Rear_Camera_Input_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[213] == 0:
            exchange_list = [reply_extend[213], 0, err_J3B_Rear_Camera_Input_Link_Lock_1]
        else:
            err_J3B_Rear_Camera_Input_Link_Lock_1 += 1
            exchange_list = [reply_extend[213], 1, err_J3B_Rear_Camera_Input_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[92]

        '''J3B_Rear_Camera_Input_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[214] == 0:
            exchange_list = [reply_extend[214], 0, err_J3B_Rear_Camera_Input_Lock_1]
        else:
            err_J3B_Rear_Camera_Input_Lock_1 += 1
            exchange_list = [reply_extend[214], 1, err_J3B_Rear_Camera_Input_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[93]

        '''J3B_Rear_Camera_Input_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[215] == 0:
            exchange_list = [reply_extend[215], 0, err_J3B_Rear_Camera_Input_Fps_1]
        else:
            err_J3B_Rear_Camera_Input_Fps_1 += 1
            exchange_list = [reply_extend[215], 1, err_J3B_Rear_Camera_Input_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[94]

        '''J3B_Rear_Camera_Input_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[216] == 0:
            exchange_list = [reply_extend[216], 0, err_J3B_Rear_Camera_Input_Crc_1]
        else:
            err_J3B_Rear_Camera_Input_Crc_1 += 1
            exchange_list = [reply_extend[216], 1, err_J3B_Rear_Camera_Input_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[95]

        '''J3B_Rear_Camera_Input_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[217] == 0:
            exchange_list = [reply_extend[217], 0, err_J3B_Rear_Camera_Input_OpenLoad_1]
        else:
            err_J3B_Rear_Camera_Input_OpenLoad_1 += 1
            exchange_list = [reply_extend[217], 1, err_J3B_Rear_Camera_Input_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[96]

        '''J3B_Rear_Camera_Input_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[218] == 0:
            exchange_list = [reply_extend[218], 0, err_J3B_Rear_Camera_Input_Short_1]
        else:
            err_J3B_Rear_Camera_Input_Short_1 += 1
            exchange_list = [reply_extend[218], 1, err_J3B_Rear_Camera_Input_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[97]

        '''Side_Camera'''
        '''J3B_Side_Camera_Input0_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[219] == 0:
            exchange_list = [reply_extend[219], 0, err_J3B_Side_Camera_Input0_Link_Lock_1]
        else:
            err_J3B_Side_Camera_Input0_Link_Lock_1 += 1
            exchange_list = [reply_extend[219], 1, err_J3B_Side_Camera_Input0_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[98]

        '''J3B_Side_Camera_Input0_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[220] == 0:
            exchange_list = [reply_extend[220], 0, err_J3B_Side_Camera_Input0_Lock_1]
        else:
            err_J3B_Side_Camera_Input0_Lock_1 += 1
            exchange_list = [reply_extend[220], 1, err_J3B_Side_Camera_Input0_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[99]

        '''J3B_Side_Camera_Input0_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[221] == 0:
            exchange_list = [reply_extend[221], 0, err_J3B_Side_Camera_Input0_Fps_1]
        else:
            err_J3B_Side_Camera_Input0_Fps_1 += 1
            exchange_list = [reply_extend[221], 1, err_J3B_Side_Camera_Input0_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[100]

        '''J3B_Side_Camera_Input0_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[222] == 0:
            exchange_list = [reply_extend[222], 0, err_J3B_Side_Camera_Input0_Crc_1]
        else:
            err_J3B_Side_Camera_Input0_Crc_1 += 1
            exchange_list = [reply_extend[222], 1, err_J3B_Side_Camera_Input0_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[101]

        '''J3B_Side_Camera_Input0_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[223] == 0:
            exchange_list = [reply_extend[223], 0, err_J3B_Side_Camera_Input0_OpenLoad_1]
        else:
            err_J3B_Side_Camera_Input0_OpenLoad_1 += 1
            exchange_list = [reply_extend[223], 1, err_J3B_Side_Camera_Input0_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[102]

        '''J3B_Side_Camera_Input0_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[224] == 0:
            exchange_list = [reply_extend[224], 0, err_J3B_Side_Camera_Input0_Short_1]
        else:
            err_J3B_Side_Camera_Input0_Short_1 += 1
            exchange_list = [reply_extend[224], 1, err_J3B_Side_Camera_Input0_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[103]

        '''J3B_Side_Camera_Input1_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[225] == 0:
            exchange_list = [reply_extend[225], 0, err_J3B_Side_Camera_Input1_Link_Lock_1]
        else:
            err_J3B_Side_Camera_Input1_Link_Lock_1 += 1
            exchange_list = [reply_extend[225], 1, err_J3B_Side_Camera_Input1_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[104]

        '''J3B_Side_Camera_Input1_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[226] == 0:
            exchange_list = [reply_extend[226], 0, err_J3B_Side_Camera_Input1_Lock_1]
        else:
            err_J3B_Side_Camera_Input1_Lock_1 += 1
            exchange_list = [reply_extend[226], 1, err_J3B_Side_Camera_Input1_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[105]

        '''J3B_Side_Camera_Input1_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[227] == 0:
            exchange_list = [reply_extend[227], 0, err_J3B_Side_Camera_Input1_Fps_1]
        else:
            err_J3B_Side_Camera_Input1_Fps_1 += 1
            exchange_list = [reply_extend[227], 1, err_J3B_Side_Camera_Input1_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[106]

        '''J3B_Side_Camera_Input1_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[228] == 0:
            exchange_list = [reply_extend[228], 0, err_J3B_Side_Camera_Input1_Crc_1]
        else:
            err_J3B_Side_Camera_Input1_Crc_1 += 1
            exchange_list = [reply_extend[228], 1, err_J3B_Side_Camera_Input1_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[107]

        '''J3B_Side_Camera_Input1_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[229] == 0:
            exchange_list = [reply_extend[229], 0, err_J3B_Side_Camera_Input1_OpenLoad_1]
        else:
            err_J3B_Side_Camera_Input1_OpenLoad_1 += 1
            exchange_list = [reply_extend[229], 1, err_J3B_Side_Camera_Input1_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[108]

        '''J3B_Side_Camera_Input1_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[230] == 0:
            exchange_list = [reply_extend[230], 0, err_J3B_Side_Camera_Input1_Short_1]
        else:
            err_J3B_Side_Camera_Input1_Short_1 += 1
            exchange_list = [reply_extend[230], 1, err_J3B_Side_Camera_Input1_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[109]

        '''J3B_Side_Camera_Input2_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[231] == 0:
            exchange_list = [reply_extend[231], 0, err_J3B_Side_Camera_Input2_Link_Lock_1]
        else:
            err_J3B_Side_Camera_Input2_Link_Lock_1 += 1
            exchange_list = [reply_extend[231], 1, err_J3B_Side_Camera_Input2_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[110]

        '''J3B_Side_Camera_Input2_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[232] == 0:
            exchange_list = [reply_extend[232], 0, err_J3B_Side_Camera_Input2_Lock_1]
        else:
            err_J3B_Side_Camera_Input2_Lock_1 += 1
            exchange_list = [reply_extend[232], 1, err_J3B_Side_Camera_Input2_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[111]

        '''J3B_Side_Camera_Input2_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[233] == 0:
            exchange_list = [reply_extend[233], 0, err_J3B_Side_Camera_Input2_Fps_1]
        else:
            err_J3B_Side_Camera_Input2_Fps_1 += 1
            exchange_list = [reply_extend[233], 1, err_J3B_Side_Camera_Input2_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[112]

        '''J3B_Side_Camera_Input2_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[234] == 0:
            exchange_list = [reply_extend[234], 0, err_J3B_Side_Camera_Input2_Crc_1]
        else:
            err_J3B_Side_Camera_Input2_Crc_1 += 1
            exchange_list = [reply_extend[234], 1, err_J3B_Side_Camera_Input2_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[113]

        '''J3B_Side_Camera_Input2_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[235] == 0:
            exchange_list = [reply_extend[235], 0, err_J3B_Side_Camera_Input2_OpenLoad_1]
        else:
            err_J3B_Side_Camera_Input2_OpenLoad_1 += 1
            exchange_list = [reply_extend[235], 1, err_J3B_Side_Camera_Input2_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[114]

        '''J3B_Side_Camera_Input2_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[236] == 0:
            exchange_list = [reply_extend[236], 0, err_J3B_Side_Camera_Input2_Short_1]
        else:
            err_J3B_Side_Camera_Input2_Short_1 += 1
            exchange_list = [reply_extend[236], 1, err_J3B_Side_Camera_Input2_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[115]

        '''J3B_Side_Camera_Input3_Link_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[237] == 0:
            exchange_list = [reply_extend[237], 0, err_J3B_Side_Camera_Input3_Link_Lock_1]
        else:
            err_J3B_Side_Camera_Input3_Link_Lock_1 += 1
            exchange_list = [reply_extend[237], 1, err_J3B_Side_Camera_Input3_Link_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[116]

        '''J3B_Side_Camera_Input3_Lock'''
        exchange_list = [0, 0, 0]
        if reply_extend[238] == 0:
            exchange_list = [reply_extend[238], 0, err_J3B_Side_Camera_Input3_Lock_1]
        else:
            err_J3B_Side_Camera_Input3_Lock_1 += 1
            exchange_list = [reply_extend[238], 1, err_J3B_Side_Camera_Input3_Lock_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[117]

        '''J3B_Side_Camera_Input3_Fps'''
        exchange_list = [0, 0, 0]
        if reply_extend[239] == 0:
            exchange_list = [reply_extend[239], 0, err_J3B_Side_Camera_Input3_Fps_1]
        else:
            err_J3B_Side_Camera_Input3_Fps_1 += 1
            exchange_list = [reply_extend[239], 1, err_J3B_Side_Camera_Input3_Fps_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[118]

        '''J3B_Side_Camera_Input3_Crc'''
        exchange_list = [0, 0, 0]
        if reply_extend[240] == 0:
            exchange_list = [reply_extend[240], 0, err_J3B_Side_Camera_Input3_Crc_1]
        else:
            err_J3B_Side_Camera_Input3_Crc_1 += 1
            exchange_list = [reply_extend[240], 1, err_J3B_Side_Camera_Input3_Crc_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[119]

        '''J3B_Side_Camera_Input3_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[241] == 0:
            exchange_list = [reply_extend[241], 0, err_J3B_Side_Camera_Input3_OpenLoad_1]
        else:
            err_J3B_Side_Camera_Input3_OpenLoad_1 += 1
            exchange_list = [reply_extend[241], 1, err_J3B_Side_Camera_Input3_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[120]

        '''J3B_Side_Camera_Input3_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[242] == 0:
            exchange_list = [reply_extend[242], 0, err_J3B_Side_Camera_Input3_Short_1]
        else:
            err_J3B_Side_Camera_Input3_Short_1 += 1
            exchange_list = [reply_extend[242], 1, err_J3B_Side_Camera_Input3_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[121]


        '''GNSS'''
        '''J3A_F9K_GNSS_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[244] == 0:
            exchange_list = [reply_extend[244], 0, err_J3A_F9K_GNSS_Status_1]
        else:
            err_J3A_F9K_GNSS_Status_1 += 1
            exchange_list = [reply_extend[244], 1, err_J3A_F9K_GNSS_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[122]

        '''MCU_F9K_GNSS_OpenLoad'''
        exchange_list = [0, 0, 0]
        if reply_extend[245] == 0:
            exchange_list = [reply_extend[245], 0, err_MCU_F9K_GNSS_OpenLoad_1]
        else:
            err_MCU_F9K_GNSS_OpenLoad_1 += 1
            exchange_list = [reply_extend[245], 1, err_MCU_F9K_GNSS_OpenLoad_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[123]

        '''MCU_F9K_GNSS_Short'''
        exchange_list = [0, 0, 0]
        if reply_extend[246] == 0:
            exchange_list = [reply_extend[246], 0, err_MCU_F9K_GNSS_Short_1]
        else:
            err_MCU_F9K_GNSS_Short_1 += 1
            exchange_list = [reply_extend[246], 1, err_MCU_F9K_GNSS_Short_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[124]

        '''MCU_F9K_GNSS_Supply'''
        exchange_list = [0, 0, 0]
        if reply_extend[247] == 0:
            exchange_list = [reply_extend[247], 0, err_MCU_F9K_GNSS_Supply_1]
        else:
            err_MCU_F9K_GNSS_Supply_1 += 1
            exchange_list = [reply_extend[247], 1, err_MCU_F9K_GNSS_Supply_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[125]

        '''J3A_F9K_IMU_Z'''
        float_list = [reply_extend[248], reply_extend[249], reply_extend[250], reply_extend[251]]
        J3A_F9K_IMU_Z = float_cvt(float_list)
        exchange_list = [J3A_F9K_IMU_Z, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[126]

        # '''J3A_IAM_20685_IMU_Z'''
        # float_list = [reply_extend[252], reply_extend[253], reply_extend[254], reply_extend[255]]
        # J3A_IAM_20685_IMU_Z = float_cvt(float_list)
        # exchange_list = [J3A_IAM_20685_IMU_Z, 0, 0]
        # data_analysed_1.append(exchange_list)  # data_analysed_1[127]

        '''J3A_F9K_CN_Value'''
        float_list = [reply_extend[252], reply_extend[253], reply_extend[254], reply_extend[255]]
        J3A_F9K_CN_Value = int.from_bytes(float_list, byteorder='little', signed=True)
        exchange_list = [J3A_F9K_CN_Value, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[127]


        '''TEMP'''
        '''MCU_TEMP'''
        float_list = [reply_extend[276], reply_extend[277], reply_extend[278], reply_extend[279]]
        MCU_TEMP = float_cvt(float_list)
        exchange_list = [MCU_TEMP, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[128]

        '''J3A_5024_TEMP'''
        float_list = [reply_extend[280], reply_extend[281], reply_extend[282], reply_extend[283]]
        J3A_5024_TEMP = float_cvt(float_list)
        exchange_list = [J3A_5024_TEMP, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[129]

        '''J3B_5024_TEMP'''
        float_list = [reply_extend[284], reply_extend[285], reply_extend[286], reply_extend[287]]
        J3B_5024_TEMP = float_cvt(float_list)
        exchange_list = [J3B_5024_TEMP, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[130]

        '''J3C_5024_TEMP'''
        float_list = [reply_extend[288], reply_extend[289], reply_extend[290], reply_extend[291]]
        J3C_5024_TEMP = float_cvt(float_list)
        exchange_list = [J3C_5024_TEMP, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[131]

        '''PCB_TEMP'''
        float_list = [reply_extend[292], reply_extend[293], reply_extend[294], reply_extend[295]]
        PCB_TEMP = float_cvt(float_list)
        exchange_list = [PCB_TEMP, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[132]


        '''J3A'''
        '''J3A_CPU_0V8_G3'''
        float_list = [reply_extend[296], reply_extend[297], reply_extend[298], reply_extend[299]]
        J3A_CPU_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.76 <= J3A_CPU_0V8_G3 <= 0.81:
            exchange_list = [J3A_CPU_0V8_G3, 0, err_J3A_CPU_0V8_G3_1]
        else:
            err_J3A_CPU_0V8_G3_1 += 1
            exchange_list = [J3A_CPU_0V8_G3, 1, err_J3A_CPU_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[133]

        '''J3A_DDR_0V8_G3'''
        float_list = [reply_extend[300], reply_extend[301], reply_extend[302], reply_extend[303]]
        J3A_DDR_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3A_DDR_0V8_G3 <= 0.81:
            exchange_list = [J3A_DDR_0V8_G3, 0, err_J3A_DDR_0V8_G3_1]
        else:
            err_J3A_DDR_0V8_G3_1 += 1
            exchange_list = [J3A_DDR_0V8_G3, 1, err_J3A_DDR_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[134]

        '''J3A_VDD_0V8_G4'''
        float_list = [reply_extend[304], reply_extend[305], reply_extend[306], reply_extend[307]]
        J3A_VDD_0V8_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3A_VDD_0V8_G4 <= 0.81:
            exchange_list = [J3A_VDD_0V8_G4, 0, err_J3A_VDD_0V8_G4_1]
        else:
            err_J3A_VDD_0V8_G4_1 += 1
            exchange_list = [J3A_VDD_0V8_G4, 1, err_J3A_VDD_0V8_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[135]

        '''J3A_COREAO_0V8_G2'''
        float_list = [reply_extend[308], reply_extend[309], reply_extend[310], reply_extend[311]]
        J3A_COREAO_0V8_G2 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3A_COREAO_0V8_G2 <= 0.81:
            exchange_list = [J3A_COREAO_0V8_G2, 0, err_J3A_COREAO_0V8_G2_1]
        else:
            err_J3A_COREAO_0V8_G2_1 += 1
            exchange_list = [J3A_COREAO_0V8_G2, 1, err_J3A_COREAO_0V8_G2_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[136]

        '''J3A_CNN0_0V8_G3'''
        float_list = [reply_extend[312], reply_extend[313], reply_extend[314], reply_extend[315]]
        J3A_CNN0_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3A_CNN0_0V8_G3 <= 0.81:
            exchange_list = [J3A_CNN0_0V8_G3, 0, err_J3A_CNN0_0V8_G3_1]
        else:
            err_J3A_CNN0_0V8_G3_1 += 1
            exchange_list = [J3A_CNN0_0V8_G3, 1, err_J3A_CNN0_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[137]

        '''J3A_CNN1_0V8_G3'''
        float_list = [reply_extend[316], reply_extend[317], reply_extend[318], reply_extend[319]]
        J3A_CNN1_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3A_CNN1_0V8_G3 <= 0.81:
            exchange_list = [J3A_CNN1_0V8_G3, 0, err_J3A_CNN1_0V8_G3_1]
        else:
            err_J3A_CNN1_0V8_G3_1 += 1
            exchange_list = [J3A_CNN1_0V8_G3, 1, err_J3A_CNN1_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[138]

        '''J3A_VDD_1V8_G1'''
        float_list = [reply_extend[320], reply_extend[321], reply_extend[322], reply_extend[323]]
        J3A_VDD_1V8_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.71 <= J3A_VDD_1V8_G1 <= 1.89:
            exchange_list = [J3A_VDD_1V8_G1, 0, err_J3A_VDD_1V8_G1_1]
        else:
            err_J3A_VDD_1V8_G1_1 += 1
            exchange_list = [J3A_VDD_1V8_G1, 1, err_J3A_VDD_1V8_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[139]

        '''J3A_EMMC_3V3_G1'''
        float_list = [reply_extend[324], reply_extend[325], reply_extend[326], reply_extend[327]]
        J3A_EMMC_3V3_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.135 <= J3A_EMMC_3V3_G1 <= 3.465:
            exchange_list = [J3A_EMMC_3V3_G1, 0, err_J3A_EMMC_3V3_G1_1]
        else:
            err_J3A_EMMC_3V3_G1_1 += 1
            exchange_list = [J3A_EMMC_3V3_G1, 1, err_J3A_EMMC_3V3_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[140]

        '''J3A_VDD_3V3_G1'''
        float_list = [reply_extend[328], reply_extend[329], reply_extend[330], reply_extend[331]]
        J3A_VDD_3V3_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.135 <= J3A_VDD_3V3_G1 <= 3.465:
            exchange_list = [J3A_VDD_3V3_G1, 0, err_J3A_VDD_3V3_G1_1]
        else:
            err_J3A_VDD_3V3_G1_1 += 1
            exchange_list = [J3A_VDD_3V3_G1, 1, err_J3A_VDD_3V3_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[141]

        '''J3A_VCC_3V3_G4'''
        float_list = [reply_extend[332], reply_extend[333], reply_extend[334], reply_extend[335]]
        J3A_VCC_3V3_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.135 <= J3A_VCC_3V3_G4 <= 3.465:
            exchange_list = [J3A_VCC_3V3_G4, 0, err_J3A_VCC_3V3_G4_1]
        else:
            err_J3A_VCC_3V3_G4_1 += 1
            exchange_list = [J3A_VCC_3V3_G4, 1, err_J3A_VCC_3V3_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[142]

        '''J3A_VCC_1V8_G4'''
        float_list = [reply_extend[336], reply_extend[337], reply_extend[338], reply_extend[339]]
        J3A_VCC_1V8_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.71 <= J3A_VCC_1V8_G4 <= 1.89:
            exchange_list = [J3A_VCC_1V8_G4, 0, err_J3A_VCC_1V8_G4_1]
        else:
            err_J3A_VCC_1V8_G4_1 += 1
            exchange_list = [J3A_VCC_1V8_G4, 1, err_J3A_VCC_1V8_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[143]

        '''J3A_TEMP_IC'''
        float_list = [reply_extend[340], reply_extend[341], reply_extend[342], reply_extend[343]]
        J3A_TEMP_IC = float_cvt(float_list)
        exchange_list = [J3A_TEMP_IC, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[144]

        '''J3A_TEMP_SW1'''
        float_list = [reply_extend[344], reply_extend[345], reply_extend[346], reply_extend[347]]
        J3A_TEMP_SW1 = float_cvt(float_list)
        exchange_list = [J3A_TEMP_SW1, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[145]

        '''J3A_TEMP_SW2'''
        float_list = [reply_extend[348], reply_extend[349], reply_extend[350], reply_extend[351]]
        J3A_TEMP_SW2 = float_cvt(float_list)
        exchange_list = [J3A_TEMP_SW2, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[146]

        '''J3A_TEMP_SW3'''
        float_list = [reply_extend[352], reply_extend[353], reply_extend[354], reply_extend[355]]
        J3A_TEMP_SW3 = float_cvt(float_list)
        exchange_list = [J3A_TEMP_SW3, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[147]

        '''J3A_TEMP_SW4'''
        float_list = [reply_extend[356], reply_extend[357], reply_extend[358], reply_extend[359]]
        J3A_TEMP_SW4 = float_cvt(float_list)
        exchange_list = [J3A_TEMP_SW4, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[148]

        '''J3A_TEMP_SW5'''
        float_list = [reply_extend[360], reply_extend[361], reply_extend[362], reply_extend[363]]
        J3A_TEMP_SW5 = float_cvt(float_list)
        exchange_list = [J3A_TEMP_SW5, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[149]

        '''J3A_TEMP_SW6'''
        float_list = [reply_extend[364], reply_extend[365], reply_extend[366], reply_extend[367]]
        J3A_TEMP_SW6 = float_cvt(float_list)
        exchange_list = [J3A_TEMP_SW6, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[150]

        '''J3A_TEMP_SW7'''
        float_list = [reply_extend[368], reply_extend[369], reply_extend[370], reply_extend[371]]
        J3A_TEMP_SW7 = float_cvt(float_list)
        exchange_list = [J3A_TEMP_SW7, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[151]

        '''J3A_TEMP_TEMP_LDO1_2'''
        float_list = [reply_extend[372], reply_extend[373], reply_extend[374], reply_extend[375]]
        J3A_TEMP_TEMP_LDO1_2 = float_cvt(float_list)
        exchange_list = [J3A_TEMP_TEMP_LDO1_2, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[152]

        '''J3A_TEMP_TEMP_LDO3_4'''
        float_list = [reply_extend[376], reply_extend[377], reply_extend[378], reply_extend[379]]
        J3A_TEMP_TEMP_LDO3_4 = float_cvt(float_list)
        exchange_list = [J3A_TEMP_TEMP_LDO3_4, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[153]

        '''J3B'''
        '''J3B_CPU_0V8_G3'''
        float_list = [reply_extend[380], reply_extend[381], reply_extend[382], reply_extend[383]]
        J3B_CPU_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3B_CPU_0V8_G3 <= 0.81:
            exchange_list = [J3B_CPU_0V8_G3, 0, err_J3B_CPU_0V8_G3_1]
        else:
            err_J3B_CPU_0V8_G3_1 += 1
            exchange_list = [J3B_CPU_0V8_G3, 1, err_J3B_CPU_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[154]

        '''J3B_DDR_0V8_G3'''
        float_list = [reply_extend[384], reply_extend[385], reply_extend[386], reply_extend[387]]
        J3B_DDR_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3B_DDR_0V8_G3 <= 0.81:
            exchange_list = [J3B_DDR_0V8_G3, 0, err_J3B_DDR_0V8_G3_1]
        else:
            err_J3B_DDR_0V8_G3_1 += 1
            exchange_list = [J3B_DDR_0V8_G3, 1, err_J3B_DDR_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[155]

        '''J3B_VDD_0V8_G4'''
        float_list = [reply_extend[388], reply_extend[389], reply_extend[390], reply_extend[391]]
        J3B_VDD_0V8_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3B_VDD_0V8_G4 <= 0.81:
            exchange_list = [J3B_VDD_0V8_G4, 0, err_J3B_VDD_0V8_G4_1]
        else:
            err_J3B_VDD_0V8_G4_1 += 1
            exchange_list = [J3B_VDD_0V8_G4, 1, err_J3B_VDD_0V8_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[156]

        '''J3B_COREAO_0V8_G2'''
        float_list = [reply_extend[392], reply_extend[393], reply_extend[394], reply_extend[395]]
        J3B_COREAO_0V8_G2 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3B_COREAO_0V8_G2 <= 0.81:
            exchange_list = [J3B_COREAO_0V8_G2, 0, err_J3B_COREAO_0V8_G2_1]
        else:
            err_J3B_COREAO_0V8_G2_1 += 1
            exchange_list = [J3B_COREAO_0V8_G2, 1, err_J3B_COREAO_0V8_G2_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[157]

        '''J3B_CNN0_0V8_G3'''
        float_list = [reply_extend[396], reply_extend[397], reply_extend[398], reply_extend[399]]
        J3B_CNN0_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3B_CNN0_0V8_G3 <= 0.81:
            exchange_list = [J3B_CNN0_0V8_G3, 0, err_J3B_CNN0_0V8_G3_1]
        else:
            err_J3B_CNN0_0V8_G3_1 += 1
            exchange_list = [J3B_CNN0_0V8_G3, 1, err_J3B_CNN0_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[158]

        '''J3B_CNN1_0V8_G3'''
        float_list = [reply_extend[400], reply_extend[401], reply_extend[402], reply_extend[403]]
        J3B_CNN1_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3B_CNN1_0V8_G3 <= 0.81:
            exchange_list = [J3B_CNN1_0V8_G3, 0, err_J3B_CNN1_0V8_G3_1]
        else:
            err_J3B_CNN1_0V8_G3_1 += 1
            exchange_list = [J3B_CNN1_0V8_G3, 1, err_J3B_CNN1_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[159]

        '''J3B_VDD_1V8_G1'''
        float_list = [reply_extend[404], reply_extend[405], reply_extend[406], reply_extend[407]]
        J3B_VDD_1V8_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.764 <= J3B_VDD_1V8_G1 <= 1.836:
            exchange_list = [J3B_VDD_1V8_G1, 0, err_J3B_VDD_1V8_G1_1]
        else:
            err_J3B_VDD_1V8_G1_1 += 1
            exchange_list = [J3B_VDD_1V8_G1, 1, err_J3B_VDD_1V8_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[160]

        '''J3B_EMMC_3V3_G1'''
        float_list = [reply_extend[408], reply_extend[409], reply_extend[410], reply_extend[411]]
        J3B_EMMC_3V3_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.201 <= J3B_EMMC_3V3_G1 <= 3.399:
            exchange_list = [J3B_EMMC_3V3_G1, 0, err_J3B_EMMC_3V3_G1_1]
        else:
            err_J3B_EMMC_3V3_G1_1 += 1
            exchange_list = [J3B_EMMC_3V3_G1, 1, err_J3B_EMMC_3V3_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[161]

        '''J3B_VDD_3V3_G1'''
        float_list = [reply_extend[412], reply_extend[413], reply_extend[414], reply_extend[415]]
        J3B_VDD_3V3_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.201<= J3B_VDD_3V3_G1 <= 3.399:
            exchange_list = [J3B_VDD_3V3_G1, 0, err_J3B_VDD_3V3_G1_1]
        else:
            err_J3B_VDD_3V3_G1_1 += 1
            exchange_list = [J3B_VDD_3V3_G1, 1, err_J3B_VDD_3V3_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[162]

        '''J3B_VCC_3V3_G4'''
        float_list = [reply_extend[416], reply_extend[417], reply_extend[418], reply_extend[419]]
        J3B_VCC_3V3_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.201 <= J3B_VCC_3V3_G4 <= 3.399:
            exchange_list = [J3B_VCC_3V3_G4, 0, err_J3B_VCC_3V3_G4_1]
        else:
            err_J3B_VCC_3V3_G4_1 += 1
            exchange_list = [J3B_VCC_3V3_G4, 1, err_J3B_VCC_3V3_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[163]

        '''J3B_VCC_1V8_G4'''
        float_list = [reply_extend[420], reply_extend[421], reply_extend[422], reply_extend[423]]
        J3B_VCC_1V8_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.746 <= J3B_VCC_1V8_G4 <= 1.854:
            exchange_list = [J3B_VCC_1V8_G4, 0, err_J3B_VCC_1V8_G4_1]
        else:
            err_J3B_VCC_1V8_G4_1 += 1
            exchange_list = [J3B_VCC_1V8_G4, 1, err_J3B_VCC_1V8_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[164]

        '''J3B_TEMP_IC'''
        float_list = [reply_extend[424], reply_extend[425], reply_extend[426], reply_extend[427]]
        J3B_TEMP_IC = float_cvt(float_list)
        exchange_list = [J3B_TEMP_IC, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[165]

        '''J3B_TEMP_SW1'''
        float_list = [reply_extend[428], reply_extend[429], reply_extend[430], reply_extend[431]]
        J3B_TEMP_SW1 = float_cvt(float_list)
        exchange_list = [J3B_TEMP_SW1, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[166]

        '''J3B_TEMP_SW2'''
        float_list = [reply_extend[432], reply_extend[433], reply_extend[434], reply_extend[435]]
        J3B_TEMP_SW2 = float_cvt(float_list)
        exchange_list = [J3B_TEMP_SW2, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[167]

        '''J3B_TEMP_SW3'''
        float_list = [reply_extend[436], reply_extend[437], reply_extend[438], reply_extend[439]]
        J3B_TEMP_SW3 = float_cvt(float_list)
        exchange_list = [J3B_TEMP_SW3, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[168]

        '''J3B_TEMP_SW4'''
        float_list = [reply_extend[440], reply_extend[441], reply_extend[442], reply_extend[443]]
        J3B_TEMP_SW4 = float_cvt(float_list)
        exchange_list = [J3B_TEMP_SW4, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[169]

        '''J3B_TEMP_SW5'''
        float_list = [reply_extend[444], reply_extend[445], reply_extend[446], reply_extend[447]]
        J3B_TEMP_SW5 = float_cvt(float_list)
        exchange_list = [J3B_TEMP_SW5, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[170]

        '''J3B_TEMP_SW6'''
        float_list = [reply_extend[448], reply_extend[449], reply_extend[450], reply_extend[451]]
        J3B_TEMP_SW6 = float_cvt(float_list)
        exchange_list = [J3B_TEMP_SW6, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[171]

        '''J3B_TEMP_SW7'''
        float_list = [reply_extend[452], reply_extend[453], reply_extend[454], reply_extend[455]]
        J3B_TEMP_SW7 = float_cvt(float_list)
        exchange_list = [J3B_TEMP_SW7, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[172]

        '''J3B_TEMP_TEMP_LDO1_2'''
        float_list = [reply_extend[456], reply_extend[457], reply_extend[458], reply_extend[459]]
        J3B_TEMP_TEMP_LDO1_2 = float_cvt(float_list)
        exchange_list = [J3B_TEMP_TEMP_LDO1_2, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[173]

        '''J3B_TEMP_TEMP_LDO3_4'''
        float_list = [reply_extend[460], reply_extend[461], reply_extend[462], reply_extend[463]]
        J3B_TEMP_TEMP_LDO3_4 = float_cvt(float_list)
        exchange_list = [J3B_TEMP_TEMP_LDO3_4, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[174]

        '''J3C'''
        '''J3C_CPU_0V8_G3'''
        float_list = [reply_extend[464], reply_extend[465], reply_extend[466], reply_extend[467]]
        J3C_CPU_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3C_CPU_0V8_G3 <= 0.81:
            exchange_list = [J3C_CPU_0V8_G3, 0, err_J3C_CPU_0V8_G3_1]
        else:
            err_J3C_CPU_0V8_G3_1 += 1
            exchange_list = [J3C_CPU_0V8_G3, 1, err_J3C_CPU_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[175]

        '''J3C_DDR_0V8_G3'''
        float_list = [reply_extend[468], reply_extend[469], reply_extend[470], reply_extend[471]]
        J3C_DDR_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3C_DDR_0V8_G3 <= 0.81:
            exchange_list = [J3C_DDR_0V8_G3, 0, err_J3C_DDR_0V8_G3_1]
        else:
            err_J3C_DDR_0V8_G3_1 += 1
            exchange_list = [J3C_DDR_0V8_G3, 1, err_J3C_DDR_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[176]

        '''J3C_VDD_0V8_G4'''
        float_list = [reply_extend[472], reply_extend[473], reply_extend[474], reply_extend[475]]
        J3C_VDD_0V8_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3C_VDD_0V8_G4 <= 0.81:
            exchange_list = [J3C_VDD_0V8_G4, 0, err_J3C_VDD_0V8_G4_1]
        else:
            err_J3C_VDD_0V8_G4_1 += 1
            exchange_list = [J3C_VDD_0V8_G4, 1, err_J3C_VDD_0V8_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[177]

        '''J3C_COREAO_0V8_G2'''
        float_list = [reply_extend[476], reply_extend[477], reply_extend[478], reply_extend[479]]
        J3C_COREAO_0V8_G2 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3C_COREAO_0V8_G2 <= 0.81:
            exchange_list = [J3C_COREAO_0V8_G2, 0, err_J3C_COREAO_0V8_G2_1]
        else:
            err_J3C_COREAO_0V8_G2_1 += 1
            exchange_list = [J3C_COREAO_0V8_G2, 1, err_J3C_COREAO_0V8_G2_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[178]

        '''J3C_CNN0_0V8_G3'''
        float_list = [reply_extend[480], reply_extend[481], reply_extend[482], reply_extend[483]]
        J3C_CNN0_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3C_CNN0_0V8_G3 <= 0.81:
            exchange_list = [J3C_CNN0_0V8_G3, 0, err_J3C_CNN0_0V8_G3_1]
        else:
            err_J3C_CNN0_0V8_G3_1 += 1
            exchange_list = [J3C_CNN0_0V8_G3, 1, err_J3C_CNN0_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[179]

        '''J3C_CNN1_0V8_G3'''
        float_list = [reply_extend[484], reply_extend[485], reply_extend[486], reply_extend[487]]
        J3C_CNN1_0V8_G3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 0.79 <= J3C_CNN1_0V8_G3 <= 0.81:
            exchange_list = [J3C_CNN1_0V8_G3, 0, err_J3C_CNN1_0V8_G3_1]
        else:
            err_J3C_CNN1_0V8_G3_1 += 1
            exchange_list = [J3C_CNN1_0V8_G3, 1, err_J3C_CNN1_0V8_G3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[180]

        '''J3C_VDD_1V8_G1'''
        float_list = [reply_extend[488], reply_extend[489], reply_extend[490], reply_extend[491]]
        J3C_VDD_1V8_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.764 <= J3C_VDD_1V8_G1 <= 1.836:
            exchange_list = [J3C_VDD_1V8_G1, 0, err_J3C_VDD_1V8_G1_1]
        else:
            err_J3C_VDD_1V8_G1_1 += 1
            exchange_list = [J3C_VDD_1V8_G1, 1, err_J3C_VDD_1V8_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[181]

        '''J3C_EMMC_3V3_G1'''
        float_list = [reply_extend[492], reply_extend[493], reply_extend[494], reply_extend[495]]
        J3C_EMMC_3V3_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.201 <= J3C_EMMC_3V3_G1 <= 3.399:
            exchange_list = [J3C_EMMC_3V3_G1, 0, err_J3C_EMMC_3V3_G1_1]
        else:
            err_J3C_EMMC_3V3_G1_1 += 1
            exchange_list = [J3C_EMMC_3V3_G1, 1, err_J3C_EMMC_3V3_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[182]

        '''J3C_VDD_3V3_G1'''
        float_list = [reply_extend[496], reply_extend[497], reply_extend[498], reply_extend[499]]
        J3C_VDD_3V3_G1 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.201 <= J3C_VDD_3V3_G1 <= 3.399:
            exchange_list = [J3C_VDD_3V3_G1, 0, err_J3C_VDD_3V3_G1_1]
        else:
            err_J3C_VDD_3V3_G1_1 += 1
            exchange_list = [J3C_VDD_3V3_G1, 1, err_J3C_VDD_3V3_G1_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[183]

        '''J3C_VCC_3V3_G4'''
        float_list = [reply_extend[500], reply_extend[501], reply_extend[502], reply_extend[503]]
        J3C_VCC_3V3_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 3.201 <= J3C_VCC_3V3_G4 <= 3.399:
            exchange_list = [J3C_VCC_3V3_G4, 0, err_J3C_VCC_3V3_G4_1]
        else:
            err_J3C_VCC_3V3_G4_1 += 1
            exchange_list = [J3C_VCC_3V3_G4, 1, err_J3C_VCC_3V3_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[184]

        '''J3C_VCC_1V8_G4'''
        float_list = [reply_extend[504], reply_extend[505], reply_extend[506], reply_extend[507]]
        J3C_VCC_1V8_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 1.746 <= J3C_VCC_1V8_G4 <= 1.854:
            exchange_list = [J3C_VCC_1V8_G4, 0, err_J3C_VCC_1V8_G4_1]
        else:
            err_J3C_VCC_1V8_G4_1 += 1
            exchange_list = [J3C_VCC_1V8_G4, 1, err_J3C_VCC_1V8_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[185]

        '''J3C_TEMP_IC'''
        float_list = [reply_extend[508], reply_extend[509], reply_extend[510], reply_extend[511]]
        J3C_TEMP_IC = float_cvt(float_list)
        exchange_list = [J3C_TEMP_IC, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[186]

        '''J3C_TEMP_SW1'''
        float_list = [reply_extend[512], reply_extend[513], reply_extend[514], reply_extend[515]]
        J3C_TEMP_SW1 = float_cvt(float_list)
        exchange_list = [J3C_TEMP_SW1, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[187]

        '''J3C_TEMP_SW2'''
        float_list = [reply_extend[516], reply_extend[517], reply_extend[518], reply_extend[519]]
        J3C_TEMP_SW2 = float_cvt(float_list)
        exchange_list = [J3C_TEMP_SW2, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[188]

        '''J3C_TEMP_SW3'''
        float_list = [reply_extend[520], reply_extend[521], reply_extend[522], reply_extend[523]]
        J3C_TEMP_SW3 = float_cvt(float_list)
        exchange_list = [J3C_TEMP_SW3, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[189]

        '''J3C_TEMP_SW4'''
        float_list = [reply_extend[524], reply_extend[525], reply_extend[526], reply_extend[527]]
        J3C_TEMP_SW4 = float_cvt(float_list)
        exchange_list = [J3C_TEMP_SW4, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[190]

        '''J3C_TEMP_SW5'''
        float_list = [reply_extend[528], reply_extend[529], reply_extend[530], reply_extend[531]]
        J3C_TEMP_SW5 = float_cvt(float_list)
        exchange_list = [J3C_TEMP_SW5, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[191]

        '''J3C_TEMP_SW6'''
        float_list = [reply_extend[532], reply_extend[533], reply_extend[534], reply_extend[535]]
        J3C_TEMP_SW6 = float_cvt(float_list)
        exchange_list = [J3C_TEMP_SW6, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[192]

        '''J3C_TEMP_SW7'''
        float_list = [reply_extend[536], reply_extend[537], reply_extend[538], reply_extend[539]]
        J3C_TEMP_SW7 = float_cvt(float_list)
        exchange_list = [J3C_TEMP_SW7, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[193]

        '''J3C_TEMP_TEMP_LDO1_2'''
        float_list = [reply_extend[540], reply_extend[541], reply_extend[542], reply_extend[543]]
        J3C_TEMP_TEMP_LDO1_2 = float_cvt(float_list)
        exchange_list = [J3C_TEMP_TEMP_LDO1_2, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[194]

        '''J3C_TEMP_TEMP_LDO3_4'''
        float_list = [reply_extend[544], reply_extend[545], reply_extend[546], reply_extend[547]]
        J3C_TEMP_TEMP_LDO3_4 = float_cvt(float_list)
        exchange_list = [J3C_TEMP_TEMP_LDO3_4, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[195]

        '''I2C'''
        '''MCU_MAX20084_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[548] == 0:
            exchange_list = [reply_extend[548], 0, err_MCU_MAX20084_I2C_Status_1]
        else:
            err_MCU_MAX20084_I2C_Status_1 += 1
            exchange_list = [reply_extend[548], 1, err_MCU_MAX20084_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[196]

        '''MCU_PMIC8100_Index0_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[549] == 0:
            exchange_list = [reply_extend[549], 0, err_MCU_PMIC8100_Index0_I2C_Status_1]
        else:
            err_MCU_PMIC8100_Index0_I2C_Status_1 += 1
            exchange_list = [reply_extend[549], 1, err_MCU_PMIC8100_Index0_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[197]

        '''MCU_PMIC8100_Index1_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[550] == 0:
            exchange_list = [reply_extend[550], 0, err_MCU_PMIC8100_Index1_I2C_Status_1]
        else:
            err_MCU_PMIC8100_Index1_I2C_Status_1 += 1
            exchange_list = [reply_extend[550], 1, err_MCU_PMIC8100_Index1_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[198]

        '''MCU_PMIC8100_Index2_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[551] == 0:
            exchange_list = [reply_extend[551], 0, err_MCU_PMIC8100_Index2_I2C_Status_1]
        else:
            err_MCU_PMIC8100_Index2_I2C_Status_1 += 1
            exchange_list = [reply_extend[551], 1, err_MCU_PMIC8100_Index2_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[199]

        '''MCU_PMIC5024_Index0_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[552] == 0:
            exchange_list = [reply_extend[552], 0, err_MCU_PMIC5024_Index0_I2C_Status_1]
        else:
            err_MCU_PMIC5024_Index0_I2C_Status_1 += 1
            exchange_list = [reply_extend[552], 1, err_MCU_PMIC5024_Index0_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[200]

        '''MCU_PMIC5024_Index1_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[553] == 0:
            exchange_list = [reply_extend[553], 0, err_MCU_PMIC5024_Index1_I2C_Status_1]
        else:
            err_MCU_PMIC5024_Index1_I2C_Status_1 += 1
            exchange_list = [reply_extend[553], 1, err_MCU_PMIC5024_Index1_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[201]

        '''MCU_PMIC5024_Index2_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[554] == 0:
            exchange_list = [reply_extend[554], 0, err_MCU_PMIC5024_Index2_I2C_Status_1]
        else:
            err_MCU_PMIC5024_Index2_I2C_Status_1 += 1
            exchange_list = [reply_extend[554], 1, err_MCU_PMIC5024_Index2_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[202]

        '''J3A_MAX9296_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[555] == 0:
            exchange_list = [reply_extend[555], 0, err_J3A_MAX9296_I2C_Status_1]
        else:
            err_J3A_MAX9296_I2C_Status_1 += 1
            exchange_list = [reply_extend[555], 1, err_J3A_MAX9296_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[203]

        '''J3A_MAX96717_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[556] == 0:
            exchange_list = [reply_extend[556], 0, err_J3A_MAX96717_I2C_Status_1]
        else:
            err_J3A_MAX96717_I2C_Status_1 += 1
            exchange_list = [reply_extend[556], 1, err_J3A_MAX96717_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[204]

        '''J3A_MAX20089_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[557] == 0:
            exchange_list = [reply_extend[557], 0, err_J3A_MAX20089_I2C_Status_1]
        else:
            err_J3A_MAX20089_I2C_Status_1 += 1
            exchange_list = [reply_extend[557], 1, err_J3A_MAX20089_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[205]

        '''J3B_MAX9296_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[558] == 0:
            exchange_list = [reply_extend[558], 0, err_J3B_MAX9296_I2C_Status_1]
        else:
            err_J3B_MAX9296_I2C_Status_1 += 1
            exchange_list = [reply_extend[558], 1, err_J3B_MAX9296_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[206]

        '''J3B_MAX96712_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[559] == 0:
            exchange_list = [reply_extend[559], 0, err_J3B_MAX96712_I2C_Status_1]
        else:
            err_J3B_MAX96712_I2C_Status_1 += 1
            exchange_list = [reply_extend[559], 1, err_J3B_MAX96712_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[207]

        '''J3B_MAX20089_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[560] == 0:
            exchange_list = [reply_extend[560], 0, err_J3B_MAX20089_I2C_Status_1]
        else:
            err_J3B_MAX20089_I2C_Status_1 += 1
            exchange_list = [reply_extend[560], 1, err_J3B_MAX20089_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[208]

        '''J3B_MAX20087_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[561] == 0:
            exchange_list = [reply_extend[561], 0, err_J3B_MAX20087_I2C_Status_1]
        else:
            err_J3B_MAX20087_I2C_Status_1 += 1
            exchange_list = [reply_extend[561], 1, err_J3B_MAX20087_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[209]

        '''J3C_MAX96712_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[562] == 0:
            exchange_list = [reply_extend[562], 0, err_J3C_MAX96712_I2C_Status_1]
        else:
            err_J3C_MAX96712_I2C_Status_1 += 1
            exchange_list = [reply_extend[562], 1, err_J3C_MAX96712_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[210]

        '''J3C_MAX96717_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[563] == 0:
            exchange_list = [reply_extend[563], 0, err_J3C_MAX96717_I2C_Status_1]
        else:
            err_J3C_MAX96717_I2C_Status_1 += 1
            exchange_list = [reply_extend[563], 1, err_J3C_MAX96717_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[211]

        '''J3C_MAX20087_I2C_Status'''
        exchange_list = [0, 0, 0]
        if reply_extend[564] == 0:
            exchange_list = [reply_extend[564], 0, err_J3C_MAX20087_I2C_Status_1]
        else:
            err_J3C_MAX20087_I2C_Status_1 += 1
            exchange_list = [reply_extend[564], 1, err_J3C_MAX20087_I2C_Status_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[212]

        '''J3A_Front_Camera_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[565] == 0:
            exchange_list = [reply_extend[565], 0, err_J3A_Front_Camera_max9295_ID_1]
        else:
            err_J3A_Front_Camera_max9295_ID_1 += 1
            exchange_list = [reply_extend[565], 1, err_J3A_Front_Camera_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[213]

        '''J3B_Rear_Camera_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[566] == 0:
            exchange_list = [reply_extend[566], 0, err_J3B_Rear_Camera_max9295_ID_1]
        else:
            err_J3B_Rear_Camera_max9295_ID_1 += 1
            exchange_list = [reply_extend[566], 1, err_J3B_Rear_Camera_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[214]

        '''J3B_Side_Camera_0_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[567] == 0:
            exchange_list = [reply_extend[567], 0, err_J3B_Side_Camera_0_max9295_ID_1]
        else:
            err_J3B_Side_Camera_0_max9295_ID_1 += 1
            exchange_list = [reply_extend[567], 1, err_J3B_Side_Camera_0_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[215]

        '''J3B_Side_Camera_1_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[568] == 0:
            exchange_list = [reply_extend[568], 0, err_J3B_Side_Camera_1_max9295_ID_1]
        else:
            err_J3B_Side_Camera_1_max9295_ID_1 += 1
            exchange_list = [reply_extend[568], 1, err_J3B_Side_Camera_1_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[216]

        '''J3B_Side_Camera_2_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[569] == 0:
            exchange_list = [reply_extend[569], 0, err_J3B_Side_Camera_2_max9295_ID_1]
        else:
            err_J3B_Side_Camera_2_max9295_ID_1 += 1
            exchange_list = [reply_extend[569], 1, err_J3B_Side_Camera_2_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[217]

        '''J3B_Side_Camera_3_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[570] == 0:
            exchange_list = [reply_extend[570], 0, err_J3B_Side_Camera_3_max9295_ID_1]
        else:
            err_J3B_Side_Camera_3_max9295_ID_1 += 1
            exchange_list = [reply_extend[570], 1, err_J3B_Side_Camera_3_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[218]

        '''J3C_Surround_Camera_0_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[571] == 0:
            exchange_list = [reply_extend[571], 0, err_J3C_Surround_Camera_0_max9295_ID_1]
        else:
            err_J3C_Surround_Camera_0_max9295_ID_1 += 1
            exchange_list = [reply_extend[571], 1, err_J3C_Surround_Camera_0_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[219]

        '''J3C_Surround_Camera_1_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[572] == 0:
            exchange_list = [reply_extend[572], 0, err_J3C_Surround_Camera_1_max9295_ID_1]
        else:
            err_J3C_Surround_Camera_1_max9295_ID_1 += 1
            exchange_list = [reply_extend[572], 1, err_J3C_Surround_Camera_1_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[220]

        '''J3C_Surround_Camera_2_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[573] == 0:
            exchange_list = [reply_extend[573], 0, err_J3C_Surround_Camera_2_max9295_ID_1]
        else:
            err_J3C_Surround_Camera_2_max9295_ID_1 += 1
            exchange_list = [reply_extend[573], 1, err_J3C_Surround_Camera_2_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[221]

        '''J3C_Surround_Camera_3_max9295_ID'''
        exchange_list = [0, 0, 0]
        if reply_extend[574] == 0:
            exchange_list = [reply_extend[574], 0, err_J3C_Surround_Camera_3_max9295_ID_1]
        else:
            err_J3C_Surround_Camera_3_max9295_ID_1 += 1
            exchange_list = [reply_extend[574], 1, err_J3C_Surround_Camera_3_max9295_ID_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[222]


        '''J3A_TEMP'''
        float_list = [reply_extend[584], reply_extend[585], reply_extend[586], reply_extend[587]]
        J3A_TEMP = float_cvt(float_list)
        exchange_list = [J3A_TEMP, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[223]

        '''J3B_TEMP'''
        float_list = [reply_extend[588], reply_extend[589], reply_extend[590], reply_extend[591]]
        J3B_TEMP = float_cvt(float_list)
        exchange_list = [J3B_TEMP, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[224]

        '''J3C_TEMP'''
        float_list = [reply_extend[592], reply_extend[593], reply_extend[594], reply_extend[595]]
        J3C_TEMP = float_cvt(float_list)
        exchange_list = [J3C_TEMP, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[225]

        '''J3A_F9K_TEMP'''
        float_list = [reply_extend[596], reply_extend[597], reply_extend[598], reply_extend[599]]
        J3A_F9K_TEMP = float_cvt(float_list)
        exchange_list = [J3A_F9K_TEMP, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[226]



        '''阈值未定'''
        '''ADC_SW1_3V3'''
        float_list = [reply_extend[600], reply_extend[601], reply_extend[602], reply_extend[603]]
        ADC_SW1_3V3 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.566 <= ADC_SW1_3V3 <= 5.528:
            exchange_list = [ADC_SW1_3V3, 0, err_ADC_SW1_3V3_1]
        else:
            err_ADC_SW1_3V3_1 += 1
            exchange_list = [ADC_SW1_3V3, 1, err_ADC_SW1_3V3_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[227]

        '''ADC_SW1_1V8'''
        float_list = [reply_extend[604], reply_extend[605], reply_extend[606], reply_extend[607]]
        ADC_SW1_1V8 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.566 <= ADC_SW1_1V8 <= 5.528:
            exchange_list = [ADC_SW1_1V8, 0, err_ADC_SW1_1V8_1]
        else:
            err_ADC_SW1_1V8_1 += 1
            exchange_list = [ADC_SW1_1V8, 1, err_ADC_SW1_1V8_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[228]

        '''ADC_SW1_1V2'''
        float_list = [reply_extend[608], reply_extend[609], reply_extend[610], reply_extend[611]]
        ADC_SW1_1V2 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.566 <= ADC_SW1_1V2 <= 5.528:
            exchange_list = [ADC_SW1_1V2, 0, err_ADC_SW1_1V2_1]
        else:
            err_ADC_SW1_1V2_1 += 1
            exchange_list = [ADC_SW1_1V2, 1, err_ADC_SW1_1V2_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[229]

        '''ADC_J3A_VDD_0V8A_G4'''
        float_list = [reply_extend[612], reply_extend[613], reply_extend[614], reply_extend[615]]
        ADC_J3A_VDD_0V8A_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.566 <= ADC_J3A_VDD_0V8A_G4 <= 5.528:
            exchange_list = [ADC_J3A_VDD_0V8A_G4, 0, err_ADC_J3A_VDD_0V8A_G4_1]
        else:
            err_ADC_J3A_VDD_0V8A_G4_1 += 1
            exchange_list = [ADC_J3A_VDD_0V8A_G4, 1, err_ADC_J3A_VDD_0V8A_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[230]

        '''ADC_J3B_VDD_0V8A_G4'''
        float_list = [reply_extend[616], reply_extend[617], reply_extend[618], reply_extend[619]]
        ADC_J3B_VDD_0V8A_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.566 <= ADC_J3B_VDD_0V8A_G4 <= 5.528:
            exchange_list = [ADC_J3B_VDD_0V8A_G4, 0, err_ADC_J3B_VDD_0V8A_G4_1]
        else:
            err_ADC_J3B_VDD_0V8A_G4_1 += 1
            exchange_list = [ADC_J3B_VDD_0V8A_G4, 1, err_ADC_J3B_VDD_0V8A_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[231]

        '''ADC_J3C_VDD_0V8A_G4'''
        float_list = [reply_extend[620], reply_extend[621], reply_extend[622], reply_extend[623]]
        ADC_J3C_VDD_0V8A_G4 = float_cvt(float_list)
        exchange_list = [0, 0, 0]
        if 4.566 <= ADC_J3C_VDD_0V8A_G4 <= 5.528:
            exchange_list = [ADC_J3C_VDD_0V8A_G4, 0, err_ADC_J3C_VDD_0V8A_G4_1]
        else:
            err_ADC_J3C_VDD_0V8A_G4_1 += 1
            exchange_list = [ADC_J3C_VDD_0V8A_G4, 1, err_ADC_J3C_VDD_0V8A_G4_1]
        data_analysed_1.append(exchange_list)  # data_analysed_1[232]


        '''USS1_PWM_PERIOD'''
        float_list = [reply_extend[624], reply_extend[625]]
        USS1_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS1_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[233]

        '''USS1_PWM_DUTY'''
        float_list = [reply_extend[626], reply_extend[627]]
        USS1_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS1_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[234]

        '''USS2_PWM_PERIOD'''
        float_list = [reply_extend[628], reply_extend[629]]
        USS2_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS2_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[235]

        '''USS2_PWM_DUTY'''
        float_list = [reply_extend[630], reply_extend[631]]
        USS2_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS2_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[236]

        '''USS3_PWM_PERIOD'''
        float_list = [reply_extend[632], reply_extend[633]]
        USS3_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS3_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[237]

        '''USS3_PWM_DUTY'''
        float_list = [reply_extend[634], reply_extend[635]]
        USS3_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS3_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[238]

        '''USS4_PWM_PERIOD'''
        float_list = [reply_extend[636], reply_extend[637]]
        USS4_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS4_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[239]

        '''USS4_PWM_DUTY'''
        float_list = [reply_extend[638], reply_extend[639]]
        USS4_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS4_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[240]

        '''USS5_PWM_PERIOD'''
        float_list = [reply_extend[640], reply_extend[641]]
        USS5_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS5_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[241]

        '''USS5_PWM_DUTY'''
        float_list = [reply_extend[642], reply_extend[643]]
        USS5_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS5_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[242]

        '''USS6_PWM_PERIOD'''
        float_list = [reply_extend[644], reply_extend[645]]
        USS6_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS6_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[243]

        '''USS6_PWM_DUTY'''
        float_list = [reply_extend[646], reply_extend[647]]
        USS6_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS6_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[244]

        '''USS7_PWM_PERIOD'''
        float_list = [reply_extend[648], reply_extend[649]]
        USS7_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS7_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[245]

        '''USS7_PWM_DUTY'''
        float_list = [reply_extend[650], reply_extend[651]]
        USS7_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS7_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[246]

        '''USS8_PWM_PERIOD'''
        float_list = [reply_extend[652], reply_extend[653]]
        USS8_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS8_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[247]

        '''USS8_PWM_DUTY'''
        float_list = [reply_extend[654], reply_extend[655]]
        USS8_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS8_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[248]

        '''USS9_PWM_PERIOD'''
        float_list = [reply_extend[656], reply_extend[657]]
        USS9_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS9_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[249]

        '''USS9_PWM_DUTY'''
        float_list = [reply_extend[658], reply_extend[659]]
        USS9_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS9_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[250]

        '''USS10_PWM_PERIOD'''
        float_list = [reply_extend[660], reply_extend[661]]
        USS10_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS10_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[251]

        '''USS10_PWM_DUTY'''
        float_list = [reply_extend[662], reply_extend[663]]
        USS10_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS10_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[252]

        '''USS11_PWM_PERIOD'''
        float_list = [reply_extend[664], reply_extend[665]]
        USS11_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS11_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[253]

        '''USS11_PWM_DUTY'''
        float_list = [reply_extend[666], reply_extend[667]]
        USS11_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS11_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[254]

        '''USS12_PWM_PERIOD'''
        float_list = [reply_extend[668], reply_extend[669]]
        USS12_PWM_PERIOD = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS12_PWM_PERIOD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[255]

        '''USS12_PWM_DUTY'''
        float_list = [reply_extend[670], reply_extend[671]]
        USS12_PWM_DUTY = int.from_bytes(float_list, byteorder='little', signed=False)
        exchange_list = [USS12_PWM_DUTY, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[256]


        '''J3A_Front_Camera_Supply_AD'''
        int_list = [reply_extend[716], reply_extend[717], reply_extend[718], reply_extend[719]]
        J3A_Front_Camera_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3A_Front_Camera_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[257]

        '''J3B_Rear_Camera_Supply_AD'''
        int_list = [reply_extend[720], reply_extend[721], reply_extend[722], reply_extend[723]]
        J3B_Rear_Camera_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3B_Rear_Camera_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[258]

        '''J3B_Side_Camera_0_Supply_AD'''
        int_list = [reply_extend[724], reply_extend[725], reply_extend[726], reply_extend[727]]
        J3B_Side_Camera_0_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3B_Side_Camera_0_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[259]

        '''J3B_Side_Camera_1_Supply_AD'''
        int_list = [reply_extend[728], reply_extend[729], reply_extend[730], reply_extend[731]]
        J3B_Side_Camera_1_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3B_Side_Camera_1_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[260]

        '''J3B_Side_Camera_2_Supply_AD'''
        int_list = [reply_extend[732], reply_extend[733], reply_extend[734], reply_extend[735]]
        J3B_Side_Camera_2_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3B_Side_Camera_2_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[261]

        '''J3B_Side_Camera_3_Supply_AD'''
        int_list = [reply_extend[736], reply_extend[737], reply_extend[738], reply_extend[739]]
        J3B_Side_Camera_3_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3B_Side_Camera_3_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[262]

        '''J3C_Surround_Camera_0_Supply_AD'''
        int_list = [reply_extend[740], reply_extend[741], reply_extend[742], reply_extend[743]]
        J3C_Surround_Camera_0_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3C_Surround_Camera_0_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[263]

        '''J3C_Surround_Camera_1_Supply_AD'''
        int_list = [reply_extend[744], reply_extend[745], reply_extend[746], reply_extend[747]]
        J3C_Surround_Camera_1_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3C_Surround_Camera_1_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[264]

        '''J3C_Surround_Camera_2_Supply_AD'''
        int_list = [reply_extend[748], reply_extend[749], reply_extend[750], reply_extend[751]]
        J3C_Surround_Camera_2_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3C_Surround_Camera_2_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[265]

        '''J3C_Surround_Camera_3_Supply_AD'''
        int_list = [reply_extend[752], reply_extend[753], reply_extend[754], reply_extend[755]]
        J3C_Surround_Camera_3_Supply_AD = int.from_bytes(int_list, byteorder='little', signed=True)
        exchange_list = [J3C_Surround_Camera_3_Supply_AD, 0, 0]
        data_analysed_1.append(exchange_list)  # data_analysed_1[266]

    else:
        pass

def listen_udp_1():
    sniff(filter = "ip src " + ip_ecu + " and ip dst " + ip_interface_1 + " and udp ", count = 0, store = 1, prn = packet_rev_udp_1, iface = net_if_1)

def listen_udp_2():
    sniff(filter = "ip src " + ip_ecu + " and ip dst " + ip_interface_2 + " and udp ", count = 0, store = 1, prn = packet_rev_udp_2, iface = net_if_2)

def listen_udp_3():
    sniff(filter = "ip src " + ip_ecu + " and ip dst " + ip_interface_3 + " and udp ", count = 0, store = 1, prn = packet_rev_udp_3, iface = net_if_3)

def listen_udp_4():
    sniff(filter = "ip src " + ip_ecu + " and ip dst " + ip_interface_4 + " and udp ", count = 0, store = 1, prn = packet_rev_udp_4, iface = net_if_4)

def listen_udp_5():
    sniff(filter = "ip src " + ip_ecu + " and ip dst " + ip_interface_5 + " and udp ", count = 0, store = 1, prn = packet_rev_udp_5, iface = net_if_5)

def listen_udp_6():
    sniff(filter = "ip src " + ip_ecu + " and ip dst " + ip_interface_6 + " and udp ", count = 0, store = 1, prn = packet_rev_udp_6, iface = net_if_6)



'''create folder and excel'''
if os.path.exists('c:\\validation_log') == False:
    os.mkdir('c:\\validation_log')

val = openpyxl.Workbook()
sheet = val.active
sheet.title = 'validation'
first_row = ['Time', 'RollingCounter', 'CAN0_Status', 'CAN1_Status', \
             'CAN2_Status', 'CAN3_Status', 'CAN4_Status',\
             'Reserved','Reserved','Reserved','Reserved','Reserved',\
             'USS1_Vout', 'USS2_Vout', 'USS3_Vout', 'ADC_VBAT_12V', 'ADC_VBoost_12V', 'ADC_TJA1043_INH', 'ADC_IG_12V',\
             'ADC_SWITCH_INH', 'ADC_SYS_1V8A', 'ADC_SYS_1V8B', 'ADC_SYS_1V2A', 'ADC_SYS_1V2B', 'ADC_SYS_3V3', 'ADC_J3A_COREPD_0V8_G3', \
             'ADC_J3A_VDDQDDR_1V1_G1', 'ADC_J3B_COREPD_0V8_G3', 'ADC_J3B_VDDQDDR_1V1_G1', 'ADC_J3C_COREPD_0V8_G3', 'ADC_J3C_VDDQDDR_1V1_G1',\
             'ADC_GPSANT_PWR_AMUX', 'ADC_GPS_3V3', 'ADC_SYS_5V', 'ADC_ETH_0V9', 'ADC_HWID', 'ADC_Proj_ID','J3A_Alive_Status', \
             'J3B_Alive_Status', 'J3C_Alive_Status', 'MCU_J3A_SPI_Status', 'MCU_J3B_SPI_Status', 'MCU_J3C_SPI_Status',\
             'MCU_J3A_Ethernet_TimeOut_Staus', 'MCU_J3B_Ethernet_TimeOut_Staus', 'MCU_J3C_Ethernet_TimeOut_Staus', \
             'MCU_J3A_Ethernet_RollingCounter_Staus', 'MCU_J3B_Ethernet_RollingCounter_Staus', 'MCU_J3C_Ethernet_RollingCounter_Staus',\
             'J3A_CPU_Load', 'J3B_CPU_Load', 'J3C_CPU_Load', 'J3A_eMMC_Status', 'J3B_eMMC_Status', 'J3C_eMMC_Status',\
             'J3C_Surround_Camera_Input0_Link_Lock', 'J3C_Surround_Camera_Input0_Lock', 'J3C_Surround_Camera_Input0_Fps', \
             'J3C_Surround_Camera_Input0_Crc', 'J3C_Surround_Camera_Input0_OpenLoad', 'J3C_Surround_Camera_Input0_Short', \
             'J3C_Surround_Camera_Input1_Link_Lock', 'J3C_Surround_Camera_Input1_Lock', 'J3C_Surround_Camera_Input1_Fps',\
             'J3C_Surround_Camera_Input1_Crc', 'J3C_Surround_Camera_Input1_OpenLoad', 'J3C_Surround_Camera_Input1_Short',\
             'J3C_Surround_Camera_Input2_Link_Lock', 'J3C_Surround_Camera_Input2_Lock', 'J3C_Surround_Camera_Input2_Fps',\
             'J3C_Surround_Camera_Input2_Crc', 'J3C_Surround_Camera_Input2_OpenLoad', 'J3C_Surround_Camera_Input2_Short',\
             'J3C_Surround_Camera_Input3_Link_Lock', 'J3C_Surround_Camera_Input3_Lock', 'J3C_Surround_Camera_Input3_Fps',\
             'J3C_Surround_Camera_Input3_Crc', 'J3C_Surround_Camera_Input3_OpenLoad', 'J3C_Surround_Camera_Input3_Short',\
             'J3C_Surround_Camera_Output_Link_Lock', 'J3C_Surround_Camera_Output_Lock', 'J3C_Surround_Camera_Output_Fps',\
             'J3C_Surround_Camera_Output_Crc', 'J3A_Front_Camera_Input_Link_Lock', 'J3A_Front_Camera_Input_Lock', \
             'J3A_Front_Camera_Input_Fps', 'J3A_Front_Camera_Input_Crc', 'J3A_Front_Camera_Input_OpenLoad', 'J3A_Front_Camera_Input_Short',\
             'J3A_Front_Camera_Output_Link_Lock', 'J3A_Front_Camera_Output_Lock','J3A_Front_Camera_Output_Fps', 'J3A_Front_Camera_Output_Crc',\
             'J3B_Rear_Camera_Input_Link_Lock', 'J3B_Rear_Camera_Input_Lock', 'J3B_Rear_Camera_Input_Fps', 'J3B_Rear_Camera_Input_Crc',\
             'J3B_Rear_Camera_Input_OpenLoad', 'J3B_Rear_Camera_Input_Short', 'J3B_Side_Camera_Input0_Link_Lock', 'J3B_Side_Camera_Input0_Lock',\
             'J3B_Side_Camera_Input0_Fps', 'J3B_Side_Camera_Input0_Crc', 'J3B_Side_Camera_Input0_OpenLoad', 'J3B_Side_Camera_Input0_Short', \
             'J3B_Side_Camera_Input1_Link_Lock', 'J3B_Side_Camera_Input1_Lock', 'J3B_Side_Camera_Input1_Fps', 'J3B_Side_Camera_Input1_Crc',\
             'J3B_Side_Camera_Input1_OpenLoad', 'J3B_Side_Camera_Input1_Short', 'J3B_Side_Camera_Input2_Link_Lock', 'J3B_Side_Camera_Input2_Lock', \
             'J3B_Side_Camera_Input2_Fps', 'J3B_Side_Camera_Input2_Crc', 'J3B_Side_Camera_Input2_OpenLoad', 'J3B_Side_Camera_Input2_Short', \
             'J3B_Side_Camera_Input3_Link_Lock','J3B_Side_Camera_Input3_Lock', 'J3B_Side_Camera_Input3_Fps', 'J3B_Side_Camera_Input3_Crc', \
             'J3B_Side_Camera_Input3_OpenLoad', 'J3B_Side_Camera_Input3_Short', 'J3A_F9K_GNSS_Status', 'MCU_F9K_GNSS_OpenLoad', \
             'MCU_F9K_GNSS_Short', 'MCU_F9K_GNSS_Supply', 'J3A_F9K_IMU_Z',  'J3A_F9K_CN_Value', 'MCU_TEMP',\
             'J3A_5024_TEMP', 'J3B_5024_TEMP', 'J3C_5024_TEMP', 'PCB_TEMP', 'J3A_CPU_0V8_G3', 'J3A_DDR_0V8_G3', 'J3A_VDD_0V8_G4',\
             'J3A_COREAO_0V8_G2', 'J3A_CNN0_0V8_G3', 'J3A_CNN1_0V8_G3', 'J3A_VDD_1V8_G1', 'J3A_EMMC_3V3_G1', 'J3A_VDD_3V3_G1', \
             'J3A_VCC_3V3_G4','J3A_VCC_1V8_G4', 'J3A_TEMP_IC', 'J3A_TEMP_SW1', 'J3A_TEMP_SW2', 'J3A_TEMP_SW3', 'J3A_TEMP_SW4', 'J3A_TEMP_SW5',\
             'J3A_TEMP_SW6', 'J3A_TEMP_SW7', 'J3A_TEMP_TEMP_LDO1_2', 'J3A_TEMP_TEMP_LDO3_4', 'J3B_CPU_0V8_G3', 'J3B_DDR_0V8_G3',\
             'J3B_VDD_0V8_G4', 'J3B_COREAO_0V8_G2', 'J3B_CNN0_0V8_G3', 'J3B_CNN1_0V8_G3', 'J3B_VDD_1V8_G1', 'J3B_EMMC_3V3_G1', 'J3B_VDD_3V3_G1',\
             'J3B_VCC_3V3_G4', 'J3B_VCC_1V8_G4', 'J3B_TEMP_IC', 'J3B_TEMP_SW1', 'J3B_TEMP_SW2', 'J3B_TEMP_SW3', 'J3B_TEMP_SW4',\
             'J3B_TEMP_SW5', 'J3B_TEMP_SW6', 'J3B_TEMP_SW7', 'J3B_TEMP_TEMP_LDO1_2', 'J3B_TEMP_TEMP_LDO3_4', 'J3C_CPU_0V8_G3', 'J3C_DDR_0V8_G3', \
             'J3C_VDD_0V8_G4', 'J3C_COREAO_0V8_G2', 'J3C_CNN0_0V8_G3', 'J3C_CNN1_0V8_G3', 'J3C_VDD_1V8_G1', 'J3C_EMMC_3V3_G1', 'J3C_VDD_3V3_G1',\
             'J3C_VCC_3V3_G4', 'J3C_VCC_1V8_G4', 'J3C_TEMP_IC', 'J3C_TEMP_SW1', 'J3C_TEMP_SW2', 'J3C_TEMP_SW3', 'J3C_TEMP_SW4', \
             'J3C_TEMP_SW5', 'J3C_TEMP_SW6', 'J3C_TEMP_SW7', 'J3C_TEMP_TEMP_LDO1_2', 'J3C_TEMP_TEMP_LDO3_4', 'MCU_MAX20084_I2C_Status',\
             'MCU_PMIC8100_Index0_I2C_Status', 'MCU_PMIC8100_Index1_I2C_Status', 'MCU_PMIC8100_Index2_I2C_Status', 'MCU_PMIC5024_Index0_I2C_Status',\
             'MCU_PMIC5024_Index1_I2C_Status', 'MCU_PMIC5024_Index2_I2C_Status', 'J3A_MAX9296_I2C_Status', 'J3A_MAX96717_I2C_Status',\
             'J3A_MAX20089_I2C_Status', 'J3B_MAX9296_I2C_Status', 'J3B_MAX96712_I2C_Status', 'J3B_MAX20089_I2C_Status', 'J3B_MAX20087_I2C_Status',\
             'J3C_MAX96712_I2C_Status', 'J3C_MAX96717_I2C_Status', 'J3C_MAX20087_I2C_Status',\
             'J3A_Front_Camera_max9295_ID', 'J3B_Rear_Camera_max9295_ID', 'J3B_Side_Camera_0_max9295_ID', 'J3B_Side_Camera_1_max9295_ID', \
             'J3B_Side_Camera_2_max9295_ID','J3B_Side_Camera_3_max9295_ID','J3C_Surround_Camera_0_max9295_ID','J3C_Surround_Camera_1_max9295_ID',\
             'J3C_Surround_Camera_2_max9295_ID','J3C_Surround_Camera_3_max9295_ID',\
             'J3A_TEMP','J3B_TEMP', 'J3C_TEMP', 'J3A_F9K_TEMP',\
             'ADC_SW1_3V3', 'ADC_SW1_1V8', 'ADC_SW1_1V2', 'ADC_J3A_VDD_0V8A_G4', 'ADC_J3B_VDD_0V8A_G4', 'ADC_J3C_VDD_0V8A_G4',\
             'USS1_PWM_PERIOD','USS1_PWM_DUTY','USS2_PWM_PERIOD','USS2_PWM_DUTY','USS3_PWM_PERIOD',\
             'USS3_PWM_DUTY','USS4_PWM_PERIOD','USS4_PWM_DUTY','USS5_PWM_PERIOD',\
             'USS5_PWM_DUTY','USS6_PWM_PERIOD','USS6_PWM_DUTY','USS7_PWM_PERIOD',\
             'USS7_PWM_DUTY','USS8_PWM_PERIOD','USS8_PWM_DUTY','USS9_PWM_PERIOD',\
             'USS9_PWM_DUTY','USS10_PWM_PERIOD','USS10_PWM_DUTY','USS11_PWM_PERIOD',\
             'USS11_PWM_DUTY','USS12_PWM_PERIOD','USS12_PWM_DUTY',\
             'J3A_Front_Camera_Supply_AD', 'J3B_Rear_Camera_Supply_AD', 'J3B_Side_Camera_0_Supply_AD', 'J3B_Side_Camera_1_Supply_AD', \
             'J3B_Side_Camera_2_Supply_AD', 'J3B_Side_Camera_3_Supply_AD', 'J3C_Surround_Camera_0_Supply_AD', 'J3C_Surround_Camera_1_Supply_AD', \
             'J3C_Surround_Camera_2_Supply_AD', 'J3C_Surround_Camera_3_Supply_AD']
column_number = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',\
                 'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT',\
                 'AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN',\
                 'BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC','CD','CE','CF','CG','CH',\
                 'CI','CJ','CK','CL','CM','CN','CO','CP','CQ','CR','CS','CT','CU','CV','CW','CX','CY','CZ','DA','DB',\
                 'DC','DD','DE','DF','DG','DH','DI','DJ','DK','DL','DM','DN','DO','DP','DQ','DR','DS','DT','DU','DV',\
                 'DW','DX','DY','DZ','EA','EB','EC','ED','EE','EF','EG','EH','EI','EJ','EK','EL','EM','EN','EO','EP',\
                 'EQ','ER','ES','ET','EU','EV','EW','EX','EY','EZ','FA','FB','FC','FD','FE','FF','FG','FH','FI','FJ',\
                 'FK','FL','FM','FN','FO','FP','FQ','FR','FS','FT','FU','FV','FW','FX','FY','FZ','GA','GB','GC','GD',\
                 'GE','GF','GG','GH','GI','GJ','GK','GL','GM','GN','GO','GP','GQ','GR','GS','GT','GU','GV','GW','GX',\
                 'GY','GZ','HA','HB','HC','HD','HE','HF','HG','HH','HI','HJ','HK','HL','HM','HN','HO','HP','HQ','HR',\
                 'HS','HT','HU','HV','HW','HX','HY','HZ','IA','IB','IC','ID','IE','IF','IG','IH','II','IJ','IK','IL',\
                 'IM','IN','IO','IP','IQ','IR','IS','IT','IU','IV','IW','IX','IY','IZ', 'JA','JB','JC','JD','JE','JF',\
                 'JG','JH','JI','JJ','JK','JL','JM','JN','JO','JP','JQ','JR','JS','JT','JU','JV','JW','JX','JY','JZ', \
                 'KA', 'KB', 'KC', 'KD', 'KE', 'KF','KG', 'KH', 'KI', 'KJ', 'KK', 'KL', 'KM', 'KN', 'KO', 'KP', 'KQ', \
                 'KR', 'KS', 'KT', 'KU', 'KV', 'KW', 'KX', 'KY','KZ', 'LA', 'LB', 'LC', 'LD', 'LE', 'LF','LG', 'LH', \
                 'LI', 'LJ', 'LK', 'LL', 'LM', 'LN', 'LO', 'LP', 'LQ']
sheet.column_dimensions['A'].width = 20
for i in column_number:
    sheet.column_dimensions[i].width = 20
sheet.append(first_row)
file_name = 'c:\\validation_log\\Validation_TestLog_%s.xlsx' % (time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()))
val.save(file_name)

ping_res_1 = ping(ip_interface_1, ip_ecu)
if ping_res_1:
    udp_sock_1 = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    udp_sock_1.bind((ip_interface_1, port_1))
    thread_1 = threading.Thread(target = listen_udp_1)
    thread_1.setDaemon(True)
    thread_1.start()
else:
    print('未PING通网卡1所连接的ECU，请检查连接是否正确！')
    time.sleep(2)


# '''Establish a TCP connection with incubator'''
# tcp_sk = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
# tcp_sk.bind((ip_tester, port_tester))
# tcp_sk.connect((ip_incubator, port_incubator))

'''tab class'''
class MY_GUI():
    def __init__(self,tab):
        self.tab = tab
    def execution_1(self):
        pass
    def init_tab(self):
        global ping_res_1, net_if_1, ip_interface_1
        Label(self.tab, text="网卡 名称：").place(x=1060, y=600)
        Label(self.tab, text=net_if_1).place(x=1160, y=600)
        Label(self.tab, text="网卡 IP：").place(x=1060, y=620)
        Label(self.tab, text=ip_interface_1).place(x=1160, y=620)
        Label(self.tab, text="ECU IP：").place(x=1060, y=640)
        Label(self.tab, text=ip_ecu).place(x=1160, y=640)
        Label(self.tab, text="网卡 MAC：").place(x=1060, y=660)
        Label(self.tab, text=mac_ecu_1).place(x=1160, y=660)
        # Button(self.tab, text="连接", command=self.execution_1).place(x=1060, y=820)

        '''CAN Status and USS_PWR'''
        self.CAN_Status = tk.LabelFrame(self.tab, text="CAN and USS_PWR", padx=10, pady=10)
        self.CAN_Status.place(x=0, y=20)
        self.ybar = Scrollbar(self.CAN_Status, orient='vertical')
        self.tree = ttk.Treeview(self.CAN_Status, yscrollcommand=self.ybar.set, show="headings", height=8)
        self.ybar['command'] = self.tree.yview
        self.tree["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree.column("Name", width=100)
        self.tree.column("Vulue", width= 50)
        self.tree.column("ErrCnt", width= 50)
        self.tree.heading("Name", text="Name")
        self.tree.heading("Vulue", text="Vulue")
        self.tree.heading("ErrCnt", text="ErrCnt")
        self.tree.insert("", 0, values=("CAN0_Status", "0", "0"))
        self.tree.insert("", 1, values=("CAN1_Status", "0", "0"))
        self.tree.insert("", 2, values=("CAN2_Status", "0", "0"))
        self.tree.insert("", 3, values=("CAN3_Status", "0", "0"))
        self.tree.insert("", 4, values=("CAN4_Status", "0", "0"))
        self.tree.insert("", 5, values=("USS1_Vout", "0", "0"))
        self.tree.insert("", 6, values=("USS2_Vout", "0", "0"))
        self.tree.insert("", 7, values=("USS3_Vout", "0", "0"))
        self.tree.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')


        '''ADC'''
        self.ADC = tk.LabelFrame(self.tab, text="ADC", padx=10, pady=10)
        self.ADC.place(x=0, y=255)
        self.ybar = Scrollbar(self.ADC, orient='vertical')
        self.tree_2 = ttk.Treeview(self.ADC, yscrollcommand=self.ybar.set, show="headings", height=22)
        self.ybar['command'] = self.tree.yview
        self.tree_2["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_2.column("Name", width=100)
        self.tree_2.column("Vulue", width=50)
        self.tree_2.column("ErrCnt", width=50)
        self.tree_2.heading("Name", text="Name")
        self.tree_2.heading("Vulue", text="Vulue")
        self.tree_2.heading("ErrCnt", text="ErrCnt")
        self.tree_2.insert("", 0, values=("VBAT_12V", "0", "0"))
        self.tree_2.insert("", 1, values=("VBoost_12V", "0", "0"))
        self.tree_2.insert("", 2, values=("TJA1043_INH", "0", "0"))
        self.tree_2.insert("", 3, values=("IG_12V", "0", "0"))
        self.tree_2.insert("", 4, values=("SWITCH_INH", "0", "0"))
        self.tree_2.insert("", 5, values=("SYS_1V8A", "0", "0"))
        self.tree_2.insert("", 6, values=("SYS_1V8B", "0", "0"))
        self.tree_2.insert("", 7, values=("SYS_1V2A", "0", "0"))
        self.tree_2.insert("", 8, values=("SYS_1V2B", "0", "0"))
        self.tree_2.insert("", 9, values=("SYS_3V3", "0", "0"))
        self.tree_2.insert("", 10, values=("J3A_COREPD_0V8_G3", "0", "0"))
        self.tree_2.insert("", 11, values=("J3A_VDDQDDR_1V1_G1", "0", "0"))
        self.tree_2.insert("", 12, values=("J3B_COREPD_0V8_G3", "0", "0"))
        self.tree_2.insert("", 13, values=("J3B_VDDQDDR_1V1_G1", "0", "0"))
        self.tree_2.insert("", 14, values=("J3C_COREPD_0V8_G3", "0", "0"))
        self.tree_2.insert("", 15, values=("J3C_VDDQDDR_1V1_G1", "0", "0"))
        self.tree_2.insert("", 16, values=("GPSANT_PWR_AMUX", "0", "0"))
        self.tree_2.insert("", 17, values=("GPS_3V3", "0", "0"))
        self.tree_2.insert("", 18, values=("SYS_5V", "0", "0"))
        self.tree_2.insert("", 19, values=("ETH_0V9", "0", "0"))
        self.tree_2.insert("", 20, values=("HWID", "0", "0"))
        self.tree_2.insert("", 21, values=("Proj_ID", "0", "0"))

        self.tree_2.insert("", 22, values=("SW1_3V3", "0", "0"))
        self.tree_2.insert("", 23, values=("SW1_1V8", "0", "0"))
        self.tree_2.insert("", 24, values=("SW1_1V2", "0", "0"))
        self.tree_2.insert("", 25, values=("J3A_VDD_0V8A_G4", "0", "0"))
        self.tree_2.insert("", 26, values=("J3B_VDD_0V8A_G4", "0", "0"))
        self.tree_2.insert("", 27, values=("J3C_VDD_0V8A_G4", "0", "0"))

        self.tree_2.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')

        '''J3_Status'''
        self.J3_Status = tk.LabelFrame(self.tab, text="J3_Status", padx=10, pady=10)
        self.J3_Status.place(x=250, y=20)
        self.ybar = Scrollbar(self.J3_Status, orient='vertical')
        self.tree_3 = ttk.Treeview(self.J3_Status, yscrollcommand=self.ybar.set, show="headings", height=10)
        self.ybar['command'] = self.tree.yview
        self.tree_3["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_3.column("Name", width=100)
        self.tree_3.column("Vulue", width=50)
        self.tree_3.column("ErrCnt", width=50)
        self.tree_3.heading("Name", text="Name")
        self.tree_3.heading("Vulue", text="Vulue")
        self.tree_3.heading("ErrCnt", text="ErrCnt")
        self.tree_3.insert("", 0, values=("J3A_Alive_Status", "0", "0"))
        self.tree_3.insert("", 1, values=("J3B_Alive_Status", "0", "0"))
        self.tree_3.insert("", 2, values=("J3C_Alive_Status", "0", "0"))
        self.tree_3.insert("", 3, values=("MCU_J3A_SPI_Status", "0", "0"))
        self.tree_3.insert("", 4, values=("MCU_J3B_SPI_Status", "0", "0"))
        self.tree_3.insert("", 5, values=("MCU_J3C_SPI_Status", "0", "0"))
        self.tree_3.insert("", 6, values=("J3A_Eth_TimeOut", "0", "0"))
        self.tree_3.insert("", 7, values=("J3B_Eth_TimeOut", "0", "0"))
        self.tree_3.insert("", 8, values=("J3C_Eth_TimeOut", "0", "0"))
        self.tree_3.insert("", 9, values=("J3A_Eth_RC", "0", "0"))
        self.tree_3.insert("", 10, values=("J3B_Eth_RC", "0", "0"))
        self.tree_3.insert("", 11, values=("J3C_Eth_RC", "0", "0"))
        self.tree_3.insert("", 12, values=("J3A_CPU_Load", "0", "0"))
        self.tree_3.insert("", 13, values=("J3B_CPU_Load", "0", "0"))
        self.tree_3.insert("", 14, values=("J3C_CPU_Load", "0", "0"))
        self.tree_3.insert("", 15, values=("J3A_eMMC_Status", "0", "0"))
        self.tree_3.insert("", 16, values=("J3B_eMMC_Status", "0", "0"))
        self.tree_3.insert("", 17, values=("J3C_eMMC_Status", "0", "0"))
        self.tree_3.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')



        '''Surround_Camera'''
        self.Surround_Camera = tk.LabelFrame(self.tab, text="J3C_Surround_Camera", padx=10, pady=10)
        self.Surround_Camera.place(x=250, y=300)
        self.ybar = Scrollbar(self.Surround_Camera, orient='vertical')
        self.tree_4 = ttk.Treeview(self.Surround_Camera, yscrollcommand=self.ybar.set, show="headings", height=10)
        self.ybar['command'] = self.tree.yview
        self.tree_4["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_4.column("Name", width=100)
        self.tree_4.column("Vulue", width=50)
        self.tree_4.column("ErrCnt", width=50)
        self.tree_4.heading("Name", text="Name")
        self.tree_4.heading("Vulue", text="Vulue")
        self.tree_4.heading("ErrCnt", text="ErrCnt")
        self.tree_4.insert("", 0, values=("Input0_Link_Lock", "0", "0"))
        self.tree_4.insert("", 1, values=("Input0_Lock", "0", "0"))
        self.tree_4.insert("", 2, values=("Input0_Fps", "0", "0"))
        self.tree_4.insert("", 3, values=("Input0_Crc", "0", "0"))
        self.tree_4.insert("", 4, values=("Input0_OpenLoad", "0", "0"))
        self.tree_4.insert("", 5, values=("Input0_Short", "0", "0"))
        self.tree_4.insert("", 6, values=("Input1_Link_Lock", "0", "0"))
        self.tree_4.insert("", 7, values=("Input1_Lock", "0", "0"))
        self.tree_4.insert("", 8, values=("Input1_Fps", "0", "0"))
        self.tree_4.insert("", 9, values=("Input1_Crc", "0", "0"))
        self.tree_4.insert("", 10, values=("Input1_OpenLoad", "0", "0"))
        self.tree_4.insert("", 11, values=("Input1_Short", "0", "0"))
        self.tree_4.insert("", 12, values=("Input2_Link_Lock", "0", "0"))
        self.tree_4.insert("", 13, values=("Input2_Lock", "0", "0"))
        self.tree_4.insert("", 14, values=("Input2_Fps", "0", "0"))
        self.tree_4.insert("", 15, values=("Input2_Crc", "0", "0"))
        self.tree_4.insert("", 16, values=("Input2_OpenLoad", "0", "0"))
        self.tree_4.insert("", 17, values=("Input2_Short", "0", "0"))
        self.tree_4.insert("", 18, values=("Input3_Link_Lock", "0", "0"))
        self.tree_4.insert("", 19, values=("Input3_Lock", "0", "0"))
        self.tree_4.insert("", 20, values=("Input3_Fps", "0", "0"))
        self.tree_4.insert("", 21, values=("Input3_Crc", "0", "0"))
        self.tree_4.insert("", 22, values=("Input3_OpenLoad", "0", "0"))
        self.tree_4.insert("", 23, values=("Input3_Short", "0", "0"))
        self.tree_4.insert("", 24, values=("Output_Link_Lock", "0", "0"))
        self.tree_4.insert("", 25, values=("Output_Lock", "0", "0"))
        self.tree_4.insert("", 26, values=("Output_Fps", "0", "0"))
        self.tree_4.insert("", 27, values=("Output_Crc", "0", "0"))
        self.tree_4.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')

        '''Rear_Camera'''
        self.Rear_Camera = tk.LabelFrame(self.tab, text="J3B_Rear_Camera", padx=10, pady=10)
        self.Rear_Camera.place(x=250, y=580)
        self.tree_5 = ttk.Treeview(self.Rear_Camera, show="headings", height=6)
        self.tree_5["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_5.column("Name", width=100)
        self.tree_5.column("Vulue", width=65)
        self.tree_5.column("ErrCnt", width=50)
        self.tree_5.heading("Name", text="Name")
        self.tree_5.heading("Vulue", text="Vulue")
        self.tree_5.heading("ErrCnt", text="ErrCnt")
        self.tree_5.insert("", 0, values=("Input_Link_Lock", "0", "0"))
        self.tree_5.insert("", 1, values=("Input_Lock", "0", "0"))
        self.tree_5.insert("", 2, values=("Input_Fps", "0", "0"))
        self.tree_5.insert("", 3, values=("Input_Crc", "0", "0"))
        self.tree_5.insert("", 4, values=("Input_OpenLoad", "0", "0"))
        self.tree_5.insert("", 5, values=("Input_Short", "0", "0"))
        self.tree_5.grid(row=0, column=0)

        '''FC'''
        self.FC = tk.LabelFrame(self.tab, text="J3A_Front_Camera", padx=10, pady=10)
        self.FC.place(x=500, y=20)
        self.tree_6 = ttk.Treeview(self.FC, show="headings", height=10)
        self.tree_6["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_6.column("Name", width=100)
        self.tree_6.column("Vulue", width=65)
        self.tree_6.column("ErrCnt", width=50)
        self.tree_6.heading("Name", text="Name")
        self.tree_6.heading("Vulue", text="Vulue")
        self.tree_6.heading("ErrCnt", text="ErrCnt")
        self.tree_6.insert("", 0, values=("Input_Link_Lock", "0", "0"))
        self.tree_6.insert("", 1, values=("Input_Lock", "0", "0"))
        self.tree_6.insert("", 2, values=("Input_Fps", "0", "0"))
        self.tree_6.insert("", 3, values=("Input_Crc", "0", "0"))
        self.tree_6.insert("", 4, values=("Input_OpenLoad", "0", "0"))
        self.tree_6.insert("", 5, values=("Input_Short", "0", "0"))
        self.tree_6.insert("", 6, values=("Output_Link_Lock", "0", "0"))
        self.tree_6.insert("", 7, values=("Output_Lock", "0", "0"))
        self.tree_6.insert("", 8, values=("Output_Fps", "0", "0"))
        self.tree_6.insert("", 9, values=("Output_Crc", "0", "0"))
        self.tree_6.grid(row=0, column=0)

        '''Side_Camera'''
        self.Side_Camera = tk.LabelFrame(self.tab, text="J3B_Side_Camera", padx=10, pady=10)
        self.Side_Camera.place(x=500, y=300)
        self.ybar = Scrollbar(self.Side_Camera, orient='vertical')
        self.tree_7 = ttk.Treeview(self.Side_Camera, yscrollcommand=self.ybar.set, show="headings", height=9)
        self.ybar['command'] = self.tree.yview
        self.tree_7["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_7.column("Name", width=100)
        self.tree_7.column("Vulue", width=50)
        self.tree_7.column("ErrCnt", width=50)
        self.tree_7.heading("Name", text="Name")
        self.tree_7.heading("Vulue", text="Vulue")
        self.tree_7.heading("ErrCnt", text="ErrCnt")
        self.tree_7.insert("", 0, values=("Input0_Link_Lock", "0", "0"))
        self.tree_7.insert("", 1, values=("Input0_Lock", "0", "0"))
        self.tree_7.insert("", 2, values=("Input0_Fps", "0", "0"))
        self.tree_7.insert("", 3, values=("Input0_Crc", "0", "0"))
        self.tree_7.insert("", 4, values=("Input0_OpenLoad", "0", "0"))
        self.tree_7.insert("", 5, values=("Input0_Short", "0", "0"))
        self.tree_7.insert("", 6, values=("Input1_Link_Lock", "0", "0"))
        self.tree_7.insert("", 7, values=("Input1_Lock", "0", "0"))
        self.tree_7.insert("", 8, values=("Input1_Fps", "0", "0"))
        self.tree_7.insert("", 9, values=("Input1_Crc", "0", "0"))
        self.tree_7.insert("", 10, values=("Input1_OpenLoad", "0", "0"))
        self.tree_7.insert("", 11, values=("Input1_Short", "0", "0"))
        self.tree_7.insert("", 12, values=("Input2_Link_Lock", "0", "0"))
        self.tree_7.insert("", 13, values=("Input2_Lock", "0", "0"))
        self.tree_7.insert("", 14, values=("Input2_Fps", "0", "0"))
        self.tree_7.insert("", 15, values=("Input2_Crc", "0", "0"))
        self.tree_7.insert("", 16, values=("Input2_OpenLoad", "0", "0"))
        self.tree_7.insert("", 17, values=("Input2_Short", "0", "0"))
        self.tree_7.insert("", 18, values=("Input3_Link_Lock", "0", "0"))
        self.tree_7.insert("", 19, values=("Input3_Lock", "0", "0"))
        self.tree_7.insert("", 20, values=("Input3_Fps", "0", "0"))
        self.tree_7.insert("", 21, values=("Input3_Crc", "0", "0"))
        self.tree_7.insert("", 22, values=("Input3_OpenLoad", "0", "0"))
        self.tree_7.insert("", 23, values=("Input3_Short", "0", "0"))
        self.tree_7.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')

        '''GNSS'''
        self.GNSS = tk.LabelFrame(self.tab, text="GNSS", padx=10, pady=10)
        self.GNSS.place(x=500, y=560)
        self.tree_8 = ttk.Treeview(self.GNSS, show="headings", height=7)
        self.tree_8["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_8.column("Name", width=100)
        self.tree_8.column("Vulue", width=65)
        self.tree_8.column("ErrCnt", width=50)
        self.tree_8.heading("Name", text="Name")
        self.tree_8.heading("Vulue", text="Vulue")
        self.tree_8.heading("ErrCnt", text="ErrCnt")
        self.tree_8.insert("", 0, values=("J3A_F9K_GNSS_Status", "0", "0"))
        self.tree_8.insert("", 1, values=("MCU_F9K_GNSS_OpenLoad", "0", "0"))
        self.tree_8.insert("", 2, values=("MCU_F9K_GNSS_Short", "0", "0"))
        self.tree_8.insert("", 3, values=("MCU_F9K_GNSS_Supply", "0", "0"))
        self.tree_8.insert("", 4, values=("J3A_F9K_IMU_Z", "0", "0"))
        self.tree_8.insert("", 5, values=("J3A_F9K_CN_Value", "0", "0"))
        self.tree_8.grid(row=0, column=0)

        '''TEMP'''
        self.TEMP = tk.LabelFrame(self.tab, text="TEMP", padx=10, pady=10)
        self.TEMP.place(x=750, y=20)
        self.ybar = Scrollbar(self.TEMP, orient='vertical')
        self.tree_9 = ttk.Treeview(self.TEMP, yscrollcommand=self.ybar.set, show="headings", height=7)
        self.ybar['command'] = self.tree.yview
        self.tree_9["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_9.column("Name", width=100)
        self.tree_9.column("Vulue", width=50)
        self.tree_9.column("ErrCnt", width=50)
        self.tree_9.heading("Name", text="Name")
        self.tree_9.heading("Vulue", text="Vulue")
        self.tree_9.heading("ErrCnt", text="ErrCnt")
        self.tree_9.insert("", 0, values=("MCU_TEMP", "0", "0"))
        self.tree_9.insert("", 1, values=("J3A_5024_TEMP", "0", "0"))
        self.tree_9.insert("", 2, values=("J3B_5024_TEMP", "0", "0"))
        self.tree_9.insert("", 3, values=("J3C_5024_TEMP", "0", "0"))
        self.tree_9.insert("", 4, values=("J3A_TEMP", "0", "0"))
        self.tree_9.insert("", 5, values=("J3B_TEMP", "0", "0"))
        self.tree_9.insert("", 6, values=("J3C_TEMP", "0", "0"))
        self.tree_9.insert("", 7, values=("J3A_F9K_TEMP", "0", "0"))
        self.tree_9.insert("", 8, values=("PCB_TEMP", "0", "0"))
        self.tree_9.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')

        '''J3A'''
        self.J3A = tk.LabelFrame(self.tab, text="J3A", padx=10, pady=10)
        self.J3A.place(x=750, y=230)
        self.ybar = Scrollbar(self.J3A, orient='vertical')
        self.tree_10 = ttk.Treeview(self.J3A, yscrollcommand=self.ybar.set, show="headings", height=10)
        self.ybar['command'] = self.tree.yview
        self.tree_10["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_10.column("Name", width=100)
        self.tree_10.column("Vulue", width=50)
        self.tree_10.column("ErrCnt", width=50)
        self.tree_10.heading("Name", text="Name")
        self.tree_10.heading("Vulue", text="Vulue")
        self.tree_10.heading("ErrCnt", text="ErrCnt")
        self.tree_10.insert("", 0, values=("CPU_0V8_G3", "0", "0"))
        self.tree_10.insert("", 1, values=("DDR_0V8_G3", "0", "0"))
        self.tree_10.insert("", 2, values=("VDD_0V8_G4", "0", "0"))
        self.tree_10.insert("", 3, values=("COREAO_0V8_G2", "0", "0"))
        self.tree_10.insert("", 4, values=("CNN0_0V8_G3", "0", "0"))
        self.tree_10.insert("", 5, values=("CNN1_0V8_G3", "0", "0"))
        self.tree_10.insert("", 6, values=("VDD_1V8_G1", "0", "0"))
        self.tree_10.insert("", 7, values=("EMMC_3V3_G1", "0", "0"))
        self.tree_10.insert("", 8, values=("VDD_3V3_G1", "0", "0"))
        self.tree_10.insert("", 9, values=("VCC_3V3_G4", "0", "0"))
        self.tree_10.insert("", 10, values=("VCC_1V8_G4", "0", "0"))
        self.tree_10.insert("", 11, values=("TEMP_IC", "0", "0"))
        self.tree_10.insert("", 12, values=("TEMP_SW1", "0", "0"))
        self.tree_10.insert("", 13, values=("TEMP_SW2", "0", "0"))
        self.tree_10.insert("", 14, values=("TEMP_SW3", "0", "0"))
        self.tree_10.insert("", 15, values=("TEMP_SW4", "0", "0"))
        self.tree_10.insert("", 16, values=("TEMP_SW5", "0", "0"))
        self.tree_10.insert("", 17, values=("TEMP_SW6", "0", "0"))
        self.tree_10.insert("", 18, values=("TEMP_SW7", "0", "0"))
        self.tree_10.insert("", 19, values=("TEMP_TEMP_LDO1_2", "0", "0"))
        self.tree_10.insert("", 20, values=("TEMP_TEMP_LDO3_4", "0", "0"))
        self.tree_10.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')


        '''J3B'''
        self.J3B = tk.LabelFrame(self.tab, text="J3B", padx=10, pady=10)
        self.J3B.place(x=750, y=500)
        self.ybar = Scrollbar(self.J3B, orient='vertical')
        self.tree_11 = ttk.Treeview(self.J3B, yscrollcommand=self.ybar.set, show="headings", height=10)
        self.ybar['command'] = self.tree.yview
        self.tree_11["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_11.column("Name", width=100)
        self.tree_11.column("Vulue", width=50)
        self.tree_11.column("ErrCnt", width=50)
        self.tree_11.heading("Name", text="Name")
        self.tree_11.heading("Vulue", text="Vulue")
        self.tree_11.heading("ErrCnt", text="ErrCnt")
        self.tree_11.insert("", 0, values=("CPU_0V8_G3", "0", "0"))
        self.tree_11.insert("", 1, values=("DDR_0V8_G3", "0", "0"))
        self.tree_11.insert("", 2, values=("VDD_0V8_G4", "0", "0"))
        self.tree_11.insert("", 3, values=("COREAO_0V8_G2", "0", "0"))
        self.tree_11.insert("", 4, values=("CNN0_0V8_G3", "0", "0"))
        self.tree_11.insert("", 5, values=("CNN1_0V8_G3", "0", "0"))
        self.tree_11.insert("", 6, values=("VDD_1V8_G1", "0", "0"))
        self.tree_11.insert("", 7, values=("EMMC_3V3_G1", "0", "0"))
        self.tree_11.insert("", 8, values=("VDD_3V3_G1", "0", "0"))
        self.tree_11.insert("", 9, values=("VCC_3V3_G4", "0", "0"))
        self.tree_11.insert("", 10, values=("VCC_1V8_G4", "0", "0"))
        self.tree_11.insert("", 11, values=("TEMP_IC", "0", "0"))
        self.tree_11.insert("", 12, values=("TEMP_SW1", "0", "0"))
        self.tree_11.insert("", 13, values=("TEMP_SW2", "0", "0"))
        self.tree_11.insert("", 14, values=("TEMP_SW3", "0", "0"))
        self.tree_11.insert("", 15, values=("TEMP_SW4", "0", "0"))
        self.tree_11.insert("", 16, values=("TEMP_SW5", "0", "0"))
        self.tree_11.insert("", 17, values=("TEMP_SW6", "0", "0"))
        self.tree_11.insert("", 18, values=("TEMP_SW7", "0", "0"))
        self.tree_11.insert("", 19, values=("TEMP_TEMP_LDO1_2", "0", "0"))
        self.tree_11.insert("", 20, values=("TEMP_TEMP_LDO3_4", "0", "0"))
        self.tree_11.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')

        '''J3C'''
        self.J3C = tk.LabelFrame(self.tab, text="J3C", padx=10, pady=10)
        self.J3C.place(x=1000, y=20)
        self.ybar = Scrollbar(self.J3C, orient='vertical')
        self.tree_12 = ttk.Treeview(self.J3C, yscrollcommand=self.ybar.set, show="headings", height=10)
        self.ybar['command'] = self.tree.yview
        self.tree_12["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_12.column("Name", width=100)
        self.tree_12.column("Vulue", width=50)
        self.tree_12.column("ErrCnt", width=50)
        self.tree_12.heading("Name", text="Name")
        self.tree_12.heading("Vulue", text="Vulue")
        self.tree_12.heading("ErrCnt", text="ErrCnt")
        self.tree_12.insert("", 0, values=("CPU_0V8_G3", "0", "0"))
        self.tree_12.insert("", 1, values=("DDR_0V8_G3", "0", "0"))
        self.tree_12.insert("", 2, values=("VDD_0V8_G4", "0", "0"))
        self.tree_12.insert("", 3, values=("COREAO_0V8_G2", "0", "0"))
        self.tree_12.insert("", 4, values=("CNN0_0V8_G3", "0", "0"))
        self.tree_12.insert("", 5, values=("CNN1_0V8_G3", "0", "0"))
        self.tree_12.insert("", 6, values=("VDD_1V8_G1", "0", "0"))
        self.tree_12.insert("", 7, values=("EMMC_3V3_G1", "0", "0"))
        self.tree_12.insert("", 8, values=("VDD_3V3_G1", "0", "0"))
        self.tree_12.insert("", 9, values=("VCC_3V3_G4", "0", "0"))
        self.tree_12.insert("", 10, values=("VCC_1V8_G4", "0", "0"))
        self.tree_12.insert("", 11, values=("TEMP_IC", "0", "0"))
        self.tree_12.insert("", 12, values=("TEMP_SW1", "0", "0"))
        self.tree_12.insert("", 13, values=("TEMP_SW2", "0", "0"))
        self.tree_12.insert("", 14, values=("TEMP_SW3", "0", "0"))
        self.tree_12.insert("", 15, values=("TEMP_SW4", "0", "0"))
        self.tree_12.insert("", 16, values=("TEMP_SW5", "0", "0"))
        self.tree_12.insert("", 17, values=("TEMP_SW6", "0", "0"))
        self.tree_12.insert("", 18, values=("TEMP_SW7", "0", "0"))
        self.tree_12.insert("", 19, values=("TEMP_TEMP_LDO1_2", "0", "0"))
        self.tree_12.insert("", 20, values=("TEMP_TEMP_LDO3_4", "0", "0"))
        self.tree_12.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')

        '''I2C'''
        self.I2C = tk.LabelFrame(self.tab, text="I2C", padx=10, pady=10)
        self.I2C.place(x=1000, y=300)
        self.ybar = Scrollbar(self.I2C, orient='vertical')
        self.tree_13 = ttk.Treeview(self.I2C, yscrollcommand=self.ybar.set, show="headings", height=11)
        self.ybar['command'] = self.tree.yview
        self.tree_13["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_13.column("Name", width=100)
        self.tree_13.column("Vulue", width=50)
        self.tree_13.column("ErrCnt", width=50)
        self.tree_13.heading("Name", text="Name")
        self.tree_13.heading("Vulue", text="Vulue")
        self.tree_13.heading("ErrCnt", text="ErrCnt")
        self.tree_13.insert("", 0, values=("MCU_MAX20084", "0", "0"))
        self.tree_13.insert("", 1, values=("MCU_PMIC8100_Index0", "0", "0"))
        self.tree_13.insert("", 2, values=("MCU_PMIC8100_Index1", "0", "0"))
        self.tree_13.insert("", 3, values=("MCU_PMIC8100_Index2", "0", "0"))
        self.tree_13.insert("", 4, values=("MCU_PMIC5024_Index0", "0", "0"))
        self.tree_13.insert("", 5, values=("MCU_PMIC5024_Index1", "0", "0"))
        self.tree_13.insert("", 6, values=("MCU_PMIC5024_Index2", "0", "0"))
        self.tree_13.insert("", 7, values=("J3A_MAX9296", "0", "0"))
        self.tree_13.insert("", 8, values=("J3A_MAX96717", "0", "0"))
        self.tree_13.insert("", 9, values=("J3A_MAX20089", "0", "0"))
        self.tree_13.insert("", 10, values=("J3B_MAX9296", "0", "0"))
        self.tree_13.insert("", 11, values=("J3B_MAX96712", "0", "0"))
        self.tree_13.insert("", 12, values=("J3B_MAX20089", "0", "0"))
        self.tree_13.insert("", 13, values=("J3B_MAX20087", "0", "0"))
        self.tree_13.insert("", 14, values=("J3C_MAX96712", "0", "0"))
        self.tree_13.insert("", 15, values=("J3C_MAX96717", "0", "0"))
        self.tree_13.insert("", 16, values=("J3C_MAX20087", "0", "0"))

        self.tree_13.insert("", 17, values=("J3A_Front_Camera_max9295_ID", "0", "0"))
        self.tree_13.insert("", 18, values=("J3B_Rear_Camera_max9295_ID", "0", "0"))
        self.tree_13.insert("", 19, values=("J3B_Side_Camera_0_max9295_ID", "0", "0"))
        self.tree_13.insert("", 20, values=("J3B_Side_Camera_1_max9295_ID", "0", "0"))
        self.tree_13.insert("", 21, values=("J3B_Side_Camera_2_max9295_ID", "0", "0"))
        self.tree_13.insert("", 22, values=("J3B_Side_Camera_3_max9295_ID", "0", "0"))
        self.tree_13.insert("", 23, values=("J3C_Surround_Camera_0_max9295_ID", "0", "0"))
        self.tree_13.insert("", 24, values=("J3C_Surround_Camera_1_max9295_ID", "0", "0"))
        self.tree_13.insert("", 25, values=("J3C_Surround_Camera_2_max9295_ID", "0", "0"))
        self.tree_13.insert("", 26, values=("J3C_Surround_Camera_3_max9295_ID", "0", "0"))

        self.tree_13.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')

        '''E521p42'''
        self.E521p42 = tk.LabelFrame(self.tab, text="E521p42", padx=10, pady=10)
        self.E521p42.place(x=1250, y=20)
        self.ybar = Scrollbar(self.E521p42, orient='vertical')
        self.tree_14 = ttk.Treeview(self.E521p42, yscrollcommand=self.ybar.set, show="headings", height=10)
        self.ybar['command'] = self.tree.yview
        self.tree_14["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_14.column("Name", width=140)
        self.tree_14.column("Vulue", width=50)
        self.tree_14.column("ErrCnt", width=50)
        self.tree_14.heading("Name", text="Name")
        self.tree_14.heading("Vulue", text="Vulue")
        self.tree_14.heading("ErrCnt", text="ErrCnt")
        self.tree_14.insert("", 0, values=("USS1_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 1, values=("USS1_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 2, values=("USS2_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 3, values=("USS2_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 4, values=("USS3_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 5, values=("USS3_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 6, values=("USS4_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 7, values=("USS4_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 8, values=("USS5_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 9, values=("USS5_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 10, values=("USS6_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 11, values=("USS6_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 12, values=("USS7_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 13, values=("USS7_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 14, values=("USS8_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 15, values=("USS8_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 16, values=("USS9_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 17, values=("USS9_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 18, values=("USS10_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 19, values=("USS10_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 20, values=("USS11_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 21, values=("USS11_PWM_DUTY", "0", "0"))
        self.tree_14.insert("", 22, values=("USS12_PWM_PERIOD", "0", "0"))
        self.tree_14.insert("", 23, values=("USS12_PWM_DUTY", "0", "0"))

        self.tree_14.grid(row=0, column=0)
        self.ybar.grid(row=0, column=1, sticky='ns')

        '''Camera_Supply'''
        self.Camera_Supply = tk.LabelFrame(self.tab, text="Camera_Supply_AD", padx=10, pady=10)
        self.Camera_Supply.place(x=1250, y=300)
        self.tree_15 = ttk.Treeview(self.Camera_Supply, show="headings", height=10)
        self.tree_15["columns"] = ("Name", "Vulue", "ErrCnt")
        self.tree_15.column("Name", width=140)
        self.tree_15.column("Vulue", width=65)
        self.tree_15.column("ErrCnt", width=50)
        self.tree_15.heading("Name", text="Name")
        self.tree_15.heading("Vulue", text="Vulue")
        self.tree_15.heading("ErrCnt", text="ErrCnt")
        self.tree_15.insert("", 0, values=("J3A_Front_Camera", "0", "0"))
        self.tree_15.insert("", 1, values=("J3B_Rear_Camera", "0", "0"))
        self.tree_15.insert("", 2, values=("J3B_Side_Camera_0", "0", "0"))
        self.tree_15.insert("", 3, values=("J3B_Side_Camera_1", "0", "0"))
        self.tree_15.insert("", 4, values=("J3B_Side_Camera_2", "0", "0"))
        self.tree_15.insert("", 5, values=("J3B_Side_Camera_3", "0", "0"))
        self.tree_15.insert("", 6, values=("J3C_Surround_Camera_0", "0", "0"))
        self.tree_15.insert("", 7, values=("J3C_Surround_Camera_1", "0", "0"))
        self.tree_15.insert("", 8, values=("J3C_Surround_Camera_2", "0", "0"))
        self.tree_15.insert("", 9, values=("J3C_Surround_Camera_3", "0", "0"))
        self.tree_15.grid(row=0, column=0)


    def tab_Refresh(self, data_analysed):
        for _ in map(self.tree.delete, self.tree.get_children("")):
            pass
        self.tree.insert("", 0, values=("CAN0_Status", str(data_analysed[1][0]), str(data_analysed[1][2])))
        self.tree.insert("", 1, values=("CAN1_Status",str(data_analysed[2][0]), str(data_analysed[2][2])))
        self.tree.insert("", 2, values=("CAN2_Status", str(data_analysed[3][0]), str(data_analysed[3][2])))
        self.tree.insert("", 3, values=("CAN3_Status", str(data_analysed[4][0]), str(data_analysed[4][2])))
        self.tree.insert("", 4, values=("CAN4_Status", str(data_analysed[5][0]), str(data_analysed[5][2])))
        self.tree.insert("", 5, values=("USS1_Vout", str(data_analysed[11][0]), str(data_analysed[11][2])))
        self.tree.insert("", 6, values=("USS2_Vout", str(data_analysed[12][0]), str(data_analysed[12][2])))
        self.tree.insert("", 7, values=("USS3_Vout", str(data_analysed[13][0]), str(data_analysed[13][2])))


        for _ in map(self.tree_2.delete, self.tree_2.get_children("")):
            pass
        self.tree_2.insert("", 0, values=("VBAT_12V", str(data_analysed[14][0]), str(data_analysed[14][2])))
        self.tree_2.insert("", 1, values=("VBoost_12V",str(data_analysed[15][0]), str(data_analysed[15][2])))
        self.tree_2.insert("", 2, values=("TJA1043_INH", str(data_analysed[16][0]), str(data_analysed[16][2])))
        self.tree_2.insert("", 3, values=("IG_12V", str(data_analysed[17][0]), str(data_analysed[17][2])))
        self.tree_2.insert("", 4, values=("SWITCH_INH",str(data_analysed[18][0]), str(data_analysed[18][2])))
        self.tree_2.insert("", 5, values=("SYS_1V8A", str(data_analysed[19][0]), str(data_analysed[19][2])))
        self.tree_2.insert("", 6, values=("SYS_1V8B", str(data_analysed[20][0]), str(data_analysed[20][2])))
        self.tree_2.insert("", 7, values=("SYS_1V2A", str(data_analysed[21][0]), str(data_analysed[21][2])))
        self.tree_2.insert("", 8, values=("SYS_1V2B", str(data_analysed[22][0]), str(data_analysed[22][2])))
        self.tree_2.insert("", 9, values=("SYS_3V3",str(data_analysed[23][0]), str(data_analysed[23][2])))
        self.tree_2.insert("", 10, values=("J3A_COREPD_0V8_G3",str(data_analysed[24][0]), str(data_analysed[24][2])))
        self.tree_2.insert("", 11, values=("J3A_VDDQDDR_1V1_G1",str(data_analysed[25][0]), str(data_analysed[25][2])))
        self.tree_2.insert("", 12, values=("J3B_COREPD_0V8_G3",str(data_analysed[26][0]), str(data_analysed[26][2])))
        self.tree_2.insert("", 13, values=("J3B_VDDQDDR_1V1_G1", str(data_analysed[27][0]), str(data_analysed[27][2])))
        self.tree_2.insert("", 14, values=("J3C_COREPD_0V8_G3", str(data_analysed[28][0]), str(data_analysed[28][2])))
        self.tree_2.insert("", 15, values=("J3C_VDDQDDR_1V1_G1", str(data_analysed[29][0]), str(data_analysed[29][2])))
        self.tree_2.insert("", 16, values=("GPSANT_PWR_AMUX", str(data_analysed[30][0]), str(data_analysed[30][2])))
        self.tree_2.insert("", 17, values=("GPS_3V3", str(data_analysed[31][0]), str(data_analysed[31][2])))
        self.tree_2.insert("", 18, values=("SYS_5V", str(data_analysed[32][0]), str(data_analysed[32][2])))
        self.tree_2.insert("", 19, values=("ETH_0V9", str(data_analysed[33][0]), str(data_analysed[33][2])))
        self.tree_2.insert("", 20, values=("HWID", str(data_analysed[34][0]), str(data_analysed[34][2])))
        self.tree_2.insert("", 21, values=("Proj_ID",str(data_analysed[35][0]), str(data_analysed[35][2])))

        self.tree_2.insert("", 22, values=("SW1_3V3", str(data_analysed[227][0]), str(data_analysed[227][2])))
        self.tree_2.insert("", 23, values=("SW1_1V8", str(data_analysed[228][0]), str(data_analysed[228][2])))
        self.tree_2.insert("", 24, values=("SW1_1V2", str(data_analysed[229][0]), str(data_analysed[229][2])))
        self.tree_2.insert("", 25, values=("J3A_VDD_0V8A_G4", str(data_analysed[230][0]), str(data_analysed[230][2])))
        self.tree_2.insert("", 26, values=("J3B_VDD_0V8A_G4", str(data_analysed[231][0]), str(data_analysed[231][2])))
        self.tree_2.insert("", 27, values=("J3C_VDD_0V8A_G4", str(data_analysed[232][0]), str(data_analysed[232][2])))


        for _ in map(self.tree_3.delete, self.tree_3.get_children("")):
            pass
        self.tree_3.insert("", 0, values=("J3A_Alive_Status", str(data_analysed[36][0]), str(data_analysed[36][2])))
        self.tree_3.insert("", 1, values=("J3B_Alive_Status", str(data_analysed[37][0]), str(data_analysed[37][2])))
        self.tree_3.insert("", 2, values=("J3C_Alive_Status", str(data_analysed[38][0]), str(data_analysed[38][2])))
        self.tree_3.insert("", 3, values=("MCU_J3A_SPI_Status", str(data_analysed[39][0]), str(data_analysed[39][2])))
        self.tree_3.insert("", 4, values=("MCU_J3B_SPI_Status", str(data_analysed[40][0]), str(data_analysed[40][2])))
        self.tree_3.insert("", 5, values=("MCU_J3C_SPI_Status", str(data_analysed[41][0]), str(data_analysed[41][2])))
        self.tree_3.insert("", 6, values=("J3A_Eth_TO", str(data_analysed[42][0]), str(data_analysed[42][2])))
        self.tree_3.insert("", 7, values=("J3B_Eth_TO", str(data_analysed[43][0]), str(data_analysed[43][2])))
        self.tree_3.insert("", 8, values=("J3C_Eth_TO", str(data_analysed[44][0]), str(data_analysed[44][2])))
        self.tree_3.insert("", 9, values=("J3A_Eth_RC", str(data_analysed[45][0]), str(data_analysed[45][2])))
        self.tree_3.insert("", 10, values=("J3B_Eth_RC", str(data_analysed[46][0]), str(data_analysed[46][2])))
        self.tree_3.insert("", 11, values=("J3C_Eth_RC", str(data_analysed[47][0]), str(data_analysed[47][2])))
        self.tree_3.insert("", 12, values=("J3A_CPU_Load", str(data_analysed[48][0]), str(data_analysed[48][2])))
        self.tree_3.insert("", 13, values=("J3B_CPU_Load", str(data_analysed[49][0]), str(data_analysed[49][2])))
        self.tree_3.insert("", 14, values=("J3C_CPU_Load", str(data_analysed[50][0]), str(data_analysed[50][2])))
        self.tree_3.insert("", 15, values=("J3A_eMMC_Status", str(data_analysed[51][0]), str(data_analysed[51][2])))
        self.tree_3.insert("", 16, values=("J3B_eMMC_Status", str(data_analysed[52][0]), str(data_analysed[52][2])))
        self.tree_3.insert("", 17, values=("J3C_eMMC_Status", str(data_analysed[53][0]), str(data_analysed[53][2])))


        for _ in map(self.tree_4.delete, self.tree_4.get_children("")):
            pass
        self.tree_4.insert("", 0, values=("Input0_Link_Lock", str(data_analysed[54][0]), str(data_analysed[54][2])))
        self.tree_4.insert("", 1, values=("Input0_Lock", str(data_analysed[55][0]), str(data_analysed[55][2])))
        self.tree_4.insert("", 2, values=("Input0_Fps", str(data_analysed[56][0]), str(data_analysed[56][2])))
        self.tree_4.insert("", 3, values=("Input0_Crc",str(data_analysed[57][0]), str(data_analysed[57][2])))
        self.tree_4.insert("", 4, values=("Input0_OpenLoad", str(data_analysed[58][0]), str(data_analysed[58][2])))
        self.tree_4.insert("", 5, values=("Input0_Short", str(data_analysed[59][0]), str(data_analysed[59][2])))
        self.tree_4.insert("", 6, values=("Input1_Link_Lock", str(data_analysed[60][0]), str(data_analysed[60][2])))
        self.tree_4.insert("", 7, values=("Input1_Lock", str(data_analysed[61][0]), str(data_analysed[61][2])))
        self.tree_4.insert("", 8, values=("Input1_Fps", str(data_analysed[62][0]), str(data_analysed[62][2])))
        self.tree_4.insert("", 9, values=("Input1_Crc", str(data_analysed[63][0]), str(data_analysed[63][2])))
        self.tree_4.insert("", 10, values=("Input1_OpenLoad", str(data_analysed[64][0]), str(data_analysed[64][2])))
        self.tree_4.insert("", 11, values=("Input1_Short", str(data_analysed[65][0]), str(data_analysed[65][2])))
        self.tree_4.insert("", 12, values=("Input2_Link_Lock", str(data_analysed[66][0]), str(data_analysed[66][2])))
        self.tree_4.insert("", 13, values=("Input2_Lock", str(data_analysed[67][0]), str(data_analysed[67][2])))
        self.tree_4.insert("", 14, values=("Input2_Fps", str(data_analysed[68][0]), str(data_analysed[68][2])))
        self.tree_4.insert("", 15, values=("Input2_Crc", str(data_analysed[69][0]), str(data_analysed[69][2])))
        self.tree_4.insert("", 16, values=("Input2_OpenLoad", str(data_analysed[70][0]), str(data_analysed[70][2])))
        self.tree_4.insert("", 17, values=("Input2_Short", str(data_analysed[71][0]), str(data_analysed[71][2])))
        self.tree_4.insert("", 18, values=("Input3_Link_Lock", str(data_analysed[72][0]), str(data_analysed[72][2])))
        self.tree_4.insert("", 19, values=("Input3_Lock", str(data_analysed[73][0]), str(data_analysed[73][2])))
        self.tree_4.insert("", 20, values=("Input3_Fps",str(data_analysed[74][0]), str(data_analysed[74][2])))
        self.tree_4.insert("", 21, values=("Input3_Crc",str(data_analysed[75][0]), str(data_analysed[75][2])))
        self.tree_4.insert("", 22, values=("Input3_OpenLoad",str(data_analysed[76][0]), str(data_analysed[76][2])))
        self.tree_4.insert("", 23, values=("Input3_Short",str(data_analysed[77][0]), str(data_analysed[77][2])))
        self.tree_4.insert("", 24, values=("Output_Link_Lock", str(data_analysed[78][0]), str(data_analysed[78][2])))
        self.tree_4.insert("", 25, values=("Output_Lock", str(data_analysed[79][0]), str(data_analysed[79][2])))
        self.tree_4.insert("", 26, values=("Output_Fps", str(data_analysed[80][0]), str(data_analysed[80][2])))
        self.tree_4.insert("", 27, values=("Output_Crc", str(data_analysed[81][0]), str(data_analysed[81][2])))

        for _ in map(self.tree_5.delete, self.tree_5.get_children("")):
            pass
        self.tree_5.insert("", 0, values=("Input_Link_Lock", str(data_analysed[92][0]), str(data_analysed[92][2])))
        self.tree_5.insert("", 1, values=("Input_Lock", str(data_analysed[93][0]), str(data_analysed[93][2])))
        self.tree_5.insert("", 2, values=("Input_Fps", str(data_analysed[94][0]), str(data_analysed[94][2])))
        self.tree_5.insert("", 3, values=("Input_Crc", str(data_analysed[95][0]), str(data_analysed[95][2])))
        self.tree_5.insert("", 4, values=("Input_OpenLoad", str(data_analysed[96][0]), str(data_analysed[96][2])))
        self.tree_5.insert("", 5, values=("Input_Short", str(data_analysed[97][0]), str(data_analysed[97][2])))

        for _ in map(self.tree_6.delete, self.tree_6.get_children("")):
            pass
        self.tree_6.insert("", 0, values=("Input_Link_Lock", str(data_analysed[82][0]), str(data_analysed[82][2])))
        self.tree_6.insert("", 1, values=("Input_Lock",str(data_analysed[83][0]), str(data_analysed[83][2])))
        self.tree_6.insert("", 2, values=("Input_Fps", str(data_analysed[84][0]), str(data_analysed[84][2])))
        self.tree_6.insert("", 3, values=("Input_Crc", str(data_analysed[85][0]), str(data_analysed[85][2])))
        self.tree_6.insert("", 4, values=("Input_OpenLoad", str(data_analysed[86][0]), str(data_analysed[86][2])))
        self.tree_6.insert("", 5, values=("Input_Short", str(data_analysed[87][0]), str(data_analysed[87][2])))
        self.tree_6.insert("", 6, values=("Output_Link_Lock", str(data_analysed[88][0]), str(data_analysed[88][2])))
        self.tree_6.insert("", 7, values=("Output_Lock", str(data_analysed[89][0]), str(data_analysed[89][2])))
        self.tree_6.insert("", 8, values=("Output_Fps", str(data_analysed[90][0]), str(data_analysed[90][2])))
        self.tree_6.insert("", 9, values=("Output_Crc", str(data_analysed[91][0]), str(data_analysed[91][2])))

        for _ in map(self.tree_7.delete, self.tree_7.get_children("")):
            pass
        self.tree_7.insert("", 0, values=("Input0_Link_Lock", str(data_analysed[98][0]), str(data_analysed[98][2])))
        self.tree_7.insert("", 1, values=("Input0_Lock", str(data_analysed[99][0]), str(data_analysed[99][2])))
        self.tree_7.insert("", 2, values=("Input0_Fps", str(data_analysed[100][0]), str(data_analysed[100][2])))
        self.tree_7.insert("", 3, values=("Input0_Crc", str(data_analysed[101][0]), str(data_analysed[101][2])))
        self.tree_7.insert("", 4, values=("Input0_OpenLoad",str(data_analysed[102][0]), str(data_analysed[102][2])))
        self.tree_7.insert("", 5, values=("Input0_Short", str(data_analysed[103][0]), str(data_analysed[103][2])))
        self.tree_7.insert("", 6, values=("Input1_Link_Lock", str(data_analysed[104][0]), str(data_analysed[104][2])))
        self.tree_7.insert("", 7, values=("Input1_Lock",str(data_analysed[105][0]), str(data_analysed[105][2])))
        self.tree_7.insert("", 8, values=("Input1_Fps", str(data_analysed[106][0]), str(data_analysed[106][2])))
        self.tree_7.insert("", 9, values=("Input1_Crc", str(data_analysed[107][0]), str(data_analysed[107][2])))
        self.tree_7.insert("", 10, values=("Input1_OpenLoad", str(data_analysed[108][0]), str(data_analysed[108][2])))
        self.tree_7.insert("", 11, values=("Input1_Short",str(data_analysed[109][0]), str(data_analysed[109][2])))
        self.tree_7.insert("", 12, values=("Input2_Link_Lock",str(data_analysed[110][0]), str(data_analysed[110][2])))
        self.tree_7.insert("", 13, values=("Input2_Lock", str(data_analysed[111][0]), str(data_analysed[111][2])))
        self.tree_7.insert("", 14, values=("Input2_Fps", str(data_analysed[112][0]), str(data_analysed[112][2])))
        self.tree_7.insert("", 15, values=("Input2_Crc", str(data_analysed[113][0]), str(data_analysed[113][2])))
        self.tree_7.insert("", 16, values=("Input2_OpenLoad", str(data_analysed[114][0]), str(data_analysed[114][2])))
        self.tree_7.insert("", 17, values=("Input2_Short", str(data_analysed[115][0]), str(data_analysed[115][2])))
        self.tree_7.insert("", 18, values=("Input3_Link_Lock", str(data_analysed[116][0]), str(data_analysed[116][2])))
        self.tree_7.insert("", 19, values=("Input3_Lock", str(data_analysed[117][0]), str(data_analysed[117][2])))
        self.tree_7.insert("", 20, values=("Input3_Fps", str(data_analysed[118][0]), str(data_analysed[118][2])))
        self.tree_7.insert("", 21, values=("Input3_Crc", str(data_analysed[119][0]), str(data_analysed[119][2])))
        self.tree_7.insert("", 22, values=("Input3_OpenLoad", str(data_analysed[120][0]), str(data_analysed[120][2])))
        self.tree_7.insert("", 23, values=("Input3_Short", str(data_analysed[121][0]), str(data_analysed[121][2])))


        for _ in map(self.tree_8.delete, self.tree_8.get_children("")):
            pass
        self.tree_8.insert("", 0, values=("J3A_F9K_GNSS_Status", str(data_analysed[122][0]), str(data_analysed[122][2])))
        self.tree_8.insert("", 1, values=("MCU_F9K_GNSS_OpenLoad", str(data_analysed[123][0]), str(data_analysed[123][2])))
        self.tree_8.insert("", 2, values=("MCU_F9K_GNSS_Short", str(data_analysed[124][0]), str(data_analysed[124][2])))
        self.tree_8.insert("", 3, values=("MCU_F9K_GNSS_Supply", str(data_analysed[125][0]), str(data_analysed[125][2])))
        self.tree_8.insert("", 4, values=("J3A_F9K_IMU_Z", str(data_analysed[126][0]), str(data_analysed[126][2])))
        self.tree_8.insert("", 5, values=("J3A_F9K_CN_Value", str(data_analysed[127][0]), str(data_analysed[127][2])))

        for _ in map(self.tree_9.delete, self.tree_9.get_children("")):
            pass
        self.tree_9.insert("", 0, values=("MCU_TEMP", str(data_analysed[128][0]), str(data_analysed[128][2])))
        self.tree_9.insert("", 1, values=("J3A_5024_TEMP", str(data_analysed[129][0]), str(data_analysed[129][2])))
        self.tree_9.insert("", 2, values=("J3B_5024_TEMP", str(data_analysed[130][0]), str(data_analysed[130][2])))
        self.tree_9.insert("", 3, values=("J3C_5024_TEMP", str(data_analysed[131][0]), str(data_analysed[131][2])))
        self.tree_9.insert("", 4, values=("J3A_TEMP", str(data_analysed[223][0]), str(data_analysed[223][2])))
        self.tree_9.insert("", 5, values=("J3B_TEMP", str(data_analysed[224][0]), str(data_analysed[224][2])))
        self.tree_9.insert("", 6, values=("J3C_TEMP", str(data_analysed[225][0]), str(data_analysed[225][2])))
        self.tree_9.insert("", 7, values=("J3A_F9K_TEMP", str(data_analysed[226][0]), str(data_analysed[226][2])))
        self.tree_9.insert("", 8, values=("PCB_TEMP", str(data_analysed[132][0]), str(data_analysed[132][2])))

        for _ in map(self.tree_10.delete, self.tree_10.get_children("")):
            pass
        self.tree_10.insert("", 0, values=("CPU_0V8_G3", str(data_analysed[133][0]), str(data_analysed[133][2])))
        self.tree_10.insert("", 1, values=("DDR_0V8_G3", str(data_analysed[134][0]), str(data_analysed[134][2])))
        self.tree_10.insert("", 2, values=("VDD_0V8_G4", str(data_analysed[135][0]), str(data_analysed[135][2])))
        self.tree_10.insert("", 3, values=("COREAO_0V8_G2", str(data_analysed[136][0]), str(data_analysed[136][2])))
        self.tree_10.insert("", 4, values=("CNN0_0V8_G3", str(data_analysed[137][0]), str(data_analysed[137][2])))
        self.tree_10.insert("", 5, values=("CNN1_0V8_G3", str(data_analysed[138][0]), str(data_analysed[138][2])))
        self.tree_10.insert("", 6, values=("VDD_1V8_G1", str(data_analysed[139][0]), str(data_analysed[139][2])))
        self.tree_10.insert("", 7, values=("EMMC_3V3_G1", str(data_analysed[140][0]), str(data_analysed[140][2])))
        self.tree_10.insert("", 8, values=("VDD_3V3_G1", str(data_analysed[141][0]), str(data_analysed[141][2])))
        self.tree_10.insert("", 9, values=("VCC_3V3_G4", str(data_analysed[142][0]), str(data_analysed[142][2])))
        self.tree_10.insert("", 10, values=("VCC_1V8_G4", str(data_analysed[143][0]), str(data_analysed[143][2])))
        self.tree_10.insert("", 11, values=("TEMP_IC", str(data_analysed[144][0]), str(data_analysed[144][2])))
        self.tree_10.insert("", 12, values=("TEMP_SW1", str(data_analysed[145][0]), str(data_analysed[145][2])))
        self.tree_10.insert("", 13, values=("TEMP_SW2", str(data_analysed[146][0]), str(data_analysed[146][2])))
        self.tree_10.insert("", 14, values=("TEMP_SW3",str(data_analysed[147][0]), str(data_analysed[147][2])))
        self.tree_10.insert("", 15, values=("TEMP_SW4", str(data_analysed[148][0]), str(data_analysed[148][2])))
        self.tree_10.insert("", 16, values=("TEMP_SW5", str(data_analysed[149][0]), str(data_analysed[149][2])))
        self.tree_10.insert("", 17, values=("TEMP_SW6", str(data_analysed[150][0]), str(data_analysed[150][2])))
        self.tree_10.insert("", 18, values=("TEMP_SW7", str(data_analysed[151][0]), str(data_analysed[151][2])))
        self.tree_10.insert("", 19, values=("TEMP_TEMP_LDO1_2", str(data_analysed[152][0]), str(data_analysed[152][2])))
        self.tree_10.insert("", 20, values=("TEMP_TEMP_LDO3_4", str(data_analysed[153][0]), str(data_analysed[153][2])))

        for _ in map(self.tree_11.delete, self.tree_11.get_children("")):
            pass
        self.tree_11.insert("", 0, values=("CPU_0V8_G3", str(data_analysed[154][0]), str(data_analysed[154][2])))
        self.tree_11.insert("", 1, values=("DDR_0V8_G3",str(data_analysed[155][0]), str(data_analysed[155][2])))
        self.tree_11.insert("", 2, values=("VDD_0V8_G4",str(data_analysed[156][0]), str(data_analysed[156][2])))
        self.tree_11.insert("", 3, values=("COREAO_0V8_G2", str(data_analysed[157][0]), str(data_analysed[157][2])))
        self.tree_11.insert("", 4, values=("CNN0_0V8_G3", str(data_analysed[158][0]), str(data_analysed[158][2])))
        self.tree_11.insert("", 5, values=("CNN1_0V8_G3", str(data_analysed[159][0]), str(data_analysed[159][2])))
        self.tree_11.insert("", 6, values=("VDD_1V8_G1", str(data_analysed[160][0]), str(data_analysed[160][2])))
        self.tree_11.insert("", 7, values=("EMMC_3V3_G1", str(data_analysed[161][0]), str(data_analysed[161][2])))
        self.tree_11.insert("", 8, values=("VDD_3V3_G1", str(data_analysed[162][0]), str(data_analysed[162][2])))
        self.tree_11.insert("", 9, values=("VCC_3V3_G4", str(data_analysed[163][0]), str(data_analysed[163][2])))
        self.tree_11.insert("", 10, values=("VCC_1V8_G4", str(data_analysed[164][0]), str(data_analysed[164][2])))
        self.tree_11.insert("", 11, values=("TEMP_IC", str(data_analysed[165][0]), str(data_analysed[165][2])))
        self.tree_11.insert("", 12, values=("TEMP_SW1", str(data_analysed[166][0]), str(data_analysed[166][2])))
        self.tree_11.insert("", 13, values=("TEMP_SW2", str(data_analysed[167][0]), str(data_analysed[167][2])))
        self.tree_11.insert("", 14, values=("TEMP_SW3", str(data_analysed[168][0]), str(data_analysed[168][2])))
        self.tree_11.insert("", 15, values=("TEMP_SW4", str(data_analysed[169][0]), str(data_analysed[169][2])))
        self.tree_11.insert("", 16, values=("TEMP_SW5", str(data_analysed[170][0]), str(data_analysed[170][2])))
        self.tree_11.insert("", 17, values=("TEMP_SW6", str(data_analysed[171][0]), str(data_analysed[171][2])))
        self.tree_11.insert("", 18, values=("TEMP_SW7",str(data_analysed[172][0]), str(data_analysed[172][2])))
        self.tree_11.insert("", 19, values=("TEMP_TEMP_LDO1_2", str(data_analysed[173][0]), str(data_analysed[173][2])))
        self.tree_11.insert("", 20, values=("TEMP_TEMP_LDO3_4", str(data_analysed[174][0]), str(data_analysed[174][2])))

        for _ in map(self.tree_12.delete, self.tree_12.get_children("")):
            pass
        self.tree_12.insert("", 0, values=("CPU_0V8_G3",str(data_analysed[175][0]), str(data_analysed[175][2])))
        self.tree_12.insert("", 1, values=("DDR_0V8_G3", str(data_analysed[176][0]), str(data_analysed[176][2])))
        self.tree_12.insert("", 2, values=("VDD_0V8_G4", str(data_analysed[177][0]), str(data_analysed[177][2])))
        self.tree_12.insert("", 3, values=("COREAO_0V8_G2", str(data_analysed[178][0]), str(data_analysed[178][2])))
        self.tree_12.insert("", 4, values=("CNN0_0V8_G3", str(data_analysed[179][0]), str(data_analysed[179][2])))
        self.tree_12.insert("", 5, values=("CNN1_0V8_G3", str(data_analysed[180][0]), str(data_analysed[180][2])))
        self.tree_12.insert("", 6, values=("VDD_1V8_G1", str(data_analysed[181][0]), str(data_analysed[181][2])))
        self.tree_12.insert("", 7, values=("EMMC_3V3_G1", str(data_analysed[182][0]), str(data_analysed[182][2])))
        self.tree_12.insert("", 8, values=("VDD_3V3_G1", str(data_analysed[183][0]), str(data_analysed[183][2])))
        self.tree_12.insert("", 9, values=("VCC_3V3_G4", str(data_analysed[184][0]), str(data_analysed[184][2])))
        self.tree_12.insert("", 10, values=("VCC_1V8_G4",str(data_analysed[185][0]), str(data_analysed[185][2])))
        self.tree_12.insert("", 11, values=("TEMP_IC", str(data_analysed[186][0]), str(data_analysed[186][2])))
        self.tree_12.insert("", 12, values=("TEMP_SW1", str(data_analysed[187][0]), str(data_analysed[187][2])))
        self.tree_12.insert("", 13, values=("TEMP_SW2", str(data_analysed[188][0]), str(data_analysed[188][2])))
        self.tree_12.insert("", 14, values=("TEMP_SW3", str(data_analysed[189][0]), str(data_analysed[189][2])))
        self.tree_12.insert("", 15, values=("TEMP_SW4", str(data_analysed[190][0]), str(data_analysed[190][2])))
        self.tree_12.insert("", 16, values=("TEMP_SW5", str(data_analysed[191][0]), str(data_analysed[191][2])))
        self.tree_12.insert("", 17, values=("TEMP_SW6", str(data_analysed[192][0]), str(data_analysed[192][2])))
        self.tree_12.insert("", 18, values=("TEMP_SW7", str(data_analysed[193][0]), str(data_analysed[193][2])))
        self.tree_12.insert("", 19, values=("TEMP_TEMP_LDO1_2",str(data_analysed[194][0]), str(data_analysed[194][2])))
        self.tree_12.insert("", 20, values=("TEMP_TEMP_LDO3_4", str(data_analysed[195][0]), str(data_analysed[195][2])))


        for _ in map(self.tree_13.delete, self.tree_13.get_children("")):
            pass
        self.tree_13.insert("", 0, values=("MCU_MAX20084", str(data_analysed[196][0]), str(data_analysed[196][2])))
        self.tree_13.insert("", 1, values=("MCU_PMIC8100_Index0", str(data_analysed[197][0]), str(data_analysed[197][2])))
        self.tree_13.insert("", 2, values=("MCU_PMIC8100_Index1", str(data_analysed[198][0]), str(data_analysed[198][2])))
        self.tree_13.insert("", 3, values=("MCU_PMIC8100_Index2", str(data_analysed[199][0]), str(data_analysed[199][2])))
        self.tree_13.insert("", 4, values=("MCU_PMIC5024_Index0", str(data_analysed[200][0]), str(data_analysed[200][2])))
        self.tree_13.insert("", 5, values=("MCU_PMIC5024_Index1", str(data_analysed[201][0]), str(data_analysed[201][2])))
        self.tree_13.insert("", 6, values=("MCU_PMIC5024_Index2", str(data_analysed[202][0]), str(data_analysed[202][2])))
        self.tree_13.insert("", 7, values=("J3A_MAX9296", str(data_analysed[203][0]), str(data_analysed[203][2])))
        self.tree_13.insert("", 8, values=("J3A_MAX96717", str(data_analysed[204][0]), str(data_analysed[204][2])))
        self.tree_13.insert("", 9, values=("J3A_MAX20089", str(data_analysed[205][0]), str(data_analysed[205][2])))
        self.tree_13.insert("", 10, values=("J3B_MAX9296", str(data_analysed[206][0]), str(data_analysed[206][2])))
        self.tree_13.insert("", 11, values=("J3B_MAX96712", str(data_analysed[207][0]), str(data_analysed[207][2])))
        self.tree_13.insert("", 12, values=("J3B_MAX20089", str(data_analysed[208][0]), str(data_analysed[208][2])))
        self.tree_13.insert("", 13, values=("J3B_MAX20087", str(data_analysed[209][0]), str(data_analysed[209][2])))
        self.tree_13.insert("", 14, values=("J3C_MAX96712", str(data_analysed[210][0]), str(data_analysed[210][2])))
        self.tree_13.insert("", 15, values=("J3C_MAX96717", str(data_analysed[211][0]), str(data_analysed[211][2])))
        self.tree_13.insert("", 16, values=("J3C_MAX20087",str(data_analysed[212][0]), str(data_analysed[212][2])))

        self.tree_13.insert("", 17, values=("J3A_Front_Camera_max9295_ID", str(data_analysed[213][0]), str(data_analysed[213][2])))
        self.tree_13.insert("", 18, values=("J3B_Rear_Camera_max9295_ID", str(data_analysed[214][0]), str(data_analysed[214][2])))
        self.tree_13.insert("", 19, values=("J3B_Side_Camera_0_max9295_ID", str(data_analysed[215][0]), str(data_analysed[215][2])))
        self.tree_13.insert("", 20, values=("J3B_Side_Camera_1_max9295_ID", str(data_analysed[216][0]), str(data_analysed[216][2])))
        self.tree_13.insert("", 21, values=("J3B_Side_Camera_2_max9295_ID", str(data_analysed[217][0]), str(data_analysed[217][2])))
        self.tree_13.insert("", 22, values=("J3B_Side_Camera_3_max9295_ID", str(data_analysed[218][0]), str(data_analysed[218][2])))
        self.tree_13.insert("", 23, values=("J3C_Surround_Camera_0_max9295_ID", str(data_analysed[219][0]), str(data_analysed[219][2])))
        self.tree_13.insert("", 24, values=("J3C_Surround_Camera_1_max9295_ID", str(data_analysed[220][0]), str(data_analysed[220][2])))
        self.tree_13.insert("", 25, values=("J3C_Surround_Camera_2_max9295_ID", str(data_analysed[221][0]), str(data_analysed[221][2])))
        self.tree_13.insert("", 26, values=("J3C_Surround_Camera_3_max9295_ID", str(data_analysed[222][0]), str(data_analysed[222][2])))


        for _ in map(self.tree_14.delete, self.tree_14.get_children("")):
            pass
        self.tree_14.insert("", 0, values=("USS1_PWM_PERIOD", str(data_analysed[233][0]), str(data_analysed[233][2])))
        self.tree_14.insert("", 1, values=("USS1_PWM_DUTY", str(data_analysed[234][0]), str(data_analysed[234][2])))
        self.tree_14.insert("", 2, values=("USS2_PWM_PERIOD", str(data_analysed[235][0]), str(data_analysed[235][2])))
        self.tree_14.insert("", 3, values=("USS2_PWM_DUTY", str(data_analysed[236][0]), str(data_analysed[236][2])))
        self.tree_14.insert("", 4, values=("USS3_PWM_PERIOD", str(data_analysed[237][0]), str(data_analysed[237][2])))
        self.tree_14.insert("", 5, values=("USS3_PWM_DUTY", str(data_analysed[238][0]), str(data_analysed[238][2])))
        self.tree_14.insert("", 6, values=("USS4_PWM_PERIOD", str(data_analysed[239][0]), str(data_analysed[239][2])))
        self.tree_14.insert("", 7, values=("USS4_PWM_DUTY", str(data_analysed[240][0]), str(data_analysed[240][2])))
        self.tree_14.insert("", 8, values=("USS5_PWM_PERIOD", str(data_analysed[241][0]), str(data_analysed[241][2])))
        self.tree_14.insert("", 9, values=("USS5_PWM_DUTY", str(data_analysed[242][0]), str(data_analysed[242][2])))
        self.tree_14.insert("", 10, values=("USS6_PWM_PERIOD", str(data_analysed[243][0]), str(data_analysed[243][2])))
        self.tree_14.insert("", 11, values=("USS6_PWM_DUTY", str(data_analysed[244][0]), str(data_analysed[244][2])))
        self.tree_14.insert("", 12, values=("USS7_PWM_PERIOD", str(data_analysed[245][0]), str(data_analysed[245][2])))
        self.tree_14.insert("", 13, values=("USS7_PWM_DUTY", str(data_analysed[246][0]), str(data_analysed[246][2])))
        self.tree_14.insert("", 14, values=("USS8_PWM_PERIOD", str(data_analysed[247][0]), str(data_analysed[247][2])))
        self.tree_14.insert("", 15, values=("USS8_PWM_DUTY", str(data_analysed[248][0]), str(data_analysed[248][2])))
        self.tree_14.insert("", 16, values=("USS9_PWM_PERIOD", str(data_analysed[249][0]), str(data_analysed[249][2])))
        self.tree_14.insert("", 17, values=("USS9_PWM_DUTY", str(data_analysed[250][0]), str(data_analysed[250][2])))
        self.tree_14.insert("", 18, values=("USS10_PWM_PERIOD", str(data_analysed[251][0]), str(data_analysed[251][2])))
        self.tree_14.insert("", 19, values=("USS10_PWM_DUTY", str(data_analysed[252][0]), str(data_analysed[252][2])))
        self.tree_14.insert("", 20, values=("USS11_PWM_PERIOD", str(data_analysed[253][0]), str(data_analysed[253][2])))
        self.tree_14.insert("", 21, values=("USS11_PWM_DUTY", str(data_analysed[254][0]), str(data_analysed[254][2])))
        self.tree_14.insert("", 22, values=("USS12_PWM_PERIOD", str(data_analysed[255][0]), str(data_analysed[255][2])))
        self.tree_14.insert("", 23, values=("USS12_PWM_DUTY", str(data_analysed[256][0]), str(data_analysed[256][2])))


        for _ in map(self.tree_15.delete, self.tree_15.get_children("")):
            pass
        self.tree_15.insert("", 0, values=("J3A_Front_Camera", str(data_analysed[257][0]), str(data_analysed[257][2])))
        self.tree_15.insert("", 1, values=("J3B_Rear_Camera", str(data_analysed[258][0]), str(data_analysed[258][2])))
        self.tree_15.insert("", 2, values=("J3B_Side_Camera_0", str(data_analysed[259][0]), str(data_analysed[259][2])))
        self.tree_15.insert("", 3, values=("J3B_Side_Camera_1", str(data_analysed[260][0]), str(data_analysed[260][2])))
        self.tree_15.insert("", 4, values=("J3B_Side_Camera_2", str(data_analysed[261][0]), str(data_analysed[261][2])))
        self.tree_15.insert("", 5, values=("J3B_Side_Camera_3", str(data_analysed[262][0]), str(data_analysed[262][2])))
        self.tree_15.insert("", 6, values=("J3C_Surround_Camera_0", str(data_analysed[263][0]), str(data_analysed[263][2])))
        self.tree_15.insert("", 7, values=("J3C_Surround_Camera_1", str(data_analysed[264][0]), str(data_analysed[264][2])))
        self.tree_15.insert("", 8, values=("J3C_Surround_Camera_2", str(data_analysed[265][0]), str(data_analysed[265][2])))
        self.tree_15.insert("", 9, values=("J3C_Surround_Camera_3", str(data_analysed[266][0]), str(data_analysed[266][2])))



init_window = tkinter.Tk()
init_window.title("DV测试上位机")
w = init_window.winfo_screenwidth()
h = init_window.winfo_screenheight()
init_window.geometry("%dx%d" % (w, h))

tabControl = ttk.Notebook(init_window)

tab1 = ttk.Frame(tabControl)
tabControl.add(tab1, text="ECU 1")
tabControl.pack(expand=1, fill="both")

tab2 = ttk.Frame(tabControl)
tabControl.add(tab2, text="ECU 2")

tab3 = ttk.Frame(tabControl)
tabControl.add(tab3, text="ECU 3")

tab4 = ttk.Frame(tabControl)
tabControl.add(tab4, text="ECU 4")

tab5 = ttk.Frame(tabControl)
tabControl.add(tab5, text="ECU 5")

tab6 = ttk.Frame(tabControl)
tabControl.add(tab6, text="ECU 6")

tab1_show = MY_GUI(tab1)
tab1_show.init_tab()

log_row = 2

def UI_Refresh():
    global data_analysed_1, data_analysed_2, data_analysed_3, data_analysed_4, data_analysed_5,\
        data_analysed_6, log_row, sheet, val, first_row,column_number, file_name
    while True:
        start_time = time.time()
        # '''Read information from incubator'''
        # print('Read information from incubator')
        # cycle_request = '0000000000060003003A0001'
        # tcp_sk.send(a2b_hex(cycle_request))
        # cycle_number = rev_msg(0x03)
        # print(cycle_number, 'cycle_number')
        # cycle_number_DEC = cycle_number[0] * 256 + cycle_number[1]
        # print(cycle_number_DEC, 'cycle_number_DEC')
        #
        # step_request = '0000000000060003003C0001'
        # tcp_sk.send(a2b_hex(step_request))
        # step_number = rev_msg(0x03)
        # print(step_number, 'step_number')
        # step_number_DEC = step_number[0] * 256 + step_number[1]
        # print(step_number_DEC, 'step_number_DEC')
        #
        # hour_request = '0000000000060003003F0002'
        # tcp_sk.send(a2b_hex(hour_request))
        # time_number = rev_msg(0x03)
        # print(time_number, 'time_number')
        # hour_number = time_number[0] * 256 + time_number[1]
        # minute_number = time_number[2] * 256 + time_number[3]
        # print(hour_number, 'hour_number')
        # print(minute_number, 'minute_number')
        # time_display = '%s:%s:00' % (hour_number, minute_number)
        #
        # curr_temp_request = '000000000006000300040001'
        # tcp_sk.send(a2b_hex(curr_temp_request))
        # curr_temp = rev_msg(0x03)
        # print(curr_temp, 'curr_temp')
        # if curr_temp[0] & 128:
        #     # curr_temp[0] = 255 - curr_temp[0]
        #     curr_temp_display = (256 * 256 - (curr_temp[0] * 256 + curr_temp[1])) / 10
        #     curr_temp_display = '-' + str(curr_temp_display) + '℃'
        # else:
        #     curr_temp_display = (curr_temp[0] * 256 + curr_temp[1]) / 10
        #     curr_temp_display = str(curr_temp_display) + '℃'
        # print(curr_temp_display, 'curr_temp_display')
        #
        # '''Relay control'''
        # if step_number_DEC == 2 and hour_number == 1 and 25 <= minute_number <= 30:
        #     ser.write(channel_Open)
        #     mode = '3.2'
        # elif step_number_DEC == 4 or step_number_DEC == 5:
        #     ser.write(channel_Open)
        #     mode = '3.2'
        # else:
        #     ser.write(channel_Close)
        #     mode = '2.1'

        '''Obtain data from ECUs'''
        if ping_res_1:
            udp_sock_1.sendto(a2b_hex('55AA55'), (ip_ecu, port_ecu))
        else:
            pass
        # if ping_res_2:
        #     udp_sock_2.sendto(a2b_hex('55AA55'), (ip_ecu, port_ecu))
        # else:
        #     pass
        # if ping_res_3:
        #     udp_sock_3.sendto(a2b_hex('55AA55'), (ip_ecu, port_ecu))
        # else:
        #     pass
        # if ping_res_4:
        #     udp_sock_4.sendto(a2b_hex('55AA55'), (ip_ecu, port_ecu))
        # else:
        #     pass
        # if ping_res_5:
        #     udp_sock_5.sendto(a2b_hex('55AA55'), (ip_ecu, port_ecu))
        # else:
        #     pass
        # if ping_res_6:
        #     udp_sock_6.sendto(a2b_hex('55AA55'), (ip_ecu, port_ecu))
        # else:
        #     pass
        time.sleep(0.2)

        # val1 = openpyxl.load_workbook(file_name)
        # sheet1 = val1['validation']
        sheet.cell(log_row,1,time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        i = 0
        while i < len(data_analysed_1):
            if data_analysed_1 != []:
                if data_analysed_1[i][1] == 0:
                    sheet.cell(log_row, i+2, str(data_analysed_1[i][0]))
                else:
                    sheet.cell(log_row, i+2, str(data_analysed_1[i][0])).fill=PatternFill('solid', fgColor = 'FF0000')
            else:
                pass
            i += 1
        val.save(file_name)
        log_row += 1
        if log_row <= 200:
            pass
        else:
            if os.path.exists('c:\\validation_log') == False:
                os.mkdir('c:\\validation_log')
            val = openpyxl.Workbook()
            sheet = val.active
            sheet.title = 'validation'
            sheet.column_dimensions['A'].width = 20
            for i in column_number:
                sheet.column_dimensions[i].width = 20
            sheet.append(first_row)
            file_name = 'c:\\validation_log\\Validation_TestLog_%s.xlsx' % (
                time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()))
            val.save(file_name)
            log_row = 2

        if data_analysed_1 != []:
            tab1_show.tab_Refresh(data_analysed_1)
        else:
            pass

        data_analysed_1 = []
        data_analysed_2 = []
        data_analysed_3 = []
        data_analysed_4 = []
        data_analysed_5 = []
        data_analysed_6 = []

        end_time = time.time()
        t_sleep = 4 - (end_time - start_time)
        if t_sleep > 0:
            time.sleep(t_sleep)
        else:
            pass

thread = Thread(target = UI_Refresh)
thread.daemon = True
thread.start()

init_window.mainloop()